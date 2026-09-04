# -*- coding: utf-8 -*-

import base64
import io
import logging
import re
import unicodedata

from markupsafe import Markup, escape
from psycopg2 import IntegrityError

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools import html2plaintext

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
except ImportError:
    Workbook = None
    PageMargins = None

_logger = logging.getLogger(__name__)

# Nhận diện bước công việc theo tên (đã bỏ dấu / bỏ ký tự đặc biệt) để không phụ
# thuộc vào khoảng trắng hay lỗi chính tả nhỏ khi đặt tên giai đoạn.
TEM_TAG_IMPORT_TASK_MARKERS = ('dobbthaytemtag',)
TEM_PHOTO_TASK_MARKERS = ('chuptemguilengroup', 'chupteamguilengroup')
# Bước 14 "Kiểm tra hình ảnh tem tag".
TEM_CHECK_TASK_MARKERS = ('kiemtrahinhanhtemtag',)
# Bước 12 "Thay tem Tag" phải khớp tuyệt đối: nhiều bước khác (bước 4, 6, 15)
# cũng chứa chuỗi "thaytemtag" trong tên nên không dùng so khớp chứa được.
TEM_REPLACE_TASK_KEYS = ('thaytemtag',)
# Bước 6 "Lập BB thay tem, bàn giao cho KT kho Kiểm tra BB thay tem tag":
# nhận diện riêng (không trùng với bước 4 / 12 / 15) để hiện bảng
# "Chi tiết tem/tag" theo Cửa hàng quản lí và mục "Xuất biên bản thay tem".
TEM_BB_REPLACE_TASK_MARKERS = ('lapbbthaytem', 'lapbienbanthaytem')
# Bước 9 "In tem, Tag".
TEM_PRINT_TASK_MARKERS = ('intemtag',)
# Bước 6 "Lập BB thay tem, bàn giao cho KT kho".
TEM_BB_REPLACE_TASK_MARKERS = ('lapbbthaytem',)
# Bước 10 "Bàn giao Tem Tag cho CH  Thu hồi tem tag cũ".
TEM_HANDOVER_TASK_MARKERS = ('bangiaotemtag', 'thuhoitentagcu')
# Bước 11 "Nhận tem tag mới".
TEM_RECEIVE_TASK_MARKERS = ('nhantemtagmoi', 'nhantemtag')
# Bước 7 "Thiết kế mẫu tem/tag, Bảng nhận diện".
TEM_DESIGN_TASK_MARKERS = ('thietkemautemtag', 'bangnhandien', 'thietkemau')
# Ưu tiên nhận diện theo giai đoạn mặc định (bền hơn khi bước bị đổi tên).
TEM_TAG_IMPORT_STAGE_XMLID = 'ctkm_core.ctkm_stage_4'
TEM_BB_REPLACE_STAGE_XMLID = 'ctkm_core.ctkm_stage_6'
TEM_PRINT_STAGE_XMLID = 'ctkm_core.ctkm_stage_9'
TEM_HANDOVER_STAGE_XMLID = 'ctkm_core.ctkm_stage_10'
TEM_RECEIVE_STAGE_XMLID = 'ctkm_core.ctkm_stage_11'
TEM_DESIGN_STAGE_XMLID = 'ctkm_core.ctkm_stage_7'
TEM_REPLACE_STAGE_XMLID = 'ctkm_core.ctkm_stage_12'
TEM_PHOTO_STAGE_XMLID = 'ctkm_core.ctkm_stage_13'
TEM_CHECK_STAGE_XMLID = 'ctkm_core.ctkm_stage_14'
TEM_POSTCHECK_STAGE_XMLID = 'ctkm_core.ctkm_stage_16'
# Các trường kho Tem/Tag được gom vào bảng "Chi tiết tem/tag".
TEM_TAG_LINE_KEY_FIELDS = ('material_code', 'store')
# Chức danh trên hồ sơ nhân viên (hr_job_title_vn) dùng để gán bước 11–14 / hậu kiểm.
STORE_MANAGER_JOB_TITLE = 'cửa hàng trưởng'
SUPERVISOR_JOB_TITLE = 'giám sát'
# Gộp Cửa hàng trưởng vào mã ngắn trên file (AETL, AEBD) chỉ khi hồ sơ
# có CẢ {store}_DDL và {store}_DNA, và không có mã trần {store}.
# Nếu file đã ghi sẵn hậu tố (AETA_DNA) thì khớp chính xác, không gộp.
STORE_DEPT_GROUP_SUFFIXES = ('_DDL', '_DNA')
# Bước 16 "Hậu kiểm CTKM Giám sát đi kiểm tra thay tem".
TEM_POSTCHECK_TASK_MARKERS = ('haukiem',)
TEM_PRICE_STAGE_XMLID = 'ctkm_core.ctkm_stage_15'
# Bước 15 "Kế toán áp giá CTKM lên PM Link Q ... Lập báo cáo cửa hàng đã áp giá".
# Không dùng "apgia" chung vì bước 5 / 8 cũng chứa "áp giá".
TEM_PRICE_TASK_MARKERS = ('ketoanapgia', 'lapbaocaocuahang',)


def normalize_step_key(value):
    """Chuẩn hóa tên bước: chữ thường, bỏ dấu, chỉ giữ chữ và số."""
    text = unicodedata.normalize('NFD', (value or '').lower())
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    return re.sub(r'[^a-z0-9]+', '', text.replace('đ', 'd'))


def _ctkm_normalize_store_key(value):
    """Chuẩn hóa mã Store / mã bộ phận để so khớp (hoa, bỏ khoảng trắng thừa)."""
    if isinstance(value, dict):
        value = next(iter(value.values()), '') if value else ''
    if not isinstance(value, str):
        value = str(value) if value else ''
    value = ' '.join(value.strip().split())
    return value.upper() if value else False


class CtkmTask(models.Model):
    _name = 'ctkm.task'
    _description = 'Công việc CTKM'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'process_date desc, id desc'

    process_date = fields.Date(
        string='Ngày xử lý',
        default=fields.Date.context_today,
        tracking=True,
    )
    name = fields.Text(
        string='Nội dung CV',
        required=True,
        tracking=True,
    )
    done_date = fields.Date(string='Ngày hoàn thành', tracking=True)
    total_days = fields.Integer(
        string='Tổng số ngày',
        compute='_compute_total_days',
        store=True,
        help='Ngày hoàn thành trừ ngày bắt đầu, tính khi bấm Hoàn thành.',
    )
    state = fields.Selection(
        selection=[
            ('todo', 'Chưa xử lý'),
            ('progress', 'Đang xử lý'),
            ('waiting_confirm', 'Chờ xác nhận'),
            ('done', 'Hoàn thành'),
        ],
        string='Trạng thái',
        default='todo',
        required=True,
        tracking=True,
    )
    duration = fields.Char(
        string='Thời gian xử lý',
        help='VD: 2 giờ',
    )
    handover_date = fields.Date(string='Ngày bàn giao', tracking=True)
    handover_employee_id = fields.Many2one(
        'hr.employee',
        string='Người nhận bàn giao',
        tracking=True,
        domain="[('active', '=', True)]",
    )
    manager_confirmed = fields.Boolean(
        string='Xác nhận quản lý',
        tracking=True,
    )
    can_confirm_as_manager = fields.Boolean(
        string='Được phép xác nhận quản lý',
        compute='_compute_can_confirm_as_manager',
    )
    checklist_need_manager_confirm = fields.Boolean(
        string='Checklist cần quản lý xác nhận',
        compute='_compute_checklist_need_manager_confirm',
    )
    is_task_assignee = fields.Boolean(
        string='Là người nhận việc',
        compute='_compute_is_task_assignee',
    )
    is_current_stage_task = fields.Boolean(
        string='Bước hiện tại',
        compute='_compute_is_current_stage_task',
        search='_search_is_current_stage_task',
        help='True khi công việc thuộc bước (giai đoạn) hiện tại của CTKM, '
             'tức là đến lượt xử lý lúc này (không phải bước tương lai).',
    )
    is_tem_tag_import_task = fields.Boolean(
        string='Bước import Tem/Tag',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Đổ BB thay tem/tag (file tổng)".',
    )
    is_tem_photo_task = fields.Boolean(
        string='Bước chụp ảnh tem/tag',
        compute='_compute_task_step_flags',
        store=True,
    )
    is_tem_check_task = fields.Boolean(
        string='Bước kiểm tra hình ảnh tem/tag',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Kiểm tra hình ảnh tem tag".',
    )
    is_tem_replace_task = fields.Boolean(
        string='Bước thay tem/tag',
        compute='_compute_task_step_flags',
        store=True,
    )
    is_tem_bb_replace_task = fields.Boolean(
        string='Bước lập BB thay tem',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc bước "Lập BB thay tem" (xuất biên bản thay tem cho KT kho; '
             'bảng "Chi tiết tem/tag" chỉ hiện cửa hàng trong "Cửa hàng quản lí").',
    )
    is_tem_print_task = fields.Boolean(
        string='Bước in tem/tag',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "In tem, Tag".',
    )
    is_tem_handover_task = fields.Boolean(
        string='Bước bàn giao / thu hồi tem/tag',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Bàn giao Tem Tag cho CH Thu hồi tem tag cũ".',
    )
    is_tem_receive_task = fields.Boolean(
        string='Bước nhận tem tag mới',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Nhận tem tag mới".',
    )
    is_tem_design_task = fields.Boolean(
        string='Bước thiết kế mẫu tem/tag',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Thiết kế mẫu tem/tag, Bảng nhận diện".',
    )
    is_tem_postcheck_task = fields.Boolean(
        string='Bước hậu kiểm CTKM',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Hậu kiểm CTKM Giám sát đi kiểm tra thay tem".',
    )
    is_tem_price_task = fields.Boolean(
        string='Bước kế toán áp giá',
        compute='_compute_task_step_flags',
        store=True,
        help='Công việc thuộc bước "Kế toán áp giá CTKM lên PM Link Q".',
    )
    show_work_processing_tab = fields.Boolean(
        string='Hiện tab Xử lý công việc',
        compute='_compute_show_work_processing_tab',
        help='Tab "Xử lý công việc" cho bước 9–11, bước 15 (Kế toán áp giá) và Hậu kiểm.',
    )
    handover_quantity = fields.Float(string='Số lượng bàn giao')
    recovery_quantity = fields.Float(string='Số lượng thu hồi')
    recover_ids = fields.One2many(
        'ctkm.task.tem.tag.recover.line',
        'task_id',
        string='Thu hồi tem',
        help='Các Tem/Tag thu hồi ở bước 10; bị xóa khỏi Kho khi bấm Hoàn thành.',
    )
    print_store_search = fields.Char(
        string='Tìm cửa hàng',
        help='Lọc danh sách cửa hàng cần in tem/tag.',
    )
    print_store_ids = fields.One2many(
        'ctkm.task.tem.print.line',
        'task_id',
        string='Cửa hàng in tem/tag',
        help='Danh sách cửa hàng cần in tem/tag ở bước 9, gom từ file tổng.',
    )
    print_store_pending_ids = fields.One2many(
        'ctkm.task.tem.print.line',
        'task_id',
        string='Thông tin cửa hàng và số lượng tem/tag cần in',
        domain=[('done', '=', False)],
        copy=False,
        help='Cửa hàng chưa tick Kết quả; tick để chuyển xuống bảng đã hoàn thành.',
    )
    print_store_done_ids = fields.One2many(
        'ctkm.task.tem.print.line',
        'task_id',
        string='Cửa hàng đã hoàn thành in tem/tag',
        domain=[('done', '=', True)],
        copy=False,
        help='Cửa hàng đã tick Kết quả. Bỏ tick để đưa lại bảng cần in.',
    )
    handover_store_search = fields.Char(string='Tìm cửa hàng bàn giao')
    collect_store_search = fields.Char(string='Tìm cửa hàng thu')
    handover_store_ids = fields.One2many(
        'ctkm.task.tem.step10.line',
        'task_id',
        string='Bàn giao tem/tag',
        domain=[('line_type', '=', 'handover')],
        context={'default_line_type': 'handover'},
        help='Tên cửa hàng và SL tem/tag lấy từ bước 9 In tem, Tag.',
    )
    collect_store_ids = fields.One2many(
        'ctkm.task.tem.step10.line',
        'task_id',
        string='Thu tem/tag',
        domain=[('line_type', '=', 'collect')],
        context={'default_line_type': 'collect'},
        help='Mã cửa hàng lấy từ bước 9; SL tem/tag điền tay.',
    )
    postcheck_store_search = fields.Char(string='Tìm cửa hàng hậu kiểm')
    postcheck_store_ids = fields.One2many(
        'ctkm.task.tem.postcheck.line',
        'task_id',
        string='Hậu kiểm thay tem',
        help='Cửa hàng và SL tem/tag lấy từ bước 9 In tem, Tag.',
    )
    price_store_search = fields.Char(string='Tìm cửa hàng kế toán áp giá')
    price_store_ids = fields.One2many(
        'ctkm.task.tem.price.line',
        'task_id',
        string='Báo cáo thay tem/tag và áp giá',
        help='Cửa hàng và SL tem/tag lấy từ bước 9 In tem, Tag.',
    )
    time_line_ids = fields.One2many(
        'ctkm.task.time.line',
        'task_id',
        string='Thời gian',
        help='Nội dung, ngày bắt đầu, ngày hoàn thành và tổng số ngày của công việc.',
    )
    verifier_ids = fields.Many2many(
        'hr.employee',
        'ctkm_task_verifier_rel',
        'task_id',
        'employee_id',
        string='Người kiểm soát',
        help='Nhân viên xác nhận bước này. Khi đặt, bước dùng những người này '
             'kiểm soát thay vì quản lý theo organization chart.',
    )
    store_verifier_ids = fields.One2many(
        'ctkm.task.store.verifier', 'task_id',
        string='Xác nhận theo cửa hàng', copy=False,
        help='Quan hệ 1 Phụ trách (Cửa hàng trưởng) – 1 Người kiểm soát '
             '(Quản lý cửa hàng) theo từng cửa hàng. Chỉ Quản lý cửa hàng '
             'ĐÚNG cửa hàng mới xác nhận được phần của Cửa hàng trưởng cửa hàng đó.',
    )
    is_verifier = fields.Boolean(
        string='Là người kiểm soát',
        compute='_compute_is_verifier',
        help='True khi người dùng hiện tại là Người kiểm soát của công việc.',
    )
    tem_tag_replace_ids = fields.One2many(
        'ctkm.task.tem.tag.replace.line',
        'task_id',
        string='Chi tiết tem/tag',
        help='Tem/Tag của CTKM này: bước 4 / 10 xem toàn bộ cửa hàng, '
             'bước 11 / 12 chỉ cửa hàng của người đang xem.',
    )
    tem_design_ids = fields.One2many(
        'ctkm.task.tem.design.line',
        'task_id',
        string='Thiết kế mẫu tem/tag',
        copy=False,
        help='Mỗi mã vật tư từ "Chi tiết tem/tag" của chương trình kèm file '
             'mẫu thiết kế (PDF / ảnh).',
    )
    support_employee_ids = fields.Many2many(
        'hr.employee',
        'ctkm_task_support_employee_rel',
        'task_id',
        'employee_id',
        string='Người hỗ trợ',
        domain="[('active', '=', True)]",
    )
    user_ids = fields.Many2many(
        'res.users',
        'ctkm_task_user_rel',
        'task_id',
        'user_id',
        string='Người nhận việc',
        default=lambda self: self.env.user,
        index=True,
        tracking=True,
    )
    completion_ids = fields.One2many(
        'ctkm.task.user.completion',
        'task_id',
        string='Hoàn thành từng người',
    )
    is_completed_by_me = fields.Boolean(
        string='Tôi đã hoàn thành',
        compute='_compute_is_completed_by_me',
    )
    all_assigned_completed = fields.Boolean(
        string='Tất cả đã hoàn thành',
        compute='_compute_all_assigned_completed',
    )
    pending_user_ids = fields.Many2many(
        'res.users',
        'ctkm_task_pending_user_rel',
        'task_id',
        'user_id',
        string='Chưa hoàn thành',
        compute='_compute_pending_user_ids',
    )
    program_id = fields.Many2one(
        'ctkm.program',
        string='Chương trình KM',
        ondelete='cascade',
        index=True,
        tracking=True,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Công ty',
        default=lambda self: self.env.company,
    )

    # --- Thông tin CTKM (readonly, lấy từ chương trình) ---
    program_name = fields.Char(
        related='program_id.name', string='Tên chương trình', readonly=True,
    )
    program_kanban_state = fields.Selection(
        related='program_id.kanban_state', string='Trạng thái giai đoạn', readonly=True,
    )
    program_stage_id = fields.Many2one(
        related='program_id.stage_id', string='Giai đoạn', readonly=True,
    )
    program_stage_progress_json = fields.Json(
        related='program_id.stage_progress_json',
        string='Tiến độ từng giai đoạn',
        readonly=True,
    )
    program_checklist_current_stage_id = fields.Many2one(
        related='program_id.checklist_current_stage_id',
        string='Giai đoạn đang làm',
        readonly=True,
    )
    program_date_begin = fields.Datetime(
        related='program_id.date_begin', string='Ngày bắt đầu', readonly=True,
    )
    program_date_end = fields.Datetime(
        related='program_id.date_end', string='Ngày kết thúc', readonly=True,
    )
    program_notify_code = fields.Char(
        related='program_id.notify_code', string='Mã số thông báo', readonly=True,
    )
    program_hour_quota = fields.Char(
        related='program_id.hour_quota', string='Định biên giờ', readonly=True,
    )
    program_organizer_id = fields.Many2one(
        related='program_id.organizer_id', string='Đơn vị tổ chức', readonly=True,
    )
    program_user_id = fields.Many2one(
        related='program_id.user_id', string='Người phụ trách', readonly=True,
    )
    program_address_id = fields.Many2one(
        related='program_id.address_id', string='Địa điểm', readonly=True,
    )
    program_event_url = fields.Char(
        related='program_id.event_url', string='URL sự kiện', readonly=True,
    )
    program_seats_limited = fields.Boolean(
        related='program_id.seats_limited', string='Giới hạn đăng ký', readonly=True,
    )
    program_seats_max = fields.Integer(
        related='program_id.seats_max', string='Số lượng tối đa', readonly=True,
    )
    program_company_id = fields.Many2one(
        related='program_id.company_id', string='Công ty CTKM', readonly=True,
    )
    program_badge_format = fields.Selection(
        related='program_id.badge_format', string='Kích thước nhãn', readonly=True,
    )
    # Binary qua sudo (không đi ACL attachment của program).
    program_badge_image = fields.Image(
        string='Ảnh nhãn',
        compute='_compute_program_badge_image',
        readonly=True,
    )
    # Bản copy gắn task — nhân viên đọc được mà không cần quyền program attachment.
    program_notify_document_ids = fields.Many2many(
        'ir.attachment',
        'ctkm_task_program_document_rel',
        'task_id',
        'attachment_id',
        string='Tài liệu gửi kèm thông báo',
        readonly=True,
        copy=False,
    )
    program_ticket_instructions = fields.Html(
        related='program_id.ticket_instructions',
        string='Hướng dẫn vé',
        readonly=True,
    )
    program_note = fields.Html(
        related='program_id.note', string='Ghi chú chương trình', readonly=True,
    )

    # Ghi chú / tài liệu của người nhận việc (được sửa)
    work_note = fields.Html(string='Ghi chú', sanitize_attributes=False)
    work_document_ids = fields.Many2many(
        'ir.attachment',
        'ctkm_task_work_document_rel',
        'task_id',
        'attachment_id',
        string='Tài liệu',
    )

    detail_tem_tag_file = fields.Binary(string='File Excel Tem/Tag')
    detail_tem_tag_filename = fields.Char(string='Tên file Excel')
    detail_tem_photo_ids = fields.Many2many(
        'ir.attachment',
        'ctkm_task_detail_tem_photo_rel',
        'task_id',
        'attachment_id',
        string='Ảnh tem/tag',
    )
    tem_photo_check_ids = fields.One2many(
        'ctkm.task.tem.photo.line',
        'task_id',
        string='Kiểm tra hình ảnh tem/tag',
        help='Bảng (Cửa hàng, Mã vật tư) lấy từ kho Tem/Tag file tổng. '
             'Bước 13 tick "Đã chụp"; bước 14 tick "Xác nhận thay".',
    )
    current_user_visible_store_keys = fields.Json(
        string='Store keys người đang xem',
        compute='_compute_current_user_visible_store_keys',
        help='Dùng để lọc bảng bước 13–14 theo mã bộ phận của người đang xem.',
    )

    # Bước phạm vi thông báo (Gửi tin tuần tự theo STT dòng)
    notify_line_id = fields.Many2one(
        'ctkm.program.notify.line',
        string='Bước phạm vi',
        ondelete='set null',
        index=True,
        copy=False,
    )
    checklist_line_id = fields.Many2one(
        'ctkm.program.checklist.line',
        string='Bước checklist',
        ondelete='set null',
        index=True,
        copy=False,
    )
    notify_step_label = fields.Char(
        related='notify_line_id.step_label',
        string='Bước xử lý',
        readonly=True,
    )
    checklist_step_name = fields.Char(
        string='Tên bước checklist',
        related='checklist_line_id.name',
        readonly=True,
    )
    forwarded = fields.Boolean(
        string='Đã chuyển tiếp bước',
        default=False,
        copy=False,
        tracking=True,
    )
    # Ghi chú / file nhận từ chương trình + bước trước (readonly)
    handover_note = fields.Html(
        string='Ghi chú nhận từ bước trước',
        sanitize_attributes=False,
        readonly=True,
        copy=False,
    )
    handover_document_ids = fields.Many2many(
        'ir.attachment',
        'ctkm_task_handover_document_rel',
        'task_id',
        'attachment_id',
        string='Tài liệu nhận từ bước trước',
        readonly=True,
        copy=False,
    )

    _user_program_line_uniq = models.Constraint(
        'UNIQUE(program_id, notify_line_id, checklist_line_id)',
        'Đã có công việc cho bước phạm vi / bước checklist này rồi.',
    )

    @api.depends_context('uid')
    def _compute_can_confirm_as_manager(self):
        for task in self:
            task.can_confirm_as_manager = task._user_can_confirm_as_manager(self.env.user)

    @api.depends('program_id', 'program_id.badge_image')
    def _compute_program_badge_image(self):
        for task in self:
            task.program_badge_image = task.program_id.sudo().badge_image or False

    @api.depends('process_date', 'done_date')
    def _compute_total_days(self):
        for task in self:
            if task.process_date and task.done_date:
                task.total_days = max(
                    0, (task.done_date - task.process_date).days
                )
            else:
                task.total_days = 0

    @api.depends('user_ids')
    @api.depends_context('uid')
    def _compute_is_task_assignee(self):
        uid = self.env.user.id
        for task in self:
            task.is_task_assignee = uid in task.user_ids.ids

    def _ctkm_ensure_receive_lines_synced(self):
        """Bước 11/12: dựng lại bảng nếu còn thiếu store so với kho Tem/Tag."""
        if self.env.context.get('ctkm_skip_receive_resync'):
            return
        targets = self.filtered(
            lambda t: (t.is_tem_receive_task or t.is_tem_replace_task) and t.program_id
        )
        if not targets or 'ctkm.inventory.tem.tag' not in self.env:
            return
        TemTag = self.env['ctkm.inventory.tem.tag'].sudo()
        for task in targets:
            inv_stores = set(TemTag.search([
                ('program_id', '=', task.program_id.id),
            ]).mapped('store_key'))
            inv_stores.discard(False)
            allowed = task._ctkm_upstream_sent_store_keys()
            if allowed is not None and not allowed:
                continue
            if allowed is not None:
                inv_stores = {
                    key for key in inv_stores
                    if task._ctkm_store_in_allowed(allowed, key)
                }
            line_stores = set(
                self.env['ctkm.task.tem.tag.replace.line'].sudo().search([
                    ('task_id', '=', task.id),
                ]).mapped('store_key')
            )
            line_stores.discard(False)
            if inv_stores != line_stores:
                task.sudo().with_context(
                    ctkm_skip_receive_resync=True,
                )._ctkm_sync_tem_tag_lines()

    @api.depends('completion_ids', 'completion_ids.done', 'user_ids')
    @api.depends_context('uid')
    def _compute_is_completed_by_me(self):
        uid = self.env.user.id
        for task in self:
            task.is_completed_by_me = bool(
                task.completion_ids.filtered(
                    lambda c: c.user_id.id == uid and c.done
                )
            )

    @api.depends('completion_ids', 'completion_ids.done', 'user_ids')
    def _compute_all_assigned_completed(self):
        for task in self:
            if not task.user_ids:
                task.all_assigned_completed = False
                continue
            done_user_ids = task.completion_ids.filtered('done').mapped('user_id.id')
            task.all_assigned_completed = set(done_user_ids) >= set(task.user_ids.ids)

    @api.depends('completion_ids', 'completion_ids.done', 'user_ids')
    def _compute_pending_user_ids(self):
        for task in self:
            done_user_ids = task.completion_ids.filtered('done').mapped('user_id.id')
            task.pending_user_ids = task.user_ids.filtered(
                lambda u: u.id not in done_user_ids
            )

    @api.depends('verifier_ids', 'verifier_ids.user_id')
    @api.depends_context('uid')
    def _compute_is_verifier(self):
        uid = self.env.user.id
        for task in self:
            # sudo: nhân viên thường không đọc được user_id của employee khác.
            task.is_verifier = bool(
                uid in task.sudo().verifier_ids.mapped('user_id.id')
            )

    @api.depends(
        'user_ids', 'state', 'forwarded', 'program_id',
        'checklist_line_id', 'checklist_line_id.state',
        'checklist_line_id.sequence', 'program_id.checklist_line_ids.state',
        'program_id.checklist_line_ids.sequence',
    )
    @api.depends_context('uid')
    def _compute_is_current_stage_task(self):
        first_pending = {}
        for program in self.mapped('program_id'):
            if not program:
                continue
            lines = program.checklist_line_ids.sorted(
                lambda line: (line.sequence, line.id)
            )
            first_pending[program.id] = lines.filtered(
                lambda line: line.state != 'done'
            )[:1]
        user = self.env.user
        is_admin = user.has_group('ctkm_core.group_ctkm_manager')
        for task in self:
            # Người kiểm soát / quản lý: việc đang chờ mình xác nhận là bước hiện tại.
            if (
                not is_admin
                and task.state == 'waiting_confirm'
                and not task.manager_confirmed
                and user not in task.user_ids
                and task._user_can_confirm_as_manager(user)
            ):
                task.is_current_stage_task = True
                continue
            if task.state == 'done' and not task.forwarded:
                task.is_current_stage_task = True
                continue
            current = first_pending.get(task.program_id.id)
            if current and task.checklist_line_id:
                task.is_current_stage_task = (
                    task.checklist_line_id.id == current.id
                    and task.state in ('todo', 'progress', 'waiting_confirm', 'done')
                )
                continue
            task.is_current_stage_task = False

    @api.model
    def _ctkm_my_tasks_domain(self, uid=None):
        """Việc của tôi: được giao, hoặc đang chờ tôi xác nhận (Người kiểm soát /
        quản lý org-chart / Quản lý cửa hàng).
        """
        uid = uid or self.env.uid
        return ['|', '|', '|',
            ('user_ids', 'in', [uid]),
            '&',
                ('state', '=', 'waiting_confirm'),
                ('user_ids.employee_ids.parent_id.user_id', 'in', [uid]),
            '&',
                ('state', '=', 'waiting_confirm'),
                ('verifier_ids.user_id', 'in', [uid]),
            '&',
                ('state', '=', 'waiting_confirm'),
                ('store_verifier_ids.verifier_user_id', 'in', [uid]),
        ]

    def _search_is_current_stage_task(self, operator, value):
        """Không gọi self.search() ở đây: Odoo 19 re-enter method này khi list
        view lọc 'Bước hiện tại', khiến việc chờ Người kiểm soát bị loại.

        Odoo 19 tối ưu boolean thành operator 'in' (không phải '=').
        """
        if operator in ('=', '=='):
            want_true = bool(value)
        elif operator in ('!=', '<>'):
            want_true = not value
        elif operator == 'in':
            try:
                want_true = True in value or 1 in value
            except TypeError:
                want_true = bool(value)
        elif operator == 'not in':
            try:
                want_true = True not in value and 1 not in value
            except TypeError:
                want_true = not value
        else:
            want_true = bool(value)
        uid = self.env.uid
        uid = self.env.uid
        cr = self.env.cr
        cr.execute("""
            SELECT t.id FROM ctkm_task t
            WHERE EXISTS (
                SELECT 1 FROM ctkm_task_user_rel r
                WHERE r.task_id = t.id AND r.user_id = %s
            )
            UNION
            SELECT t.id FROM ctkm_task t
            WHERE t.state = 'waiting_confirm'
              AND EXISTS (
                SELECT 1 FROM ctkm_task_verifier_rel r
                JOIN hr_employee e ON e.id = r.employee_id
                WHERE r.task_id = t.id AND e.user_id = %s
              )
            UNION
            SELECT t.id FROM ctkm_task t
            JOIN ctkm_task_store_verifier sv ON sv.task_id = t.id
            JOIN hr_employee e ON e.id = sv.verifier_id
            WHERE t.state = 'waiting_confirm' AND e.user_id = %s
            UNION
            SELECT t.id FROM ctkm_task t
            JOIN ctkm_task_user_rel r ON r.task_id = t.id
            JOIN hr_employee worker ON worker.user_id = r.user_id
            JOIN hr_employee mgr ON mgr.id = worker.parent_id
            WHERE t.state = 'waiting_confirm' AND mgr.user_id = %s
        """, (uid, uid, uid, uid))
        candidate_ids = [row[0] for row in cr.fetchall()]
        tasks = self.browse(candidate_ids)
        matching_ids = []
        user = self.env.user
        for task in tasks:
            if task.is_current_stage_task:
                matching_ids.append(task.id)
                continue
            if (
                task.state == 'waiting_confirm'
                and not task.manager_confirmed
                and task._user_can_confirm_as_manager(user)
            ):
                matching_ids.append(task.id)
        if want_true:
            return [('id', 'in', matching_ids)]
        return [('id', 'not in', matching_ids)]

    @api.depends('checklist_line_id', 'checklist_line_id.need_manager_confirm')
    def _compute_checklist_need_manager_confirm(self):
        for task in self:
            task.checklist_need_manager_confirm = (
                task.checklist_line_id.need_manager_confirm
                if task.checklist_line_id
                else True
            )

    @api.depends(
        'is_tem_print_task', 'is_tem_handover_task', 'is_tem_receive_task',
        'is_tem_postcheck_task', 'is_tem_price_task',
    )
    def _compute_show_work_processing_tab(self):
        for task in self:
            task.show_work_processing_tab = bool(
                task.is_tem_print_task
                or task.is_tem_handover_task
                or task.is_tem_receive_task
                or task.is_tem_postcheck_task
                or task.is_tem_price_task
            )

    @api.depends('name', 'checklist_line_id.name', 'checklist_line_id.stage_id')
    def _compute_task_step_flags(self):
        stage_ids = {
            flag: self._ctkm_step_stage_id(xmlid)
            for flag, xmlid in (
                ('import', TEM_TAG_IMPORT_STAGE_XMLID),
                ('bb_replace', TEM_BB_REPLACE_STAGE_XMLID),
                ('print', TEM_PRINT_STAGE_XMLID),
                ('handover', TEM_HANDOVER_STAGE_XMLID),
                ('receive', TEM_RECEIVE_STAGE_XMLID),
                ('replace', TEM_REPLACE_STAGE_XMLID),
                ('photo', TEM_PHOTO_STAGE_XMLID),
                ('check', TEM_CHECK_STAGE_XMLID),
                ('design', TEM_DESIGN_STAGE_XMLID),
                ('postcheck', TEM_POSTCHECK_STAGE_XMLID),
                ('price', TEM_PRICE_STAGE_XMLID),
            )
        }
        for task in self:
            keys = [
                key
                for key in (
                    normalize_step_key(task.name),
                    normalize_step_key(task.checklist_step_name),
                )
                if key
            ]
            stage_id = task.checklist_line_id.stage_id.id
            is_import = (
                (stage_id and stage_id == stage_ids['import'])
                or any(marker in key for key in keys for marker in TEM_TAG_IMPORT_TASK_MARKERS)
            )
            is_photo = (
                (stage_id and stage_id == stage_ids['photo'])
                or any(marker in key for key in keys for marker in TEM_PHOTO_TASK_MARKERS)
            )
            is_check = (
                (stage_id and stage_id == stage_ids['check'])
                or any(marker in key for key in keys for marker in TEM_CHECK_TASK_MARKERS)
            )
            # So khớp tuyệt đối để bước 4 / 6 / 15 (tên cũng chứa "thay tem tag")
            # không bị nhận nhầm thành bước 12.
            is_replace = (
                (stage_id and stage_id == stage_ids['replace'])
                or any(key in TEM_REPLACE_TASK_KEYS for key in keys)
            )
            is_bb_replace = (
                (stage_id and stage_id == stage_ids['bb_replace'])
                or any(marker in key for key in keys for marker in TEM_BB_REPLACE_TASK_MARKERS)
            )
            is_print = (
                (stage_id and stage_id == stage_ids['print'])
                or any(marker in key for key in keys for marker in TEM_PRINT_TASK_MARKERS)
            )
            is_handover = (
                (stage_id and stage_id == stage_ids['handover'])
                or any(marker in key for key in keys for marker in TEM_HANDOVER_TASK_MARKERS)
            )
            is_receive = (
                (stage_id and stage_id == stage_ids['receive'])
                or any(marker in key for key in keys for marker in TEM_RECEIVE_TASK_MARKERS)
            )
            is_design = (
                (stage_id and stage_id == stage_ids['design'])
                or any(marker in key for key in keys for marker in TEM_DESIGN_TASK_MARKERS)
            )
            is_postcheck = (
                (stage_id and stage_id == stage_ids['postcheck'])
                or any(marker in key for key in keys for marker in TEM_POSTCHECK_TASK_MARKERS)
            )
            is_price = (
                (stage_id and stage_id == stage_ids['price'])
                or any(marker in key for key in keys for marker in TEM_PRICE_TASK_MARKERS)
            )
            # Bước 6 "Lập BB thay tem": nhận diện riêng (không trùng bước 4 / 12 / 15).
            is_bb_replace = (
                (stage_id and stage_id == stage_ids['bb_replace'])
                or any(marker in key for key in keys for marker in TEM_BB_REPLACE_TASK_MARKERS)
            )
            task.is_tem_tag_import_task = bool(is_import)
            task.is_tem_photo_task = bool(is_photo)
            task.is_tem_check_task = bool(is_check)
            task.is_tem_replace_task = bool(
                is_replace and not is_import and not is_bb_replace and not is_price
            )
            task.is_tem_bb_replace_task = bool(is_bb_replace and not is_import and not is_replace)
            task.is_tem_print_task = bool(is_print and not is_handover and not is_receive)
            task.is_tem_handover_task = bool(is_handover)
            task.is_tem_receive_task = bool(is_receive and not is_handover)
            task.is_tem_postcheck_task = bool(
                is_postcheck
                and not is_import
                and not is_print
                and not is_handover
                and not is_receive
                and not is_replace
                and not is_price
            )
            task.is_tem_price_task = bool(
                is_price
                and not is_import
                and not is_print
                and not is_handover
                and not is_receive
                and not is_replace
                and not is_postcheck
            )
            task.is_tem_design_task = bool(
                is_design
                and not is_import
                and not is_bb_replace
                and not is_print
                and not is_handover
                and not is_receive
                and not is_replace
                and not is_photo
                and not is_postcheck
                and not is_price
            )

    @api.model
    def _ctkm_step_stage_id(self, xmlid):
        stage = self.env.ref(xmlid, raise_if_not_found=False)
        return stage.id if stage else False

    # --- Bảng "Chi tiết tem/tag" (bước 4, 9–12) ---

    def _ctkm_tem_tag_store_keys(self):
        """Mã cửa hàng (HRM 'Cửa hàng' / 'Mã bộ phận') của người nhận việc."""
        self.ensure_one()
        tem_tag = self.env['ctkm.inventory.tem.tag'].sudo()
        if hasattr(tem_tag, 'store_keys_for_user'):
            return tem_tag.store_keys_for_user(self.user_ids[:1])
        return tem_tag.current_user_store_keys()

    def _ctkm_tem_tag_receive_store_keys(self):
        """Mã cửa hàng bước 11 "Nhận tem tag mới" (theo người đang xem)."""
        self.ensure_one()
        return list(self._ctkm_current_user_visible_store_keys())

    @api.model
    def _ctkm_user_department_store_keys(self, user):
        """Mã bộ phận / cửa hàng trên hồ sơ của một user, đã chuẩn hóa."""
        user = user.sudo() if user else self.env['res.users']
        if not user:
            return set()
        employees = self.env['hr.employee'].sudo()
        if 'employee_id' in user._fields and user.employee_id:
            employees |= user.employee_id.sudo()
        if user.employee_ids:
            employees |= user.employee_ids.sudo()
        keys = set()
        for employee in employees:
            keys |= self._ctkm_employee_department_store_keys(employee)
        return keys

    @api.model
    def _ctkm_current_user_visible_store_keys(self):
        return self._ctkm_visible_store_keys_for_user(self.env.user)

    @api.model
    def _ctkm_visible_store_keys_for_user(self, user):
        """Mã Store trên biên bản mà *user* được xem (bước 11–14).

        Khớp chính xác mã bộ phận. Chỉ gộp ``{store}_DDL`` / ``{store}_DNA``
        vào mã ngắn trên file (AETL) khi hồ sơ có cả hai hậu tố và không có
        mã trần; cột file đã có hậu tố (AETA_DNA) không gộp.
        Không lấy LUG Permission — tránh lộ store của cửa hàng khác.
        """
        dept_keys = self._ctkm_user_department_store_keys(user)
        visible = set(dept_keys)
        grouped_parents = self._ctkm_grouped_parent_store_keys()
        for dept in list(dept_keys):
            parent = self._ctkm_grouped_store_key(dept)
            if parent and parent in grouped_parents:
                visible.add(parent)
        if 'ctkm.inventory.tem.tag' in self.env and dept_keys:
            tem_tag = self.env['ctkm.inventory.tem.tag'].sudo()
            groups = tem_tag._read_group(
                [('store_key', '!=', False)],
                ['store_key'],
                ['id:count'],
            )
            for store_key, _count in groups:
                key = _ctkm_normalize_store_key(store_key)
                if key and any(
                    self._ctkm_department_matches_store(dept, key)
                    for dept in dept_keys
                ):
                    visible.add(key)
        return [key for key in visible if key]

    @api.depends('is_tem_photo_task', 'is_tem_check_task')
    def _compute_current_user_visible_store_keys(self):
        """Danh sách store_key dùng domain lọc bảng bước 13–14 trên form."""
        user = self.env.user
        is_manager = user.has_group('ctkm_core.group_ctkm_manager')
        visible = [] if is_manager else list(
            self._ctkm_visible_store_keys_for_user(user)
        )
        for task in self:
            if is_manager:
                keys = {
                    line.store_key or _ctkm_normalize_store_key(line.store)
                    for line in task.sudo().tem_photo_check_ids
                    if line.store_key or line.store
                }
                task.current_user_visible_store_keys = [key for key in keys if key]
            else:
                task.current_user_visible_store_keys = visible

    def _ctkm_store_visible_to_user(self, store_value):
        """True nếu mã Store trên dòng thuộc cửa hàng của user đang xem."""
        key = _ctkm_normalize_store_key(store_value)
        if not key:
            return False
        if self.env.user.has_group('ctkm_core.group_ctkm_manager'):
            return True
        if key in set(self._ctkm_current_user_visible_store_keys()):
            return True
        if key in set(self._ctkm_tem_tag_managed_store_keys()):
            return True
        dept_keys = self._ctkm_user_department_store_keys(self.env.user)
        return any(
            self._ctkm_department_matches_store(dept, key)
            for dept in dept_keys
        )

    def _ctkm_tem_tag_managed_store_keys(self):
        """Mã cửa hàng theo 'Cửa hàng quản lí' của người nhận việc (bước 6 và bước 14)."""
        self.ensure_one()
        tem_tag = self.env['ctkm.inventory.tem.tag'].sudo()
        if hasattr(tem_tag, 'managed_store_keys_for_user'):
            return tem_tag.managed_store_keys_for_user(self.user_ids)
        return []

    def _ctkm_bb_store_keys(self):
        """Mã Store trên biên bản bước 4 (bảng Chi tiết tem/tag / kho Tem-Tag)."""
        self.ensure_one()
        keys = set()
        for line in self.tem_tag_replace_ids:
            key = _ctkm_normalize_store_key(line.store)
            if key:
                keys.add(key)
        if self.program_id and 'ctkm.inventory.tem.tag' in self.env:
            rows = self.env['ctkm.inventory.tem.tag'].sudo().search([
                ('program_id', '=', self.program_id.id),
            ])
            for rec in rows:
                key = rec.store_key or _ctkm_normalize_store_key(rec.store)
                if key:
                    keys.add(key)
        return keys

    def _ctkm_employee_department_store_keys(self, employee):
        """Mã bộ phận / cửa hàng trên hồ sơ nhân viên, đã chuẩn hóa.

        Chỉ lấy *mã* (code), không lấy tên cửa hàng — tránh khớp nhầm
        nhiều store LUG_* khi tên chứa tiền tố chung.
        """
        codes = []
        if 'ma_bo_phan' in employee._fields:
            codes.append(employee.ma_bo_phan)
        if 'ma_bo_phan_id' in employee._fields and employee.ma_bo_phan_id:
            codes.append(employee.ma_bo_phan_id.sudo().code)
        if 'store_id' in employee._fields and employee.store_id:
            codes.append(employee.store_id.sudo().code)
        keys = set()
        for code in codes:
            key = _ctkm_normalize_store_key(code)
            if key:
                keys.add(key)
        return keys

    @api.model
    def _ctkm_grouped_store_key(self, store_key):
        """Mã gốc nếu *store_key* đã có hậu tố ``_DDL`` / ``_DNA`` (AETA_DNA → AETA)."""
        if not store_key:
            return False
        for suffix in STORE_DEPT_GROUP_SUFFIXES:
            if store_key.endswith(suffix):
                parent = store_key[:-len(suffix)]
                if parent and len(parent) >= 3:
                    return parent
        return False

    @api.model
    def _ctkm_hr_ma_bo_phan_keys(self):
        """Mã bộ phận đang dùng trên hồ sơ nhân viên (active), đã chuẩn hóa."""
        keys = set()
        Employee = self.env['hr.employee'].sudo()
        if 'ma_bo_phan_id' not in Employee._fields:
            return keys
        groups = Employee._read_group(
            [('active', '=', True), ('ma_bo_phan_id', '!=', False)],
            ['ma_bo_phan_id'],
            ['id:count'],
        )
        for rec, _count in groups:
            if not rec:
                continue
            rec = rec.sudo()
            key = _ctkm_normalize_store_key(rec.code)
            if key:
                keys.add(key)
        return keys

    @api.model
    def _ctkm_grouped_parent_store_keys(self):
        """Mã ngắn trên file được gộp (AETL, AEBD).

        Chỉ khi hồ sơ có cả ``{store}_DDL`` và ``{store}_DNA``, và không có
        mã trần ``{store}``. Cache theo cursor (một lần mỗi request).
        """
        cached = getattr(self.env.cr, '_ctkm_grouped_parent_store_keys', None)
        if cached is not None:
            return cached
        dept_keys = self._ctkm_hr_ma_bo_phan_keys()
        parents = set()
        candidates = set()
        for key in dept_keys:
            parent = self._ctkm_grouped_store_key(key)
            if parent:
                candidates.add(parent)
        for parent in candidates:
            if parent in dept_keys:
                continue
            if all(
                f'{parent}{suffix}' in dept_keys
                for suffix in STORE_DEPT_GROUP_SUFFIXES
            ):
                parents.add(parent)
        cached = frozenset(parents)
        self.env.cr._ctkm_grouped_parent_store_keys = cached
        return cached

    @api.model
    def _ctkm_department_matches_store(self, dept_key, store_key):
        """Khớp mã bộ phận HRM với mã Store trên biên bản / file Excel.

        * Khớp chính xác luôn được (AETA_DNA = AETA_DNA).
        * Cột file đã có ``_DDL`` / ``_DNA`` → không gộp thêm.
        * Cột mã ngắn (AETL) chỉ gộp người ``AETL_DDL`` / ``AETL_DNA`` khi
          hồ sơ có cả hai hậu tố và không có mã trần AETL.
        """
        if not dept_key or not store_key:
            return False
        if dept_key == store_key:
            return True
        if self._ctkm_grouped_store_key(store_key):
            return False
        if store_key not in self._ctkm_grouped_parent_store_keys():
            return False
        return any(
            dept_key == f'{store_key}{suffix}'
            for suffix in STORE_DEPT_GROUP_SUFFIXES
        )

    def _ctkm_employee_matches_bb_stores(self, employee, store_keys):
        dept_keys = self._ctkm_employee_department_store_keys(employee)
        if not dept_keys or not store_keys:
            return False
        return any(
            self._ctkm_department_matches_store(dept_key, store_key)
            for dept_key in dept_keys
            for store_key in store_keys
        )

    def _ctkm_find_store_manager_users(self, store_keys):
        """Cửa hàng trưởng có mã bộ phận khớp Store trên biên bản.

        Khớp chính xác; chỉ gộp ``{store}_DDL`` + ``{store}_DNA`` vào mã ngắn
        khi hồ sơ có cả hai và không có mã trần (AETL). Cột file đã có hậu tố
        (AETA_DNA) giữ khớp đúng mã.
        """
        if not store_keys:
            return self.env['res.users']
        Employee = self.env['hr.employee'].sudo()
        domain = [
            ('active', '=', True),
            ('user_id', '!=', False),
        ]
        if 'job_title' in Employee._fields:
            domain.append(('job_title', '=', STORE_MANAGER_JOB_TITLE))
        employees = Employee.search(domain)
        matched = employees.filtered(
            lambda emp: self._ctkm_employee_matches_bb_stores(emp, store_keys)
        )
        return matched.mapped('user_id').filtered(
            lambda user: user.active and not user.share
        )

    def _ctkm_checklist_line_is_store_manager_step(self, line):
        """Bước 11 'Nhận tem tag mới' và bước 12 'Thay tem Tag': Phụ trách là
        Cửa hàng trưởng (theo chức danh trên hồ sơ).

        Bước 13 'Chụp team...' và 14 'Kiểm tra hình ảnh' KHÔNG dùng cơ chế này:
        bước 13 lấy Quản lý cửa hàng làm Phụ trách (xem
        _ctkm_assign_photo_store_managers_after_bb_import), bước 14 giữ Phụ trách
        cấu hình sẵn trong Giai đoạn.
        """
        if not line:
            return False
        stage_id = line.stage_id.id if line.stage_id else False
        target_stage_ids = {
            self._ctkm_step_stage_id(TEM_RECEIVE_STAGE_XMLID),
            self._ctkm_step_stage_id(TEM_REPLACE_STAGE_XMLID),
        }
        target_stage_ids.discard(False)
        if stage_id and stage_id in target_stage_ids:
            return True
        key = normalize_step_key(line.name)
        if not key:
            return False
        if any(marker in key for marker in TEM_RECEIVE_TASK_MARKERS):
            return True
        if key in TEM_REPLACE_TASK_KEYS:
            return True
        return False

    def _ctkm_checklist_line_is_photo_step(self, line):
        """Bước 13 'Chụp team gửi lên group / chụp từng con tem'."""
        if not line:
            return False
        stage_id = line.stage_id.id if line.stage_id else False
        photo_id = self._ctkm_step_stage_id(TEM_PHOTO_STAGE_XMLID)
        if stage_id and photo_id and stage_id == photo_id:
            return True
        key = normalize_step_key(line.name)
        return bool(key and any(marker in key for marker in TEM_PHOTO_TASK_MARKERS))

    def _ctkm_assign_store_managers_after_bb_import(self):
        """Khi hoàn thành bước 4: gán Cửa hàng trưởng vào người thực hiện bước 11–12.

        Chỉ áp dụng cho 'Nhận tem tag mới' (11) và 'Thay tem Tag' (12). Bước 13
        (Chụp ảnh) và 14 (Kiểm tra ảnh) không dùng cơ chế này — xem
        _ctkm_assign_photo_store_managers_after_bb_import.

        Khớp mã bộ phận trên hồ sơ với mã Store trên file: chính xác mặc định.
        Chỉ gộp khi file ghi mã ngắn (AETL / AEBD) và hồ sơ có cả
        ``{store}_DDL`` lẫn ``{store}_DNA``, không có mã trần. Cột đã có hậu tố
        (AETA_DNA) không gộp. Chức danh phải là Cửa hàng trưởng.
        """
        self.ensure_one()
        if not self.is_tem_tag_import_task or not self.program_id:
            return self.env['res.users']
        store_keys = self._ctkm_bb_store_keys()
        if not store_keys:
            return self.env['res.users']
        users = self._ctkm_find_store_manager_users(store_keys)
        target_lines = self.program_id.checklist_line_ids.filtered(
            lambda line: (
                self._ctkm_checklist_line_is_store_manager_step(line)
                and line.state != 'done'
            )
        )
        if users:
            user_ids = users.ids
            for line in target_lines:
                if set(line.user_ids.ids) == set(user_ids):
                    continue
                line.sudo().write({'user_ids': [(6, 0, user_ids)]})
            # Công việc đã tạo theo cờ bước (checklist đổi tên / chưa gắn stage).
            flag_tasks = self.env['ctkm.task'].sudo().search([
                ('program_id', '=', self.program_id.id),
                ('state', '!=', 'done'),
                '|',
                ('is_tem_receive_task', '=', True),
                ('is_tem_replace_task', '=', True),
            ])
            for other in flag_tasks:
                line = other.checklist_line_id
                if line and line.state != 'done':
                    if set(line.user_ids.ids) != set(user_ids):
                        line.sudo().write({'user_ids': [(6, 0, user_ids)]})
                    continue
                if set(other.user_ids.ids) == set(user_ids):
                    continue
                other.with_context(
                    ctkm_task_sync=True,
                    ctkm_internal_state_write=True,
                ).write({'user_ids': [(6, 0, user_ids)]})
        store_label = ', '.join(sorted(store_keys))
        if users:
            body = _(
                'Đã gán cửa hàng trưởng <b>%(users)s</b> làm người thực hiện '
                'bước 11–12 (Nhận tem tag mới, Thay tem Tag) theo mã Store trên '
                'biên bản: %(stores)s.'
            ) % {
                'users': escape(', '.join(users.mapped('name'))),
                'stores': escape(store_label),
            }
        else:
            body = _(
                'Không tìm thấy nhân viên chức danh Cửa hàng trưởng có '
                'mã bộ phận khớp Store trên biên bản thay tem/tag (%s) '
                '(khớp đúng mã, hoặc gộp _DDL+_DNA vào mã ngắn khi hồ sơ '
                'có cả hai và không có mã trần). '
                'Bước 11–12 giữ nguyên người thực hiện hiện tại.'
            ) % escape(store_label)
        self.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return users

    def _ctkm_assign_photo_store_managers_after_bb_import(self):
        """Khi hoàn thành bước 4: gán Quản lý cửa hàng (hr.store.manager_id) của
        từng Store trên biên bản làm Phụ trách bước 13 'Chụp team gửi lên group /
        chụp từng con tem'.

        Khác với bước 11–12 (Phụ trách = Cửa hàng trưởng theo chức danh), bước 13
        lấy Quản lý cửa hàng được cấu hình tại từng Cửa hàng
        (Employee → Cấu hình → Cửa hàng) của mọi mã Store có trên biên bản tổng.
        Người kiểm soát (verifier_ids) của bước 13 giữ nguyên theo cấu hình.
        """
        self.ensure_one()
        if not self.is_tem_tag_import_task or not self.program_id:
            return self.env['res.users']
        store_keys = self._ctkm_bb_store_keys()
        if not store_keys:
            return self.env['res.users']
        # Quản lý cửa hàng của từng Store trên biên bản (hr.store.manager_id).
        managers = self._ctkm_store_manager_employees_for_bb_stores(store_keys)
        users = managers.mapped('user_id').filtered(
            lambda user: user and user.active and not user.share
        )
        store_label = ', '.join(sorted(store_keys))
        if not users:
            self.message_post(
                body=_(
                    'Không tìm thấy nhân viên Quản lý cửa hàng (Employee → '
                    'Cấu hình → Cửa hàng) tương ứng với mã Store trên biên bản: '
                    '%(stores)s. Bước 13 (Chụp team...) giữ nguyên người phụ '
                    'trách hiện tại.'
                ) % {'stores': escape(store_label)},
                subtype_xmlid='mail.mt_note',
                body_is_html=True,
            )
            return users
        user_ids = users.ids
        target_lines = self.program_id.checklist_line_ids.filtered(
            lambda line: (
                self._ctkm_checklist_line_is_photo_step(line)
                and line.state != 'done'
            )
        )
        if target_lines:
            for line in target_lines:
                if set(line.user_ids.ids) == set(user_ids):
                    continue
                line.sudo().write({'user_ids': [(6, 0, user_ids)]})
        # Công việc đã tạo theo cờ bước (checklist đổi tên / chưa gắn stage).
        photo_tasks = self.env['ctkm.task'].sudo().search([
            ('program_id', '=', self.program_id.id),
            ('state', '!=', 'done'),
            ('is_tem_photo_task', '=', True),
        ])
        for other in photo_tasks:
            line = other.checklist_line_id
            if line and line.state != 'done':
                if set(line.user_ids.ids) != set(user_ids):
                    line.sudo().write({'user_ids': [(6, 0, user_ids)]})
                continue
            if set(other.user_ids.ids) == set(user_ids):
                continue
            other.with_context(
                ctkm_task_sync=True,
                ctkm_internal_state_write=True,
            ).write({'user_ids': [(6, 0, user_ids)]})
        self.message_post(
            body=_(
                'Đã gán Quản lý cửa hàng <b>%(users)s</b> làm người phụ trách '
                'bước 13 (Chụp team gửi lên group / chụp từng con tem) theo mã '
                'Store trên biên bản: %(stores)s.'
            ) % {
                'users': escape(', '.join(users.mapped('name'))),
                'stores': escape(store_label),
            },
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return users

    def _ctkm_miens_for_bb_stores(self, store_keys):
        """Miền của các mã Store trên biên bản, lấy từ danh mục cửa hàng (hr.store)."""
        if not store_keys:
            return set()
        miens = set()
        Store = (
            self.env['hr.store'].sudo()
            if 'hr.store' in self.env
            else self.env['hr.store.code'].sudo()
        )
        stores = Store.search([('code', '!=', False)])
        for store in stores:
            code = _ctkm_normalize_store_key(store.code)
            name = _ctkm_normalize_store_key(store.name)
            if not store.mien:
                continue
            if any(
                self._ctkm_department_matches_store(code, key)
                or (name and self._ctkm_department_matches_store(name, key))
                or key == code
                for key in store_keys
            ):
                miens.add(store.mien)
        return miens

    def _ctkm_find_supervisor_users(self, miens):
        """Giám sát thuộc một trong các miền của cửa hàng trên biên bản."""
        if not miens:
            return self.env['res.users']
        Employee = self.env['hr.employee'].sudo()
        domain = [
            ('active', '=', True),
            ('user_id', '!=', False),
        ]
        if 'job_title' in Employee._fields:
            domain.append(('job_title', '=', SUPERVISOR_JOB_TITLE))
        if 'mien' in Employee._fields:
            domain.append(('mien', 'in', list(miens)))
        employees = Employee.search(domain)
        return employees.mapped('user_id').filtered(
            lambda user: user.active and not user.share
        )

    def _ctkm_checklist_line_is_postcheck_step(self, line):
        """Bước Hậu kiểm (Giám sát đi kiểm tra thay tem)."""
        if not line:
            return False
        stage_id = line.stage_id.id if line.stage_id else False
        postcheck_id = self._ctkm_step_stage_id(TEM_POSTCHECK_STAGE_XMLID)
        if stage_id and postcheck_id and stage_id == postcheck_id:
            return True
        key = normalize_step_key(line.name)
        return bool(key and any(marker in key for marker in TEM_POSTCHECK_TASK_MARKERS))

    def _ctkm_assign_supervisors_after_bb_import(self):
        """Khi hoàn thành bước 4: gán Giám sát (theo miền của Store) vào bước Hậu kiểm."""
        self.ensure_one()
        if not self.is_tem_tag_import_task or not self.program_id:
            return self.env['res.users']
        store_keys = self._ctkm_bb_store_keys()
        if not store_keys:
            return self.env['res.users']
        miens = self._ctkm_miens_for_bb_stores(store_keys)
        users = self._ctkm_find_supervisor_users(miens)
        target_lines = self.program_id.checklist_line_ids.filtered(
            lambda line: (
                self._ctkm_checklist_line_is_postcheck_step(line)
                and line.state != 'done'
            )
        )
        if users:
            user_ids = users.ids
            for line in target_lines:
                if set(line.user_ids.ids) == set(user_ids):
                    continue
                line.sudo().write({'user_ids': [(6, 0, user_ids)]})
            flag_tasks = self.env['ctkm.task'].sudo().search([
                ('program_id', '=', self.program_id.id),
                ('state', '!=', 'done'),
            ]).filtered(
                lambda t: self._ctkm_checklist_line_is_postcheck_step(t.checklist_line_id)
                or any(
                    marker in (normalize_step_key(t.name) or '')
                    for marker in TEM_POSTCHECK_TASK_MARKERS
                )
            )
            for other in flag_tasks:
                line = other.checklist_line_id
                if line and line.state != 'done':
                    if set(line.user_ids.ids) != set(user_ids):
                        line.sudo().write({'user_ids': [(6, 0, user_ids)]})
                    continue
                if set(other.user_ids.ids) == set(user_ids):
                    continue
                other.with_context(
                    ctkm_task_sync=True,
                    ctkm_internal_state_write=True,
                ).write({'user_ids': [(6, 0, user_ids)]})
        store_label = ', '.join(sorted(store_keys))
        mien_label = ', '.join(sorted(miens)) if miens else _('(chưa xác định)')
        if users:
            body = _(
                'Đã gán giám sát <b>%(users)s</b> làm người thực hiện bước '
                'Hậu kiểm theo miền %(mien)s của mã Store trên biên bản: %(stores)s.'
            ) % {
                'users': escape(', '.join(users.mapped('name'))),
                'mien': escape(mien_label),
                'stores': escape(store_label),
            }
        else:
            body = _(
                'Không tìm thấy nhân viên chức danh Giám sát thuộc miền %(mien)s '
                '(suy ra từ Store trên biên bản: %(stores)s). '
                'Bước Hậu kiểm giữ nguyên người thực hiện hiện tại.'
            ) % {
                'mien': escape(mien_label),
                'stores': escape(store_label),
            }
        self.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return users

    def _ctkm_checklist_line_is_receive_or_replace_step(self, line):
        """Bước 11 'Nhận tem tag mới', bước 12 'Thay tem Tag' và bước 10
        'Bàn giao Tem Tag / Thu hồi tem tag cũ' (các bước dùng Quản lý cửa
        hàng làm Người kiểm soát theo từng cửa hàng)."""
        if not line:
            return False
        stage_id = line.stage_id.id if line.stage_id else False
        target_stage_ids = {
            self._ctkm_step_stage_id(TEM_HANDOVER_STAGE_XMLID),
            self._ctkm_step_stage_id(TEM_RECEIVE_STAGE_XMLID),
            self._ctkm_step_stage_id(TEM_REPLACE_STAGE_XMLID),
        }
        target_stage_ids.discard(False)
        if stage_id and stage_id in target_stage_ids:
            return True
        key = normalize_step_key(line.name)
        if not key:
            return False
        if any(marker in key for marker in TEM_HANDOVER_TASK_MARKERS):
            return True
        if any(marker in key for marker in TEM_RECEIVE_TASK_MARKERS):
            return True
        if key in TEM_REPLACE_TASK_KEYS:
            return True
        return False

    def _ctkm_store_manager_employees_for_bb_stores(self, store_keys):
        """Quản lý cửa hàng (hr.store.manager_id) của từng Store trên biên bản.

        'Quản lý cửa hàng' được cấu hình tại Employee → Cấu hình → Cửa hàng
        (model hr.store, trường manager_id), định nghĩa trong module hr_store.

        Mỗi mã Store trên biên bản chỉ lấy ĐÚNG 1 Quản lý cửa hàng (cửa hàng
        khớp mã) để đảm bảo quan hệ 1 Phụ trách (Cửa hàng trưởng) – 1 Người kiểm
        soát (Quản lý cửa hàng) theo CÙNG một cửa hàng: chief của store X chỉ
        được xác nhận bởi Quản lý cửa hàng của store X.
        """
        if not store_keys or 'hr.store' not in self.env:
            return self.env['hr.employee']
        return self.env['hr.employee'].sudo().browse({
            mgr.id for mgr in self._ctkm_store_manager_map(store_keys).values()
        })

    def _ctkm_store_manager_map(self, store_keys):
        """Map mã Store (trên biên bản) → Quản lý cửa hàng (hr.store.manager_id).

        Mỗi mã Store chỉ lấy ĐÚNG 1 Quản lý cửa hàng (khớp chính xác mã cửa
        hàng, fallback ``{store}_DDL`` / ``{store}_DNA``) để đảm bảo chief của
        store X chỉ được xác nhận bởi Quản lý cửa hàng của store X.
        """
        if not store_keys or 'hr.store' not in self.env:
            return {}
        Store = self.env['hr.store'].sudo()
        stores = Store.search([('code', '!=', False)])
        by_exact = {}
        by_contains = {}
        for store in stores:
            if not store.manager_id:
                continue
            code = _ctkm_normalize_store_key(store.code)
            name = _ctkm_normalize_store_key(store.name)
            for key in store_keys:
                if key == code or (name and key == name):
                    by_exact.setdefault(key, store.manager_id)
                elif self._ctkm_department_matches_store(code, key) or (
                    name and self._ctkm_department_matches_store(name, key)
                ):
                    by_contains.setdefault(key, store.manager_id)
        result = {}
        for key in store_keys:
            if key in by_exact:
                result[key] = by_exact[key]
            elif key in by_contains:
                result[key] = by_contains[key]
        return result

    def _ctkm_users_for_store_key(self, store_key):
        """Các người nhận việc (user_ids) thuộc cửa hàng có mã = store_key."""
        self.ensure_one()
        result = self.env['res.users']
        for user in self.user_ids:
            emp_keys = self._ctkm_user_department_store_keys(user)
            if any(
                self._ctkm_department_matches_store(dk, store_key)
                for dk in emp_keys
            ):
                result |= user
        return result

    def _ctkm_assign_store_managers_verifier_after_bb_import(self):
        """Khi hoàn thành bước 4: gán Quản lý cửa hàng (hr.store.manager_id)
        làm Người kiểm soát bước 10 'Bàn giao Tem Tag / Thu hồi tem tag cũ',
        bước 11 'Nhận tem tag mới' và bước 12 'Thay tem Tag' theo từng cửa
        hàng trên biên bản tổng.

        Khác với 'Phụ trách' (Cửa hàng trưởng - chức danh trên hồ sơ),
        'Người kiểm soát' ở đây lấy từ cấu hình Cửa hàng (Quản lý cửa hàng).
        """
        self.ensure_one()
        if not self.is_tem_tag_import_task or not self.program_id:
            return self.env['hr.employee']
        store_keys = self._ctkm_bb_store_keys()
        if not store_keys:
            return self.env['hr.employee']
        store_manager_map = self._ctkm_store_manager_map(store_keys)
        managers = self.env['hr.employee'].sudo().browse({
            m.id for m in store_manager_map.values()
        })
        store_label = ', '.join(sorted(store_keys))
        if not managers:
            self.message_post(
                body=_(
                    'Không tìm thấy nhân viên Quản lý cửa hàng (Employee → '
                    'Cấu hình → Cửa hàng) tương ứng với mã Store trên biên bản: '
                    '%(stores)s. Bước 10–12 giữ nguyên Người kiểm soát hiện tại.'
                ) % {'stores': escape(store_label)},
                subtype_xmlid='mail.mt_note',
                body_is_html=True,
            )
            return managers
        manager_ids = managers.ids
        target_lines = self.program_id.checklist_line_ids.filtered(
            lambda line: (
                self._ctkm_checklist_line_is_receive_or_replace_step(line)
                and line.state != 'done'
            )
        )
        if target_lines:
            for line in target_lines:
                if set(line.verifier_ids.ids) == set(manager_ids):
                    continue
                line.sudo().write({'verifier_ids': [(6, 0, manager_ids)]})
        flag_tasks = self.env['ctkm.task'].sudo().search([
            ('program_id', '=', self.program_id.id),
            ('state', '!=', 'done'),
            '|', '|',
            ('is_tem_handover_task', '=', True),
            ('is_tem_receive_task', '=', True),
            ('is_tem_replace_task', '=', True),
        ])
        for other in flag_tasks:
            line = other.checklist_line_id
            if line and line.state != 'done':
                if set(line.verifier_ids.ids) != set(manager_ids):
                    line.sudo().write({'verifier_ids': [(6, 0, manager_ids)]})
            elif set(other.verifier_ids.ids) != set(manager_ids):
                other.with_context(
                    ctkm_task_sync=True,
                    ctkm_internal_state_write=True,
                ).write({'verifier_ids': [(6, 0, manager_ids)]})
            # Quan hệ 1 Phụ trách – 1 Người kiểm soát theo TỪNG cửa hàng.
            # Bước 10 (Bàn giao Tem Tag) không có Phụ trách riêng → mỗi cửa
            # hàng một dòng, Quản lý cửa hàng xác nhận trực tiếp (no_assignee).
            other.sudo()._ctkm_build_store_verifier_lines(store_keys, store_manager_map)
        uncovered = self.env['res.users']
        for other in flag_tasks:
            if other.is_tem_handover_task:
                # Bước 10 không có Phụ trách riêng nên bỏ qua cảnh báo thiếu.
                continue
            covered = other.store_verifier_ids.mapped('assignee_user_id')
            uncovered |= (other.user_ids - covered)
        if uncovered:
            self.message_post(
                body=_(
                    'Lưu ý: một số Phụ trách không có Quản lý cửa hàng tương ứng '
                    '(thiếu cấu hình Cửa hàng) nên không được xác nhận riêng: '
                    '%(users)s. CTKM Admin có thể xác nhận thay.'
                ) % {'users': escape(', '.join(uncovered.mapped('name')))},
                subtype_xmlid='mail.mt_note',
                body_is_html=True,
            )
        self.message_post(
            body=_(
                'Đã gán <b>%(users)s</b> (Quản lý cửa hàng) làm Người kiểm soát '
                'bước 10–12 (Bàn giao Tem Tag, Nhận tem tag mới, Thay tem Tag) '
                'theo mã Store trên biên bản: %(stores)s. Mỗi Cửa hàng trưởng / '
                'bước bàn giao chỉ được xác nhận bởi Quản lý cửa hàng đúng cửa '
                'hàng đó.'
            ) % {
                'users': escape(', '.join(managers.mapped('display_name'))),
                'stores': escape(store_label),
            },
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return managers

    def _ctkm_build_store_verifier_lines(self, store_keys, store_manager_map):
        """(Tái) tạo ctkm.task.store.verifier.

        - Bước có Phụ trách (Cửa hàng trưởng) theo cửa hàng (bước 11/12): mỗi
          chief gắn với Quản lý cửa hàng ĐÚNG cửa hàng của mình.
        - Bước không có Phụ trách riêng (Bàn giao Tem Tag - bước 10): mỗi cửa
          hàng một dòng, Quản lý cửa hàng xác nhận trực tiếp phần cửa hàng đó
          (no_assignee = True).
        """
        self.ensure_one()
        if not store_keys or not store_manager_map:
            return
        StoreVerifier = self.env['ctkm.task.store.verifier'].sudo()
        self.store_verifier_ids.sudo().unlink()
        no_assignee = bool(self.is_tem_handover_task)
        lines = []
        for key in store_keys:
            manager = store_manager_map.get(key)
            if not manager:
                continue
            if no_assignee:
                lines.append({
                    'task_id': self.id,
                    'store_key': key,
                    'store_label': key,
                    'verifier_id': manager.id,
                    'no_assignee': True,
                })
                continue
            chiefs = self._ctkm_users_for_store_key(key)
            if not chiefs:
                continue
            for chief in chiefs:
                lines.append({
                    'task_id': self.id,
                    'store_key': key,
                    'store_label': key,
                    'assignee_user_id': chief.id,
                    'verifier_id': manager.id,
                })
        if lines:
            StoreVerifier.create(lines)

    def _ctkm_tem_tag_line_values(self):
        """Gom kho Tem/Tag của CTKM theo (Mã vật tư, Store) cho bảng chi tiết."""
        self.ensure_one()
        if 'ctkm.inventory.tem.tag' not in self.env:
            return []
        if not self.program_id:
            return []
        if not (
            self.is_tem_tag_import_task
            or self.is_tem_bb_replace_task
            or self.is_tem_handover_task
            or self.is_tem_receive_task
            or self.is_tem_replace_task
        ):
            return []
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            classify_tem_tag_kinds,
        )
        tem_tag = self.env['ctkm.inventory.tem.tag'].sudo()
        domain = [('program_id', '=', self.program_id.id)]
        # Bước 11 / 12: lưu TẤT CẢ cửa hàng của file tổng. Mỗi cửa hàng trưởng
        # chỉ thấy store của mình nhờ ir.rule (không lọc theo người nhận việc đầu).
        if self.is_tem_bb_replace_task:
            # Bước 6 "Lập BB thay tem": chỉ cửa hàng trong "Cửa hàng quản lí"
            # (Nhân viên → Cấu hình) của người nhận việc.
            store_keys = self._ctkm_tem_tag_managed_store_keys()
            if not store_keys:
                return []
            domain = domain + [('store_key', 'in', store_keys)]
        groups = tem_tag._read_group(
            domain,
            ['material_code', 'store', 'store_key', 'tem_tag'],
            ['quantity:sum', 'replaced_quantity:sum', 'date:max', 'ctkm_name:max'],
        )
        by_key = {}
        for (
            material_code, store, store_key, kind_value, quantity,
            replaced_quantity, last_date, ctkm_name,
        ) in groups:
            kinds = classify_tem_tag_kinds(kind_value)
            is_tag = bool(kinds == {'tag'})
            kind = 'tag' if is_tag else 'tem'
            key = (material_code or '', store or '', kind)
            rec = by_key.setdefault(key, {
                'material_code': material_code or False,
                'store': store or False,
                'store_key': store_key or False,
                'date': last_date or False,
                'total_quantity': 0.0,
                'replaced_quantity': 0.0,
                'is_tem': not is_tag,
                'is_tag': is_tag,
                'ctkm_name': ctkm_name or False,
            })
            if store_key and not rec.get('store_key'):
                rec['store_key'] = store_key
            rec['total_quantity'] += quantity or 0.0
            rec['replaced_quantity'] += replaced_quantity or 0.0
            if last_date and (not rec['date'] or last_date > rec['date']):
                rec['date'] = last_date
        values = list(by_key.values())
        values.sort(key=lambda vals: (
            vals['material_code'] or '',
            vals['store'] or '',
            0 if vals['is_tem'] else 1,
        ))
        if self.is_tem_receive_task or self.is_tem_replace_task:
            allowed = self._ctkm_upstream_sent_store_keys()
            if allowed is not None:
                values = [
                    vals for vals in values
                    if self._ctkm_store_in_allowed(
                        allowed, vals.get('store'), vals.get('store_key'),
                    )
                ]
        # store_key trên dòng replace là field computed — chỉ dùng để lọc.
        for vals in values:
            vals.pop('store_key', None)
        return values

    def _ctkm_sync_tem_tag_lines(self):
        """Dựng lại bảng "Chi tiết tem/tag" từ dữ liệu kho Tem/Tag hiện tại."""
        Line = self.env['ctkm.task.tem.tag.replace.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        programs = self.mapped('program_id')
        dispatch_push = self.env.context.get('ctkm_dispatch_push')
        if programs and not dispatch_push:
            print_tasks = self.sudo().search([
                ('program_id', 'in', programs.ids),
                ('is_tem_print_task', '=', True),
            ])
            # Cũng sync các task bước in đang mở (cờ stored có thể lệch tên bước).
            print_tasks |= self.sudo().filtered('is_tem_print_task')
            print_tasks._ctkm_sync_print_store_lines()
            self.sudo().search([
                ('program_id', 'in', programs.ids),
                ('is_tem_handover_task', '=', True),
            ])._ctkm_sync_step10_lines()
            self.sudo().search([
                ('program_id', 'in', programs.ids),
                ('is_tem_postcheck_task', '=', True),
            ])._ctkm_sync_postcheck_lines()
            photo_tasks = self.sudo().search([
                ('program_id', 'in', programs.ids),
                '|',
                ('is_tem_photo_task', '=', True),
                ('is_tem_check_task', '=', True),
            ])
            photo_tasks |= self.sudo().filtered(
                lambda t: t.is_tem_photo_task or t.is_tem_check_task
            )
            photo_tasks._ctkm_sync_tem_photo_lines()
        elif not programs:
            self.sudo().filtered(
                lambda t: t.is_tem_photo_task or t.is_tem_check_task
            )._ctkm_sync_tem_photo_lines()
        for task in self:
            if task.is_tem_receive_task or task.is_tem_replace_task:
                allowed = task._ctkm_upstream_sent_store_keys()
                if allowed is not None and not allowed:
                    continue
            values = task._ctkm_tem_tag_line_values()
            if (
                not values
                and (task.is_tem_receive_task or task.is_tem_replace_task)
                and task.tem_tag_replace_ids
            ):
                continue
            existing = {}
            obsolete = Line.browse()
            for line in task.sudo().tem_tag_replace_ids:
                key = (
                    line.material_code or '',
                    line.store or '',
                    'tag' if line.is_tag and not line.is_tem else 'tem',
                )
                if key in existing:
                    obsolete |= line
                else:
                    existing[key] = line
            to_create = []
            for vals in values:
                key = (
                    vals['material_code'] or '',
                    vals['store'] or '',
                    'tag' if vals.get('is_tag') and not vals.get('is_tem') else 'tem',
                )
                line = existing.pop(key, None)
                if not line:
                    to_create.append(dict(vals, task_id=task.id))
                    continue
                changes = {
                    field: value
                    for field, value in vals.items()
                    if line[field] != value
                }
                if changes:
                    line.with_context(ctkm_tem_tag_line_sync=True).write(changes)
            if to_create:
                Line.create(to_create)
            obsolete |= Line.browse([line.id for line in existing.values()])
            if obsolete:
                obsolete.unlink()
        if programs and not dispatch_push:
            self.sudo()._ctkm_sync_price_lines_for_programs(programs)

    def _ctkm_tem_design_line_values(self):
        """Lấy các Mã vật tư phân biệt từ bảng 'Chi tiết tem/tag' của chương trình.

        Nguồn là ``ctkm.task.tem.tag.replace.line`` của mọi công việc thuộc cùng
        chương trình (chủ yếu từ bước 'Đổ BB thay tem/tag' / 'Lập BB thay tem').
        Trả về danh sách ``{'material_code': <mã>}`` đã sắp xếp, mỗi mã một dòng.
        """
        self.ensure_one()
        if not self.program_id:
            return []
        if 'ctkm.task.tem.tag.replace.line' not in self.env:
            return []
        program_tasks = self.sudo().search([('program_id', '=', self.program_id.id)])
        replace_lines = self.env['ctkm.task.tem.tag.replace.line'].sudo().search(
            [('task_id', 'in', program_tasks.ids)]
        )
        codes = sorted({line.material_code for line in replace_lines if line.material_code})
        return [{'material_code': code} for code in codes]

    def _ctkm_sync_tem_design_lines(self):
        """Dựng lại bảng 'Thiết kế mẫu tem/tag' từ Mã vật tư của chương trình.

        Mỗi Mã vật tư phân biệt từ bảng 'Chi tiết tem/tag' ứng với một dòng.
        Dòng đã có (có thể đã tải file) được giữ lại; dòng không còn trong nguồn
        (và chưa tải file) bị xóa.
        """
        Line = self.env['ctkm.task.tem.design.line'].sudo().with_context(
            ctkm_tem_design_sync=True,
        )
        for task in self:
            if not task.is_tem_design_task:
                continue
            values = task._ctkm_tem_design_line_values()
            value_codes = {vals['material_code'] for vals in values}
            existing = {line.material_code: line for line in task.sudo().tem_design_ids}
            to_create = [
                dict(vals, task_id=task.id)
                for vals in values
                if vals['material_code'] not in existing
            ]
            if to_create:
                Line.create(to_create)
            # Xóa các dòng không còn trong nguồn và chưa tải file (tránh mất file đã up).
            obsolete = Line.browse()
            for code, line in existing.items():
                if code not in value_codes and not line.file:
                    obsolete |= line
            if obsolete:
                obsolete.unlink()

    def action_refresh_tem_design_lines(self):
        """Nút 'Làm mới' bảng Thiết kế mẫu tem/tag trên form bước 7."""
        task = self[:1]
        if not task.is_tem_design_task:
            return self._ctkm_notify_reload(
                _('Không áp dụng'),
                _('Chỉ bước "Thiết kế mẫu tem/tag, Bảng nhận diện" mới có bảng này.'),
            )
        self.sudo()._ctkm_sync_tem_design_lines()
        return self._ctkm_notify_reload(
            _('Đã làm mới'),
            _('Đã cập nhật danh sách Mã vật tư từ bảng "Chi tiết tem/tag" của chương trình.'),
        )

    def action_preview_tem_photos(self):
        """Xem trước ảnh tem/tag đã tải lên của bước Chụp ảnh."""
        self.ensure_one()
        if not self.detail_tem_photo_ids:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Chưa có ảnh'),
                    'message': _('Chưa có hình ảnh nào được tải lên để xem.'),
                    'type': 'warning',
                },
            }
        if len(self.detail_tem_photo_ids) == 1:
            photo = self.detail_tem_photo_ids[0]
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/image/%s?download=0' % photo.id,
                'target': 'new',
            }
        return {
            'name': _('Xem ảnh tem/tag'),
            'type': 'ir.actions.act_window',
            'res_model': 'ir.attachment',
            'view_mode': 'kanban,list,form',
            'domain': [('id', 'in', self.detail_tem_photo_ids.ids)],
            'target': 'new',
            'context': {'create': False, 'edit': False, 'delete': False},
        }

    def _ctkm_print_store_line_values(self):
        """Gom file tổng Tem/Tag theo cửa hàng: cộng SL tem/tag cho bước In tem, Tag.

        Không phụ thuộc hr.store: mỗi cửa hàng là một store_key phân biệt lấy trực
        tiếp từ file tổng, tên hiển thị là ``store`` (tên trên file). Ví dụ file tổng
        có AETL / N21080_0_NAVY = 1 và AETL / N21080_0_YELLOW = 1 → một dòng cửa hàng
        AETL với SL tem = 2.
        """
        self.ensure_one()
        if 'ctkm.inventory.tem.tag' not in self.env:
            return []
        if not self.program_id or not self.is_tem_print_task:
            return []
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            classify_tem_tag_kinds,
            normalize_store_key,
        )
        rows = self.env['ctkm.inventory.tem.tag'].sudo().search([
            ('program_id', '=', self.program_id.id),
        ])
        by_store = {}
        for row in rows:
            store_name = row.store or ''
            store_key = row.store_key or normalize_store_key(store_name)
            if not store_key:
                continue
            rec = by_store.setdefault(store_key, {
                'store': store_name or store_key,
                'store_key': store_key,
                'tem_quantity': 0.0,
                'tag_quantity': 0.0,
            })
            if store_name and rec['store'] in (store_key, '', False):
                rec['store'] = store_name
            amount = row.quantity or 0.0
            kinds = classify_tem_tag_kinds(row.tem_tag)
            if 'tag' in kinds:
                rec['tag_quantity'] += amount
            if 'tem' in kinds or not kinds:
                rec['tem_quantity'] += amount
        values = [
            vals for vals in by_store.values()
            if vals['tem_quantity'] or vals['tag_quantity']
        ]
        values.sort(key=lambda vals: (vals['store'] or '', vals['store_key'] or ''))
        for index, vals in enumerate(values, start=1):
            vals['sequence'] = index
        return values

    def _ctkm_hr_store_map(self, keys):
        """Khớp mã/tên kho với cửa hàng trên Cấu hình nhân viên (hr.store)."""
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            match_hr_store,
            normalize_store_key,
        )
        mapping = {}
        if 'hr.store' not in self.env:
            return mapping
        wanted = {normalize_store_key(key) for key in keys if key}
        if not wanted:
            return mapping
        stores = self.env['hr.store'].sudo().search([])
        for key in wanted:
            store = match_hr_store(stores, key)
            if store:
                mapping[key] = store
        return mapping

    def _ctkm_print_line_alias_keys(self, line=None, vals=None):
        """Các mã cửa hàng dùng để nhận diện cùng một dòng (AETL = LUG_AETL)."""
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            normalize_store_key,
        )
        keys = []

        def add(raw):
            key = normalize_store_key(raw)
            if not key or key in keys:
                return
            keys.append(key)
            if key.startswith('LUG') and len(key) > 3:
                short = key[3:]
                if short and short not in keys:
                    keys.append(short)

        if line is not None:
            add(line.store_key)
            add(line.store)
            add(line.store_code)
            if line.store_id:
                add(line.store_id.code)
                add(line.store_id.name)
        if vals:
            add(vals.get('store_key'))
            add(vals.get('store'))
            add(vals.get('store_code'))
        return keys

    def _ctkm_sync_print_store_lines(self):
        """Dựng lại bảng cửa hàng in tem/tag, giữ nguyên ô tích và dòng thêm tay."""
        Line = self.env['ctkm.task.tem.print.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        for task in self:
            if not task.is_tem_print_task:
                continue
            values = task._ctkm_print_store_line_values()
            existing = {}
            obsolete = Line.browse()
            for line in task.sudo().print_store_ids:
                line_keys = task._ctkm_print_line_alias_keys(line=line)
                if not line_keys:
                    if line.is_manual:
                        continue
                    obsolete |= line
                    continue
                already = any(
                    existing.get(key) and existing.get(key).id == line.id
                    for key in line_keys
                )
                if already:
                    continue
                duplicate = next((existing[key] for key in line_keys if key in existing), None)
                if duplicate:
                    if line.is_manual:
                        continue
                    obsolete |= line
                    continue
                for key in line_keys:
                    existing.setdefault(key, line)
            to_create = []
            for vals in values:
                line = None
                for key in task._ctkm_print_line_alias_keys(vals=vals):
                    line = existing.get(key)
                    if line:
                        break
                if line:
                    for alias, other in list(existing.items()):
                        if other.id == line.id:
                            existing.pop(alias, None)
                    changes = {}
                    for field, value in vals.items():
                        if field in ('is_manual', 'done', 'done_date'):
                            continue
                        if field == 'store_id' and line.store_id:
                            continue
                        if line[field] != value:
                            changes[field] = value
                    if changes:
                        line.write(changes)
                    continue
                to_create.append(dict(vals, task_id=task.id, is_manual=False))
            if to_create:
                Line.create(to_create)
            leftover_ids = list({
                line.id: line for line in existing.values()
            }.values())
            leftover = Line.browse([line.id for line in leftover_ids])
            obsolete |= leftover.filtered(lambda line: not line.is_manual)
            if obsolete:
                obsolete.unlink()
            task._ctkm_renumber_print_store_lines()

    def _ctkm_renumber_print_store_lines(self):
        self.ensure_one()
        lines = self.sudo().print_store_ids.sorted(
            lambda line: (line.sequence or 0, line.store or '', line.id)
        )
        Line = self.env['ctkm.task.tem.print.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        for index, line in enumerate(lines, start=1):
            if line.sequence != index:
                Line.browse(line.id).write({'sequence': index})

    def _ctkm_source_print_task(self):
        """Công việc bước 9 (In tem, Tag) cùng chương trình."""
        self.ensure_one()
        if not self.program_id:
            return self.browse()
        return self.sudo().search([
            ('program_id', '=', self.program_id.id),
            ('is_tem_print_task', '=', True),
        ], order='id desc', limit=1)

    def _ctkm_source_replace_task(self):
        """Công việc bước 12 (Thay tem Tag) cùng chương trình."""
        self.ensure_one()
        if not self.program_id:
            return self.browse()
        return self.sudo().search([
            ('program_id', '=', self.program_id.id),
            ('is_tem_replace_task', '=', True),
        ], order='id desc', limit=1)

    def _ctkm_source_check_task(self):
        """Công việc bước 14 (Kiểm tra hình ảnh tem tag) cùng chương trình."""
        self.ensure_one()
        if not self.program_id:
            return self.browse()
        return self.sudo().search([
            ('program_id', '=', self.program_id.id),
            ('is_tem_check_task', '=', True),
        ], order='id desc', limit=1)

    def _ctkm_source_kt_apply_task(self):
        """Công việc bước 8 (KT áp giá lên phần mềm linkq) cùng chương trình."""
        self.ensure_one()
        if not self.program_id:
            return self.browse()
        stage_id = self._ctkm_step_stage_id('ctkm_core.ctkm_stage_8')
        domain = [('program_id', '=', self.program_id.id)]
        if stage_id:
            found = self.sudo().search(
                domain + [('checklist_line_id.stage_id', '=', stage_id)],
                order='id desc',
                limit=1,
            )
            if found:
                return found
        tasks = self.sudo().search(domain)
        return tasks.filtered(
            lambda task: 'ktapgia' in (normalize_step_key(task.name) or '')
            or 'ktapgia' in (normalize_step_key(task.checklist_step_name) or '')
        )[:1]

    def _ctkm_store_canonical_key(self, *values):
        """Mã cửa hàng để khớp GLLA_AEBD / AEBD, không gộp nhầm LUG_AEHD với LUG_THID."""
        for value in values:
            key = _ctkm_normalize_store_key(value)
            if not key:
                continue
            parent = self._ctkm_grouped_store_key(key)
            if parent:
                return parent
            if '_' in key:
                last = key.rsplit('_', 1)[-1]
                if len(last) >= 3:
                    return last
            return key
        return False

    def _ctkm_store_key_aliases(self, *values):
        """Tập mã cửa hàng để khớp GLLA_AEBD / LUG_AEHD / AEBD."""
        aliases = set()
        for value in values:
            key = _ctkm_normalize_store_key(value)
            if not key:
                continue
            aliases.add(key)
            parent = self._ctkm_grouped_store_key(key)
            if parent:
                aliases.add(parent)
            canon = self._ctkm_store_canonical_key(key)
            if canon:
                aliases.add(canon)
        return aliases

    def _ctkm_build_previous_confirm_index(self):
        """ASM từ cột Đã thay bước 12; KTDT từ SL chưa thay cùng bước."""
        self.ensure_one()
        asm = {}
        remaining = set()
        remaining_info = {}
        replace_task = self._ctkm_source_replace_task()
        if replace_task:
            by_store = {}
            for line in replace_task.sudo().tem_tag_replace_ids:
                canon = self._ctkm_store_canonical_key(line.store_key, line.store)
                if not canon:
                    continue
                by_store.setdefault(canon, []).append(line)
            for canon, lines in by_store.items():
                if lines and all(line.replaced_done for line in lines):
                    src = lines[-1]
                    for line in lines:
                        if line.write_date and (
                            not src.write_date or line.write_date >= src.write_date
                        ):
                            src = line
                    asm[canon] = (
                        src.write_uid.id if src.write_uid else False,
                        src.write_date.date() if src.write_date else False,
                    )
                remain_lines = [
                    line for line in lines
                    if (line.remaining_quantity or 0.0) > 0.000001
                ]
                if remain_lines:
                    remaining.add(canon)
                    src = remain_lines[-1]
                    remaining_info[canon] = (
                        src.write_uid.id if src.write_uid else False,
                        src.write_date.date() if src.write_date else False,
                    )
        return {
            'asm': asm,
            'remaining': remaining,
            'remaining_info': remaining_info,
        }

    def _ctkm_price_confirm_vals(self, index, *store_values):
        """ASM từ Đã thay bước 12; KTDT từ SL chưa thay."""
        canon = self._ctkm_store_canonical_key(*store_values)
        asm_hit = index['asm'].get(canon) if canon else None
        ktdt_hit = index.get('remaining_info', {}).get(canon) if canon else None
        vals = {
            'replaced': bool(canon and canon in index['asm']),
            'not_replaced': bool(canon and canon in index['remaining']),
            'replaced_user_id': asm_hit[0] if asm_hit else False,
            'replaced_date': asm_hit[1] if asm_hit else False,
            'not_replaced_user_id': ktdt_hit[0] if ktdt_hit else False,
            'not_replaced_date': ktdt_hit[1] if ktdt_hit else False,
        }
        return vals

    def _ctkm_print_line_store_key(self, line):
        """Mã cửa hàng dùng để copy bước 9 → 10 (ưu tiên store_key, rồi hr.store)."""
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            normalize_store_key,
        )
        if line.store_key:
            return line.store_key
        if line.store_id:
            return normalize_store_key(line.store_id.code or line.store_id.name)
        return normalize_store_key(line.store)

    def _ctkm_sync_step10_lines(self):
        """Bàn giao: copy tên + SL từ bước 9. Thu: copy mã cửa hàng, giữ SL đã điền."""
        Line = self.env['ctkm.task.tem.step10.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        for task in self:
            if not task.is_tem_handover_task:
                continue
            print_task = task._ctkm_source_print_task()
            sources = print_task.sudo().print_store_ids
            task._ctkm_sync_step10_type(Line, 'handover', sources, copy_qty=True)

    def _ctkm_sync_step10_type(self, Line, line_type, sources, copy_qty):
        self.ensure_one()
        existing = {}
        obsolete = Line.browse()
        current = Line.search([
            ('task_id', '=', self.id),
            ('line_type', '=', line_type),
        ])
        for line in current:
            key = line.store_key or ''
            if not key:
                if line.is_manual:
                    continue
                obsolete |= line
                continue
            if key in existing:
                if line.is_manual:
                    continue
                obsolete |= line
            else:
                existing[key] = line
        to_create = []
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            match_hr_store,
        )
        stores = (
            self.env['hr.store'].sudo().search([])
            if 'hr.store' in self.env else self.env['hr.store']
        )
        for source in sources.sorted(lambda line: (line.sequence, line.id)):
            key = self._ctkm_print_line_store_key(source)
            if not key:
                continue
            store_rec = match_hr_store(stores, source.store_key, source.store)
            store_name = source.store or source.store_code or ''
            if store_rec:
                name = store_rec.name
                if isinstance(name, dict):
                    name = next(iter(name.values()), '') if name else ''
                if name:
                    store_name = name
            vals = {
                'sequence': source.sequence or 1,
                'store_id': store_rec.id if store_rec else False,
                'store': store_name,
                'store_key': key,
                'print_line_id': source.id,
            }
            if copy_qty:
                vals['tem_quantity'] = float(source.tem_quantity or 0.0)
                vals['tag_quantity'] = float(source.tag_quantity or 0.0)
            line = existing.pop(key, None)
            if not line:
                create_vals = dict(
                    vals, task_id=self.id, line_type=line_type, is_manual=False,
                )
                if not copy_qty:
                    create_vals['tem_quantity'] = 0.0
                    create_vals['tag_quantity'] = 0.0
                to_create.append(create_vals)
                continue
            changes = {}
            for field, value in vals.items():
                if line[field] != value:
                    changes[field] = value
            if changes:
                line.write(changes)
        leftover = Line.browse([line.id for line in existing.values()])
        obsolete |= leftover.filtered(lambda line: not line.is_manual)
        if to_create:
            Line.create(to_create)
            if obsolete:
                obsolete.unlink()

    def _ctkm_sync_postcheck_lines(self):
        """Hậu kiểm: copy cửa hàng + SL tem/tag từ bước 9 In tem, Tag."""
        Line = self.env['ctkm.task.tem.postcheck.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        for task in self:
            if not task.is_tem_postcheck_task:
                continue
            print_task = task._ctkm_source_print_task()
            sources = print_task.sudo().print_store_ids
            allowed = task._ctkm_upstream_sent_store_keys()
            if allowed is not None and not allowed:
                continue
            if allowed is not None:
                sources = sources.filtered(
                    lambda line: task._ctkm_store_in_allowed(
                        allowed,
                        task._ctkm_print_line_store_key(line),
                        line.store,
                    )
                )
            existing = {}
            obsolete = Line.browse()
            for line in Line.search([('task_id', '=', task.id)]):
                key = line.store_key or ''
                if not key:
                    obsolete |= line
                    continue
                if key in existing:
                    obsolete |= line
                else:
                    existing[key] = line
            to_create = []
            from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
                match_hr_store,
            )
            stores = (
                self.env['hr.store'].sudo().search([])
                if 'hr.store' in self.env else self.env['hr.store']
            )
            for source in sources.sorted(lambda line: (line.sequence, line.id)):
                key = task._ctkm_print_line_store_key(source)
                if not key:
                    continue
                store_rec = match_hr_store(stores, source.store_key, source.store)
                store_name = source.store or source.store_code or ''
                if store_rec:
                    name = store_rec.name
                    if isinstance(name, dict):
                        name = next(iter(name.values()), '') if name else ''
                    if name:
                        store_name = name
                vals = {
                    'sequence': source.sequence or 1,
                    'store_id': store_rec.id if store_rec else False,
                    'store': store_name,
                    'store_key': key,
                    'print_line_id': source.id,
                }
                line = existing.pop(key, None)
                if not line:
                    to_create.append(dict(vals, task_id=task.id))
                    continue
                changes = {}
                for field, value in vals.items():
                    current = line[field]
                    if line._fields[field].type == 'many2one':
                        current = current.id if current else False
                    if current != value:
                        changes[field] = value
                if changes:
                    line.write(changes)
            leftover = Line.browse([line.id for line in existing.values()])
            obsolete |= leftover
            if to_create:
                Line.create(to_create)
            if obsolete:
                obsolete.unlink()

    def _ctkm_sync_price_lines(self):
        """Kế toán áp giá: copy cửa hàng + SL từ bước 9, xác nhận từ bước 8/12/14."""
        Line = self.env['ctkm.task.tem.price.line'].sudo().with_context(
            ctkm_tem_tag_line_sync=True,
        )
        for task in self:
            if not task.is_tem_price_task:
                continue
            print_task = task._ctkm_source_print_task()
            sources = print_task.sudo().print_store_ids
            allowed = task._ctkm_upstream_sent_store_keys()
            if allowed is not None and not allowed:
                continue
            if allowed is not None:
                sources = sources.filtered(
                    lambda line: task._ctkm_store_in_allowed(
                        allowed,
                        task._ctkm_print_line_store_key(line),
                        line.store,
                    )
                )
            confirm_index = task._ctkm_build_previous_confirm_index()
            existing = {}
            obsolete = Line.browse()
            for line in Line.search([('task_id', '=', task.id)]):
                key = line.store_key or ''
                if not key:
                    obsolete |= line
                    continue
                if key in existing:
                    obsolete |= line
                else:
                    existing[key] = line
            to_create = []
            from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
                match_hr_store,
            )
            stores = (
                self.env['hr.store'].sudo().search([])
                if 'hr.store' in self.env else self.env['hr.store']
            )
            for source in sources.sorted(lambda line: (line.sequence, line.id)):
                key = task._ctkm_print_line_store_key(source)
                if not key:
                    continue
                store_rec = match_hr_store(stores, source.store_key, source.store)
                store_name = source.store or source.store_code or ''
                if store_rec:
                    name = store_rec.name
                    if isinstance(name, dict):
                        name = next(iter(name.values()), '') if name else ''
                    if name:
                        store_name = name
                vals = {
                    'sequence': source.sequence or 1,
                    'store_id': store_rec.id if store_rec else False,
                    'store': store_name,
                    'store_key': key,
                    'print_line_id': source.id,
                }
                vals.update(task._ctkm_price_confirm_vals(
                    confirm_index, key, source.store, source.store_code,
                    store_rec.code if store_rec else False,
                    store_name,
                ))
                can_apply = bool(vals.get('replaced') and vals.get('not_replaced'))
                if not can_apply:
                    vals['price_applied'] = False
                    vals['price_applied_user_id'] = False
                    vals['price_applied_date'] = False
                line = existing.pop(key, None)
                if not line:
                    to_create.append(dict(vals, task_id=task.id))
                    continue
                if can_apply:
                    vals.pop('price_applied', None)
                    vals.pop('price_applied_user_id', None)
                    vals.pop('price_applied_date', None)
                changes = {}
                for field, value in vals.items():
                    current = line[field]
                    ftype = line._fields[field].type
                    if ftype == 'many2one':
                        current = current.id if current else False
                        if isinstance(value, models.BaseModel):
                            value = value.id if value else False
                    if current != value:
                        changes[field] = value
                if changes:
                    line.write(changes)
            leftover = Line.browse([line.id for line in existing.values()])
            obsolete |= leftover
            if to_create:
                Line.create(to_create)
            if obsolete:
                obsolete.unlink()

    @api.model
    def _ctkm_sync_price_lines_for_programs(self, programs):
        if self.env.context.get('ctkm_price_line_syncing'):
            return
        programs = programs.exists() if programs else programs
        if not programs:
            return
        self.sudo().with_context(ctkm_price_line_syncing=True).search([
            ('program_id', 'in', programs.ids),
            ('is_tem_price_task', '=', True),
        ])._ctkm_sync_price_lines()

    def _ctkm_is_photo_or_check_task(self):
        return bool(self.is_tem_photo_task or self.is_tem_check_task)

    def _ctkm_tem_photo_line_values(self):
        """Gom kho Tem/Tag (file tổng) theo (Cửa hàng, Mã vật tư).

        Dùng cho bước 13 (Đã chụp) và bước 14 (Xác nhận thay). Nguồn là
        ``ctkm.inventory.tem.tag`` của chương trình — bước
        "Đổ BB thay tem/tag (file tổng)" chứa tất cả cửa hàng.
        """
        self.ensure_one()
        if not self._ctkm_is_photo_or_check_task() or not self.program_id:
            return []
        if 'ctkm.inventory.tem.tag' not in self.env:
            return []
        rows = self.env['ctkm.inventory.tem.tag'].sudo().search([
            ('program_id', '=', self.program_id.id),
        ])
        by_key = {}
        for rec in rows:
            key = (rec.store or '', rec.material_code or '')
            if not key[0] or not key[1]:
                continue
            if key not in by_key:
                by_key[key] = {
                    'store': rec.store or '',
                    'store_key': rec.store_key or '',
                    'material_code': rec.material_code or '',
                    'total_quantity': 0.0,
                }
            by_key[key]['total_quantity'] += rec.quantity or 0.0
        values = list(by_key.values())
        values.sort(key=lambda vals: (vals['store'] or '', vals['material_code'] or ''))
        allowed = self._ctkm_upstream_sent_store_keys()
        if allowed is not None:
            values = [
                vals for vals in values
                if self._ctkm_store_in_allowed(
                    allowed, vals.get('store_key'), vals.get('store'),
                )
            ]
        return values

    def _ctkm_sync_tem_photo_lines(self):
        """Dựng lại bảng chụp / kiểm tra ảnh tem-tag từ kho Tem/Tag (file tổng)."""
        Line = self.env['ctkm.task.tem.photo.line'].sudo().with_context(
            ctkm_tem_photo_sync=True,
        )
        for task in self:
            if not task._ctkm_is_photo_or_check_task():
                continue
            allowed = task._ctkm_upstream_sent_store_keys()
            if allowed is not None and not allowed:
                continue
            values = task._ctkm_tem_photo_line_values()
            existing = {}
            for line in task.sudo().tem_photo_check_ids:
                existing[(line.store or '', line.material_code or '')] = line
            to_create = []
            for vals in values:
                key = (vals['store'] or '', vals['material_code'] or '')
                line = existing.pop(key, None)
                if line:
                    changes = {
                        field: value
                        for field, value in vals.items()
                        if line[field] != value
                    }
                    if changes:
                        line.write(changes)
                else:
                    to_create.append(dict(vals, task_id=task.id))
            if to_create:
                Line.create(to_create)
            removed = Line.browse([line.id for line in existing.values()])
            if removed:
                removed.unlink()
        self.sudo()._ctkm_sync_price_lines_for_programs(self.mapped('program_id'))

    def action_add_collect_store_line(self):
        """Nút Tạo dòng trên bảng Thu tem/tag."""
        self.ensure_one()
        if not self.is_tem_handover_task:
            raise UserError(_(
                'Chỉ bước bàn giao / thu hồi tem tag mới được thêm dòng thu.'
            ))
        is_ctkm_manager = self.env.user.has_group('ctkm_core.group_ctkm_manager')
        if not is_ctkm_manager and self.env.user not in self.user_ids:
            raise UserError(_('Chỉ người nhận việc mới được thêm dòng thu tem/tag.'))
        sequence = max(self.collect_store_ids.mapped('sequence') or [0]) + 1
        self.env['ctkm.task.tem.step10.line'].create({
            'task_id': self.id,
            'line_type': 'collect',
            'sequence': sequence,
            'is_manual': True,
        })
        return self._ctkm_notify_reload(
            _('Đã tạo dòng'),
            _('Chọn cửa hàng trên dòng mới; điền SL tem / SL tag thu về.'),
        )

    def action_add_print_store_line(self):
        """Nút Tạo dòng: thêm một dòng trống để chọn cửa hàng."""
        self.ensure_one()
        if not self.is_tem_print_task:
            raise UserError(_('Chỉ bước "In tem, Tag" mới được thêm cửa hàng in.'))
        is_ctkm_manager = self.env.user.has_group('ctkm_core.group_ctkm_manager')
        if not is_ctkm_manager and self.env.user not in self.user_ids:
            raise UserError(_(
                'Chỉ người nhận việc mới được thêm cửa hàng in tem/tag.'
            ))
        sequence = max(self.print_store_ids.mapped('sequence') or [0]) + 1
        self.env['ctkm.task.tem.print.line'].create({
            'task_id': self.id,
            'sequence': sequence,
            'is_manual': True,
        })
        return self._ctkm_notify_reload(
            _('Đã tạo dòng'),
            _('Chọn cửa hàng trên dòng mới (Cấu hình nhân viên → Cửa hàng).'),
        )

    def action_export_thu_tem_file(self):
        """Xuất biên bản thu tem/tag (bảng Thu tem/tag) theo mẫu thutem.xlsx."""
        self.ensure_one()
        if not self.is_tem_handover_task:
            raise UserError(_(
                'Chỉ công việc bước "Bàn giao Tem Tag / Thu hồi tem tag cũ" '
                'mới được xuất biên bản thu tem.'
            ))
        if Workbook is None:
            raise UserError(_('Thiếu thư viện openpyxl để xuất file Excel.'))
        data = self._build_thu_tem_xlsx()
        date_str = fields.Date.context_today(self).strftime('%d_%m_%Y')
        filename = 'Bien_ban_thu_tem_%s_%s.xlsx' % (
            self.program_id.notify_code or self.program_id.id or 'CTKM',
            date_str,
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(data),
            'res_model': 'ctkm.task',
            'res_id': self.id,
            'type': 'binary',
            'public': True,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    # Bố cục sheet TH (khớp mẫu thutem.xlsx, sheet TH).
    TH_PRICE_FORMAT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
    TH_QTY_FORMAT = '#,##0'

    def _build_thu_tem_xlsx(self):
        """Xuất bảng Thu tem/tag ra sheet TH giống hệt mẫu thutem.xlsx.

        Bố cục (dựng hoàn toàn bằng openpyxl, không đọc file mẫu):
            Dòng 1: "CÔNG TY TNHH SÁNG TÂM" (bold 10)
            Dòng 2: địa chỉ công ty (bold 10)
            Dòng 3: "Tổng hợp nhập xuất tồn theo cột" (bold 18, height 23.4)
            Dòng 4: "Ngày dd/mm/yyyy" (bold 10)
            Dòng 6: tiêu đề cột: Mã vật tư | Giá KM | Ghi chú | Tem/tag |
                    <các mã cửa hàng>... | Tổng cộng | Mã gốc
                    (bold, căn giữa, viền thin, height 108.6)
            Dòng 7+: mỗi Mã vật tư một dòng; ô cửa hàng = SL tem + SL tag,
                    cột Tổng cộng = SUM(E..cột CH cuối), cột Mã gốc = phần
                    đầu của Mã vật tư (trước dấu "_").
            Dòng cuối: "Tổng cộng:" + SUM từng cột cửa hàng.
        Freeze panes tại E7.
        """
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = 'TH'

        thin = Side(style='thin', color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        bold10 = Font(bold=True, size=10)
        bold11 = Font(bold=True, size=11)
        title_font = Font(bold=True, size=18)
        center = Alignment(horizontal='center', vertical='center')

        today = fields.Date.context_today(self)

        ws['A1'] = 'CÔNG TY TNHH SÁNG TÂM'
        ws['A1'].font = bold10
        ws['A2'] = '32-34 đường 74, Phường 10, Quận 6, TP. Hồ Chí Minh'
        ws['A2'].font = bold10
        ws['A3'] = 'Tổng hợp nhập xuất tồn theo cột'
        ws['A3'].font = title_font
        ws.row_dimensions[3].height = 23.4
        ws['A4'] = 'Ngày %s' % today.strftime('%d/%m/%Y')
        ws['A4'].font = bold10

        collect = self.collect_store_ids.sudo()
        materials = sorted({line.material_code for line in collect if line.material_code})
        stores = sorted(
            {line.store_key or '' for line in collect if line.store_key},
        )
        grid = {}
        kinds = {}  # material_code -> set('tem'/'tag') có số lượng > 0
        for line in collect:
            mat = line.material_code or ''
            key = (mat, line.store_key or '')
            qty = ((line.tem_quantity or 0.0) + (line.tag_quantity or 0.0))
            if qty:
                grid[key] = grid.get(key, 0.0) + qty
                if (line.tem_quantity or 0.0) > 0:
                    kinds.setdefault(mat, set()).add('tem')
                if (line.tag_quantity or 0.0) > 0:
                    kinds.setdefault(mat, set()).add('tag')

        # Giá KM + Ghi chú lấy từ kho Tem/Tag (import bước 4) theo Mã vật tư.
        prices, notes = {}, {}
        program = self.program_id
        if program and 'ctkm.inventory.tem.tag' in self.env:
            Inventory = self.env['ctkm.inventory.tem.tag'].sudo()
            for rec in Inventory.search([('program_id', '=', program.id)]):
                code = rec.material_code or ''
                if not code:
                    continue
                if rec.promo_price:
                    prices.setdefault(code, rec.promo_price)
                note = (rec.ctkm_name or '').strip()
                if note:
                    notes.setdefault(code, note)

        # Cột: A Mã vật tư | B Giá KM | C Ghi chú | D Tem/tag |
        #      E.. stores | Tổng cộng | Mã gốc
        store_start_col = 5
        total_col = store_start_col + len(stores)
        origin_col = total_col + 1

        header_row = 6
        headers = ['Mã vật tư', 'Giá KM', 'Ghi chú', 'Tem/tag']
        headers += list(stores) + ['Tổng cộng', 'Mã gốc']
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = bold11
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[header_row].height = 108.6

        last_store_col_letter = get_column_letter(total_col - 1)
        r = header_row + 1
        for mat in materials:
            kind = kinds.get(mat) or set()
            label = (
                'TEM+TAG' if kind == {'tem', 'tag'}
                else 'TEM' if kind == {'tem'}
                else 'TAG' if kind == {'tag'}
                else ''
            )
            row_vals = [
                mat, prices.get(mat), notes.get(mat, ''), label,
            ] + [grid.get((mat, skey)) for skey in stores]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.border = border
                if col_idx == 2:
                    cell.number_format = self.TH_PRICE_FORMAT
                elif col_idx >= store_start_col:
                    cell.number_format = self.TH_QTY_FORMAT
            # Cột "Tổng cộng": công thức SUM như mẫu.
            tc = ws.cell(
                row=r, column=total_col,
                value='=SUM(%s%d:%s%d)' % (
                    get_column_letter(store_start_col), r,
                    last_store_col_letter, r,
                ),
            )
            tc.font = bold11
            tc.border = border
            tc.number_format = self.TH_QTY_FORMAT
            # Cột "Mã gốc": phần trước dấu "_" của Mã vật tư.
            oc = ws.cell(row=r, column=origin_col, value=mat.split('_')[0])
            oc.border = border
            r += 1

        # Dòng cộng dồn cuối bảng.
        first_data_row = header_row + 1
        last_data_row = r - 1
        trow = ws.cell(row=r, column=1, value='Tổng cộng:')
        trow.font = bold11
        trow.border = border
        for col_idx in range(2, origin_col + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            if col_idx >= store_start_col:
                letter = get_column_letter(col_idx)
                cell.value = '=SUM(%s%d:%s%d)' % (
                    letter, first_data_row, letter, last_data_row,
                )
                cell.font = bold11
                cell.number_format = self.TH_QTY_FORMAT

        # Độ rộng cột & freeze panes đúng mẫu.
        widths = {'A': 29.33, 'B': 13.44, 'C': 80.66, 'D': 15.55}
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        for col_idx in range(store_start_col, origin_col - 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8.0
        ws.column_dimensions[get_column_letter(origin_col)].width = 25.33
        ws.freeze_panes = 'E7'

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    # Biên bản in và bàn giao tem (bước 10) — khớp mẫu giấy in.
    HANDOVER_BB_MIN_ROWS = 15
    HANDOVER_BB_LAST_COL = 8
    HANDOVER_BB_NOTES = (
        '* Lưu ý: Sau khi thay tem theo biên bản này:\n'
        '- Tem cũ sau khi thay phải giữ lại đầy đủ, nộp về theo đúng số lượng trên biên bản.\n'
        '- Kiểm tra ngay tem thiếu / tem sai thông tin; báo về công ty nếu có sai lệch.\n'
        '- Nhân viên bán hàng nào làm mất tem sẽ bị phạt 50.000đ/1 con tem.\n'
        '- Cửa hàng trưởng chịu trách nhiệm kiểm tra số lượng tem giao nhận trước khi ký.\n'
        '- Kiểm tra kỹ thông tin trên tem (mã, giá) trước khi dán.\n'
        '- Chụp ảnh tem đã thay xong gửi về công ty.'
    )
    _HANDOVER_BB_INVALID_SHEET = (':', '\\', '/', '?', '*', '[', ']')

    def action_export_handover_bb_file(self):
        """Xuất biên bản in và bàn giao tem theo mẫu giấy, mỗi cửa hàng một sheet."""
        self.ensure_one()
        if not self.is_tem_handover_task:
            raise UserError(_(
                'Chỉ công việc bước "Bàn giao Tem Tag / Thu hồi tem tag cũ" '
                'mới được xuất biên bản bàn giao.'
            ))
        if Workbook is None:
            raise UserError(_('Thiếu thư viện openpyxl để xuất file Excel.'))
        if not self.handover_store_ids:
            raise UserError(_(
                'Chưa có cửa hàng trên bảng Bàn giao tem/tag. '
                'Bấm Làm mới danh sách trước khi xuất biên bản.'
            ))
        data = self._build_handover_bb_xlsx()
        date_str = fields.Date.context_today(self).strftime('%d_%m_%Y')
        filename = 'Bien_ban_ban_giao_tem_%s_%s.xlsx' % (
            self.program_id.notify_code or self.program_id.id or 'CTKM',
            date_str,
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(data),
            'res_model': 'ctkm.task',
            'res_id': self.id,
            'type': 'binary',
            'public': True,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def _handover_bb_sheet_title(self, store_code):
        title = (store_code or 'CH').strip()
        for ch in self._HANDOVER_BB_INVALID_SHEET:
            title = title.replace(ch, '_')
        return title[:31] or 'CH'

    def _handover_bb_norm_keys(self, store_key):
        from odoo.addons.ctkm_core.models.ctkm_task_tem_print_line import (
            normalize_store_key,
        )
        key = normalize_store_key(store_key) or ''
        keys = set()
        if key:
            keys.add(key)
            stripped = key
            if stripped.startswith('LUG_'):
                stripped = stripped[4:]
            elif stripped.startswith('LUG'):
                stripped = stripped[3:]
            if stripped:
                keys.add(stripped)
                keys.add('LUG' + stripped)
                keys.add('LUG_' + stripped)
        return keys

    def _handover_bb_keys_overlap(self, key_a, key_b):
        keys_a = self._handover_bb_norm_keys(key_a)
        keys_b = self._handover_bb_norm_keys(key_b)
        if keys_a & keys_b:
            return True
        for left in keys_a:
            for right in keys_b:
                if len(left) >= 3 and len(right) >= 3 and (
                    left.endswith(right) or right.endswith(left)
                ):
                    return True
        return False

    def _handover_bb_store_label(self, line):
        code = line.store_id.code if line.store_id else False
        if isinstance(code, dict):
            code = next(iter(code.values()), '') if code else ''
        return (code or line.store_code or line.store_key or line.store or 'CH').strip()

    def _handover_bb_materials_for_store(self, store_key, inv_records):
        """Gom mã vật tư giao mới (kho bước 4) và SL thu (bảng Thu tem/tag)."""
        materials = {}
        work_contents = []
        for rec in inv_records:
            if not self._handover_bb_keys_overlap(store_key, rec.store_key or rec.store):
                continue
            code = rec.material_code
            if not code:
                continue
            mat = materials.setdefault(code, {
                'promo': rec.promo_price or 0.0,
                'qty': 0.0,
                'collect_qty': 0.0,
            })
            mat['qty'] += rec.quantity or 0.0
            if rec.promo_price and not mat['promo']:
                mat['promo'] = rec.promo_price
            work_content = (rec.bb_work_content or '').strip()
            if work_content and work_content not in work_contents:
                work_contents.append(work_content)
        for line in self.collect_store_ids.sudo():
            if not self._handover_bb_keys_overlap(store_key, line.store_key or line.store):
                continue
            code = (line.material_code or '').strip()
            if not code:
                continue
            qty = (line.tem_quantity or 0.0) + (line.tag_quantity or 0.0)
            mat = materials.setdefault(code, {
                'promo': 0.0,
                'qty': 0.0,
                'collect_qty': 0.0,
            })
            mat['collect_qty'] += qty
        return materials, work_contents[:1]

    def _build_handover_bb_xlsx(self):
        """Mỗi cửa hàng trên bảng Bàn giao tem/tag → một sheet biên bản."""
        self.ensure_one()
        program = self.program_id
        today = fields.Date.context_today(self)
        date_vn = 'Hôm nay, ngày %02d tháng %02d năm %s' % (
            today.day, today.month, today.year,
        )
        date_str = today.strftime('%d/%m/%Y')
        notify_code = (program.notify_code or '').strip()

        store_list = []
        seen = set()
        for line in self.handover_store_ids.sudo().sorted(
            lambda rec: (rec.sequence, rec.store or '', rec.id)
        ):
            key = line.store_key or ''
            if not key or key in seen:
                continue
            seen.add(key)
            store_list.append((key, self._handover_bb_store_label(line)))
        if not store_list:
            raise UserError(_(
                'Không có cửa hàng hợp lệ trên bảng Bàn giao tem/tag để xuất biên bản.'
            ))

        inv_records = self.env['ctkm.inventory.tem.tag']
        if program and 'ctkm.inventory.tem.tag' in self.env:
            inv_records = self.env['ctkm.inventory.tem.tag'].sudo().search([
                ('program_id', '=', program.id),
            ])

        wb = Workbook()
        used_titles = set()
        first = True
        for store_key, store_code in store_list:
            materials, work_contents = self._handover_bb_materials_for_store(
                store_key, inv_records,
            )
            ws = wb.active if first else wb.create_sheet()
            first = False
            base_title = self._handover_bb_sheet_title(store_code)
            sheet_title = base_title
            suffix = 1
            while sheet_title in used_titles:
                suffix += 1
                extra = '_%s' % suffix
                sheet_title = '%s%s' % (base_title[:31 - len(extra)], extra)
            used_titles.add(sheet_title)
            ws.title = sheet_title
            self._build_handover_bb_sheet(
                ws, store_code, materials, work_contents,
                date_vn, date_str, notify_code,
            )

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _build_handover_bb_sheet(
        self, ws, store_code, materials, work_contents,
        date_vn, date_str, notify_code,
    ):
        """Vẽ một sheet biên bản in/bàn giao tem — bố cục giống mẫu giấy."""
        last_col = self.HANDOVER_BB_LAST_COL
        last_letter = get_column_letter(last_col)
        thin = Side(style='thin', color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_name = 'Times New Roman'
        bold = Font(name=font_name, bold=True, size=11)
        motto_font = Font(name=font_name, bold=True, size=12)
        motto_line_font = Font(name=font_name, bold=True, size=12, underline='single')
        title_font = Font(name=font_name, bold=True, size=16)
        cell_font = Font(name=font_name, size=11)
        note_font = Font(name=font_name, size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
        left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)

        def merge_value(start, end, value, font, align, height=None):
            ws.merge_cells('%s:%s' % (start, end))
            cell = ws[start]
            cell.value = value
            cell.font = font
            cell.alignment = align
            if height:
                ws.row_dimensions[cell.row].height = height

        merge_value('A1', '%s1' % last_letter, 'Cộng Hòa Xã Hội Chủ Nghĩa Việt Nam', motto_font, center)
        merge_value(
            'A2', '%s2' % last_letter,
            'Độc lập - Tự do - Hạnh phúc', motto_line_font, center,
        )
        merge_value(
            'A3', '%s3' % last_letter,
            'BIÊN BẢN IN VÀ BÀN GIAO TEM', title_font, center, 24,
        )
        merge_value('A4', '%s4' % last_letter, 'Số: ..............', bold, center)
        ws['A5'] = date_vn
        ws['A5'].font = cell_font
        ws['A6'] = 'TẠI CH: %s' % store_code
        ws['A6'].font = cell_font
        work_content = work_contents[0] if work_contents else ''
        if not work_content:
            work_content = 'IN DÁN BÀN GIAO TEM THEO TB %s NGÀY %s' % (
                notify_code or '........', date_str,
            )
        merge_value(
            'A7', '%s7' % last_letter,
            'Nội dung công việc: %s' % work_content,
            cell_font, left_center, 28,
        )

        group_row = 8
        header_row = 9
        merge_value('A8', 'D8', 'GIAO TEM MỚI', bold, center)
        merge_value('E8', 'H8', 'TEM THU HỒI', bold, center)
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=group_row, column=col_idx)
            cell.font = bold
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[group_row].height = 22

        headers = [
            'STT', 'MÃ VẬT TƯ', 'GIÁ', 'SỐ LƯỢNG',
            'MÃ VẬT TƯ', 'SL THU', 'SL MẤT', 'GHI CHÚ',
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = bold
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[header_row].height = 24

        codes = sorted(
            code for code, mat in materials.items()
            if (mat.get('qty') or 0) > 0 or (mat.get('collect_qty') or 0) > 0
        )
        data_rows = max(len(codes), self.HANDOVER_BB_MIN_ROWS)
        first_data = header_row + 1
        last_data = first_data + data_rows - 1

        def qty_value(value):
            if not value:
                return None
            if abs(value - round(value)) < 0.0001:
                return int(round(value))
            return value

        for offset in range(data_rows):
            row = first_data + offset
            code = codes[offset] if offset < len(codes) else ''
            mat = materials.get(code) or {}
            ws.cell(row=row, column=1, value=offset + 1)
            ws.cell(row=row, column=2, value=code or None)
            price = mat.get('promo') or None
            price_cell = ws.cell(row=row, column=3, value=price or None)
            price_cell.number_format = '#,##0'
            qty_cell = ws.cell(row=row, column=4, value=qty_value(mat.get('qty')))
            qty_cell.number_format = '#,##0'
            ws.cell(row=row, column=5, value=code or None)
            collect_cell = ws.cell(
                row=row, column=6, value=qty_value(mat.get('collect_qty')),
            )
            collect_cell.number_format = '#,##0'
            ws.cell(row=row, column=7, value=None)
            ws.cell(row=row, column=8, value=None)
            for col_idx in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = cell_font
                cell.border = border
                if col_idx in (1, 3, 4, 6, 7):
                    cell.alignment = center
                else:
                    cell.alignment = left_center
            ws.row_dimensions[row].height = 20

        total_row = last_data + 1
        ws.merge_cells('A%s:C%s' % (total_row, total_row))
        total_label = ws['A%s' % total_row]
        total_label.value = 'TỔNG CỘNG'
        total_label.font = bold
        total_label.alignment = center
        qty_sum = ws.cell(
            row=total_row, column=4,
            value='=SUM(D%s:D%s)' % (first_data, last_data),
        )
        qty_sum.font = bold
        qty_sum.number_format = '#,##0'
        collect_sum = ws.cell(
            row=total_row, column=6,
            value='=SUM(F%s:F%s)' % (first_data, last_data),
        )
        collect_sum.font = bold
        collect_sum.number_format = '#,##0'
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font = bold
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[total_row].height = 22

        note_row = total_row + 2
        merge_value(
            'A%s' % note_row, '%s%s' % (last_letter, note_row),
            self.HANDOVER_BB_NOTES, note_font, left_top, 96,
        )

        sig_title_row = note_row + 2
        sig_title_row2 = sig_title_row + 3
        sig_space_row2 = sig_title_row2 + 1
        groups = [
            (sig_title_row, ['Người lập', 'Giám sát', 'Người in tem']),
            (sig_title_row2, ['Người giao', 'Người nhận', 'Kiểm soát']),
        ]
        spans = [('A', 'C'), ('D', 'E'), ('F', 'H')]
        for title_row, labels in groups:
            space_row = title_row + 1
            for (start, end), label in zip(spans, labels):
                merge_value(
                    '%s%s' % (start, title_row),
                    '%s%s' % (end, title_row),
                    label, bold, center,
                )
                ws.merge_cells('%s%s:%s%s' % (start, space_row, end, space_row))
            ws.row_dimensions[title_row].height = 18
            ws.row_dimensions[space_row].height = 48

        widths = {1: 6, 2: 26, 3: 12, 4: 11, 5: 26, 6: 10, 7: 10, 8: 16}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.sheet_view.showGridLines = False
        ws.print_area = 'A1:%s%s' % (last_letter, sig_space_row2)
        if PageMargins is not None:
            ws.page_margins = PageMargins(
                left=0.4, right=0.4, top=0.5, bottom=0.5,
                header=0.2, footer=0.2,
            )

    @api.model
    def _ctkm_sync_tem_tag_lines_for_programs(self, program_ids):
        """Cập nhật bảng chi tiết của mọi công việc bước 4 / 6 / 9–12 của CTKM."""
        program_ids = [program_id for program_id in (program_ids or []) if program_id]
        if not program_ids:
            return self.browse()
        tasks = self.sudo().search([
            ('program_id', 'in', program_ids),
            '|', '|', '|', '|', '|', '|', '|', '|', '|',
            ('is_tem_tag_import_task', '=', True),
            ('is_tem_bb_replace_task', '=', True),
            ('is_tem_print_task', '=', True),
            ('is_tem_handover_task', '=', True),
            ('is_tem_receive_task', '=', True),
            ('is_tem_replace_task', '=', True),
            ('is_tem_photo_task', '=', True),
            ('is_tem_check_task', '=', True),
            ('is_tem_postcheck_task', '=', True),
            ('is_tem_price_task', '=', True),
        ])
        tasks._ctkm_sync_tem_tag_lines()
        tasks.filtered(
            lambda t: t.is_tem_photo_task or t.is_tem_check_task
        )._ctkm_sync_tem_photo_lines()
        tasks.filtered('is_tem_postcheck_task')._ctkm_sync_postcheck_lines()
        tasks.filtered('is_tem_price_task')._ctkm_sync_price_lines()
        return tasks

    def _ctkm_clear_program_tem_tag_inventory(self):
        """Xóa kho Tem/Tag của CTKM này (dữ liệu import file tổng)."""
        programs = self.mapped('program_id')
        if not programs or 'ctkm.inventory.tem.tag' not in self.env:
            return self.env['ctkm.inventory.tem.tag']
        return self.env['ctkm.inventory.tem.tag'].sudo().search([
            ('program_id', 'in', programs.ids),
        ]).unlink()

    def action_refresh_tem_tag_lines(self):
        """Nút 'Làm mới' bảng Chi tiết tem/tag trên form công việc."""
        task = self[:1]
        if task.is_tem_tag_import_task:
            # Bảng này chỉ phản ánh kho đã import. Dữ liệu cũ (sheet ẩn, lần import
            # trước) phải xóa khỏi kho thì danh sách mới trống / đúng file mới.
            task._ctkm_clear_program_tem_tag_inventory()
        self.sudo()._ctkm_sync_tem_tag_lines()
        if task.is_tem_tag_import_task:
            return self._ctkm_notify_reload(
                _('Đã xóa dữ liệu cũ'),
                _('Đã xóa tem/tag đã import của chương trình này. '
                  'Nhấn Import Tem/Tag để tải lại file Excel (sheet TEM và TAG).'),
            )
        if task.is_tem_print_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('Đã gom số lượng tem/tag theo cửa hàng từ file tổng.'),
            )
        if task.is_tem_handover_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('Đã lấy danh sách cửa hàng từ bước In tem, Tag.'),
            )
        if task.is_tem_postcheck_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('Đã lấy danh sách cửa hàng từ bước In tem, Tag.'),
            )
        if task.is_tem_photo_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('Đã lấy danh sách cửa hàng / mã vật tư từ file tổng để tick Đã chụp.'),
            )
        if task.is_tem_check_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('Đã lấy danh sách cửa hàng / mã vật tư từ file tổng để xác nhận ảnh.'),
            )
        if task.is_tem_price_task:
            return self._ctkm_notify_reload(
                _('Đã làm mới'),
                _('ASM xác nhận lấy từ cột Đã thay bước Thay tem Tag. '
                  'KT áp giá chỉ tick khi đã có đủ ASM xác nhận và KTDT xác nhận.'),
            )
        return self._ctkm_notify_reload(
            _('Đã làm mới'),
            _('Bảng Chi tiết tem/tag đã cập nhật theo dữ liệu kho Tem/Tag.'),
        )

    def action_tick_all_print_done(self):
        """Tick Đã in cho mọi cửa hàng đang ở bảng cần in (theo ô tìm kiếm nếu có)."""
        self.ensure_one()
        if not self.is_tem_print_task:
            raise UserError(_('Chỉ bước "In tem, Tag" mới dùng nút này.'))
        lines = self.print_store_pending_ids
        search = (self.print_store_search or '').strip()
        if search:
            key = search.lower()
            lines = lines.filtered(
                lambda line: key in (line.store or '').lower()
            )
        if not lines:
            return self._ctkm_notify_reload(
                _('Không có dòng'),
                _('Không còn cửa hàng nào để tick Đã in.'),
                notif_type='warning',
            )
        lines.write({'done': True})
        return self._ctkm_notify_reload(
            _('Đã tick tất cả'),
            _('Đã đánh dấu Đã in cho %s cửa hàng.') % len(lines),
        )

    def web_read(self, specification):
        # Mở form bước 9: gom SL tem/tag theo cửa hàng từ file tổng.
        if self.ids and not self.env.context.get('ctkm_skip_print_autosync'):
            print_tasks = self.filtered('is_tem_print_task')
            if print_tasks:
                print_tasks.sudo().with_context(
                    ctkm_skip_print_autosync=True,
                )._ctkm_sync_print_store_lines()
        # Mở form bước 10: copy ngay từ bước 9 (không cần bấm Làm mới).
        if self.ids and not self.env.context.get('ctkm_skip_step10_autosync'):
            handover = self.filtered('is_tem_handover_task')
            if handover:
                handover.sudo().with_context(
                    ctkm_skip_step10_autosync=True,
                )._ctkm_sync_step10_lines()
        # Mở form bước Hậu kiểm: copy cửa hàng + SL từ bước 9.
        if self.ids and not self.env.context.get('ctkm_skip_postcheck_autosync'):
            postcheck = self.filtered('is_tem_postcheck_task')
            if postcheck:
                postcheck.sudo().with_context(
                    ctkm_skip_postcheck_autosync=True,
                )._ctkm_sync_postcheck_lines()
        # Mở form bước 15 Kế toán áp giá: copy cửa hàng + SL từ bước 9.
        if self.ids and not self.env.context.get('ctkm_skip_price_autosync'):
            price_tasks = self.filtered('is_tem_price_task')
            if price_tasks:
                price_tasks.sudo().with_context(
                    ctkm_skip_price_autosync=True,
                )._ctkm_sync_price_lines()
        # Mở form bước 7: dựng bảng thiết kế mẫu tem/tag từ "Chi tiết tem/tag".
        if self.ids and not self.env.context.get('ctkm_skip_tem_design_sync'):
            design_tasks = self.filtered('is_tem_design_task')
            if design_tasks:
                design_tasks.sudo().with_context(
                    ctkm_skip_tem_design_sync=True,
                )._ctkm_sync_tem_design_lines()
        # Mở form bước 13/14: đồng bộ bảng (Cửa hàng, Mã vật tư) từ file tổng.
        if self.ids and not self.env.context.get('ctkm_skip_tem_photo_sync'):
            photo_tasks = self.filtered(
                lambda t: t.is_tem_photo_task or t.is_tem_check_task
            )
            if photo_tasks:
                photo_tasks.sudo().with_context(
                    ctkm_skip_tem_photo_sync=True,
                )._ctkm_sync_tem_photo_lines()
        if self.ids and not self.env.context.get('ctkm_skip_time_line_ensure'):
            self.sudo().with_context(
                ctkm_skip_time_line_ensure=True,
            )._ctkm_ensure_time_lines()
        # Mở form bước 11/12: đủ store từ file tổng, ir.rule lọc theo cửa hàng user.
        if self.ids and not self.env.context.get('ctkm_skip_receive_resync'):
            receive_tasks = self.filtered(
                lambda t: t.is_tem_receive_task or t.is_tem_replace_task
            )
            if receive_tasks:
                receive_tasks._ctkm_ensure_receive_lines_synced()
                receive_tasks.invalidate_recordset(['tem_tag_replace_ids'])
        try:
            return super().web_read(specification)
        except KeyError:
            # Odoo framework bug (web/models.py:153-154): trong web_read, bản ghi
            # many2one được đọc bởi user (line 149) bị lọc bởi ir.rule, nhưng vòng
            # lặp sau đó chạy dưới .sudo() (line 153) lại gồm cả bản ghi đó →
            # KeyError. Bước 11/12 (Nhận tem / Thay tem) mang store_verifier_ids
            # chứa Nhân viên / Quản lý cửa hàng của CÁC cửa hàng khác mà người đọc
            # không có quyền xem (record rule hr.employee/res.users) → lỗi này.
            # Đọc lại dưới sudo để display_name được phân giải (không lộ dữ liệu:
            # self đã là công việc người dùng có quyền mở, chỉ nới quyền đọc liên
            # kết nhân viên/người dùng nội bộ cùng CTKM).
            if any(t.is_tem_receive_task or t.is_tem_replace_task for t in self):
                _logger.warning(
                    'ctkm_core: web_read KeyError, retry as sudo (task_ids=%s)',
                    self.ids,
                )
                return self.sudo().web_read(specification)
            raise

    def _ctkm_ensure_time_lines(self):
        """Tạo dòng thời gian đầu tiên từ nội dung / ngày của công việc."""
        Line = self.env['ctkm.task.time.line'].sudo()
        for task in self:
            if task.time_line_ids:
                continue
            Line.create({
                'task_id': task.id,
                'name': task.checklist_step_name or task.name or '',
                'date_start': task.process_date or fields.Date.context_today(task),
                'date_end': task.done_date or False,
                'is_main': True,
                'sequence': 1,
            })

    def _ctkm_touch_main_time_line(self, vals):
        """Khi hoàn thành / đổi ngày xử lý: cập nhật dòng thời gian chính nếu còn trống."""
        for task in self:
            main = task.time_line_ids.filtered('is_main')[:1] or task.time_line_ids[:1]
            if not main:
                task._ctkm_ensure_time_lines()
                continue
            updates = {}
            if vals.get('done_date') and not main.date_end:
                updates['date_end'] = vals['done_date']
            if vals.get('process_date') and not main.date_start:
                updates['date_start'] = vals['process_date']
            if updates:
                main.with_context(ctkm_time_line_sync=True).write(updates)

    def _get_worker_employee(self):
        """Nhân viên gắn với người tạo công việc."""
        self.ensure_one()
        user = self.user_ids[:1]
        if not user:
            return self.env['hr.employee']
        employee = user.sudo().employee_id
        if employee:
            return employee
        return self.env['hr.employee'].sudo().search(
            [('user_id', '=', user.id)], limit=1
        )

    def _get_org_chart_manager_user(self):
        """Người nhận yêu cầu xác nhận: Người kiểm soát bước, hoặc quản lý org-chart.

        Khi bước đã cấu hình 'Người kiểm soát', KHÔNG fallback sang quản lý
        organization chart — người đó không được tick xác nhận trên form.
        """
        self.ensure_one()
        if self.sudo().verifier_ids:
            return self._get_verifier_user()
        employee = self._get_worker_employee()
        manager = employee.parent_id.sudo() if employee else self.env['hr.employee']
        user = manager.user_id
        if user and user.active and not user.share and user.partner_id:
            return user
        return self.env['res.users']

    def _get_verifier_users(self):
        """Tài khoản Odoo của các Người kiểm soát (sudo: nhân viên thường
        không đọc được user_id / employee của người khác).
        """
        self.ensure_one()
        users = self.env['res.users']
        for verifier in self.sudo().verifier_ids:
            user = verifier.sudo().user_id
            if user and user.active and not user.share and user.partner_id:
                users |= user
        return users

    def _get_verifier_user(self):
        """User của Người kiểm soát đầu tiên (nếu bước có cấu hình)."""
        return self._get_verifier_users()[:1]

    def _user_can_confirm_as_manager(self, user):
        """Người kiểm soát của bước (nếu có), quản lý org-chart, hoặc CTKM Admin.

        Với xác nhận theo cửa hàng (bước 11–12): chỉ Quản lý cửa hàng ĐÚNG
        cửa hàng của một Phụ trách (đã hoàn thành, chưa xác nhận) mới được.
        """
        self.ensure_one()
        if not user or user.share:
            return False
        if user.has_group('ctkm_core.group_ctkm_manager'):
            return True
        if self.store_verifier_ids:
            return bool(self.store_verifier_ids.filtered(
                lambda l: l.verifier_user_id == user
                and l.assignee_completed and not l.verified
            ))
        if self.sudo().verifier_ids:
            verifier_user_ids = set(self._get_verifier_users().ids)
            return user.id in verifier_user_ids
        manager_user = self._get_org_chart_manager_user()
        return bool(manager_user and manager_user.id == user.id)

    def _ctkm_task_form_url(self):
        self.ensure_one()
        menu = self.env.ref('ctkm_core.menu_ctkm_my_tasks', raise_if_not_found=False)
        app_menu_id = False
        if menu:
            app_menu = menu
            while app_menu.parent_id:
                app_menu = app_menu.parent_id
            app_menu_id = app_menu.id
        # Dùng model path (ctkm.task), không dùng action path (ctkm-my-tasks)
        # để tránh webclient gọi RPC với model sai. App CTKM trên thanh menu
        # do menu_id + ctkm_navbar_app_patch.js đảm nhiệm.
        url = '/odoo/ctkm.task/%s' % self.id
        if app_menu_id:
            url = '%s?menu_id=%s' % (url, app_menu_id)
        return url, app_menu_id

    def _ctkm_manager_confirm_button_markup(self):
        self.ensure_one()
        href, _app_menu_id = self._ctkm_task_form_url()
        return Markup(
            '<div class="o_ctkm_notify_detail mt-2">'
            '<a class="btn btn-primary btn-sm o_ctkm_manager_confirm_btn" '
            'href="%s" data-task-id="%s" contenteditable="false">'
            'Bấm để xác nhận hoàn thành'
            '</a>'
            '</div>'
        ) % (escape(href), self.id)

    def _ctkm_manager_confirm_message_body(self):
        self.ensure_one()
        worker_name = self.user_ids[:1].name or ''
        program_name = self.program_name or self.name or ''
        if self._ctkm_is_store_dispatch_step():
            pending = self.sudo().store_dispatch_ids.filtered(
                lambda rec: rec.state == 'pending'
            ).sorted('id')[:1]
            store_label = pending.store_label or pending.store_key or ''
            lines = [
                Markup('<b>Yêu cầu xác nhận gửi dữ liệu cửa hàng CTKM</b>'),
                Markup('Nhân viên <b>%s</b> đã bấm Gửi dữ liệu.') % escape(worker_name),
                Markup('Cửa hàng: <b>%s</b>') % escape(store_label),
                Markup('Công việc: <b>%s</b>') % escape(program_name),
                Markup(
                    'Vui lòng vào form và bấm <b>Xác nhận quản lý</b> '
                    'cho cửa hàng này.'
                ),
                self._ctkm_manager_confirm_button_markup(),
            ]
            return Markup('<br/>').join(lines)
        lines = [
            Markup('<b>Yêu cầu xác nhận hoàn thành công việc CTKM</b>'),
            Markup('Nhân viên <b>%s</b> đã bấm Hoàn thành.') % escape(worker_name),
            Markup('Công việc: <b>%s</b>') % escape(program_name),
            Markup(
                'Vui lòng vào form và tick <b>Xác nhận quản lý</b> '
                'để hoàn tất trạng thái.'
            ),
            self._ctkm_manager_confirm_button_markup(),
        ]
        return Markup('<br/>').join(lines)

    def _ctkm_worker_confirmed_button_markup(self):
        self.ensure_one()
        href, _app_menu_id = self._ctkm_task_form_url()
        return Markup(
            '<div class="o_ctkm_notify_detail mt-2">'
            '<a class="btn btn-primary btn-sm o_ctkm_task_open_btn" '
            'href="%s" data-task-id="%s" contenteditable="false">'
            'Bấm để xem công việc'
            '</a>'
            '</div>'
        ) % (escape(href), self.id)

    def _ctkm_worker_manager_confirmed_message_body(self, manager_user=None):
        self.ensure_one()
        manager_user = manager_user or self.env.user
        program_name = self.program_name or self.name or ''
        lines = [
            Markup('<b>Quản lý đã xác nhận hoàn thành công việc CTKM</b>'),
            Markup('Quản lý <b>%s</b> đã xác nhận.') % escape(manager_user.name or ''),
            Markup('Công việc: <b>%s</b>') % escape(program_name),
            Markup('Trạng thái công việc đã chuyển sang <b>Hoàn thành</b>.'),
            self._ctkm_worker_confirmed_button_markup(),
        ]
        return Markup('<br/>').join(lines)

    def _post_ctkm_bot_dm(self, recipient_user, body):
        """Gửi DM Discuss từ OdooBot CTKM tới một user."""
        self.ensure_one()
        Message = self.env['mail.message']
        recipient_user = recipient_user.sudo() if recipient_user else recipient_user
        if not recipient_user or recipient_user.share or not recipient_user.partner_id:
            return Message
        bot_user = self.env.ref(
            'business_discuss_bots.user_bot_ctkm', raise_if_not_found=False
        )
        if not bot_user or not bot_user.partner_id:
            raise UserError(_('Chưa cấu hình OdooBot CTKM trên hệ thống.'))
        try:
            chat = (
                self.env['discuss.channel']
                .sudo()
                .with_user(recipient_user)
                ._get_or_create_chat([bot_user.partner_id.id], pin=True)
            )
            return chat.with_user(bot_user).sudo().message_post(
                body=body,
                message_type='comment',
                subtype_xmlid='mail.mt_comment',
                author_id=bot_user.partner_id.id,
            )
        except Exception:
            _logger.exception(
                'ctkm_core: OdooBot CTKM DM failed task_id=%s recipient_user_id=%s',
                self.id,
                recipient_user.id,
            )
            return Message

    def _notify_org_manager_confirm(self):
        """Gửi tin OdooBot CTKM tới Người kiểm soát (hoặc quản lý org-chart)."""
        self.ensure_one()
        verifier_employees = self.sudo().verifier_ids
        if verifier_employees:
            recipients = self._get_verifier_users().filtered(
                lambda u: u not in self.user_ids
            )
            if not recipients:
                names = ', '.join(verifier_employees.mapped('display_name'))
                if self._get_verifier_users():
                    raise UserError(_(
                        'Người kiểm soát trùng với người làm việc. '
                        'Vui lòng kiểm tra lại cấu hình Người kiểm soát.'
                    ))
                raise UserError(_(
                    'Người kiểm soát (%s) chưa có tài khoản Odoo '
                    'hoặc tài khoản không hoạt động. '
                    'Không thể gửi yêu cầu xác nhận.'
                ) % names)
            role_label = _('Người kiểm soát')
        else:
            manager_user = self._get_org_chart_manager_user()
            if not manager_user:
                raise UserError(_(
                    'Không tìm thấy quản lý trực tiếp trên organization chart '
                    '(hoặc quản lý chưa có tài khoản Odoo). '
                    'Không thể gửi yêu cầu xác nhận.'
                ))
            if manager_user in self.user_ids:
                raise UserError(_(
                    'Quản lý trực tiếp trùng với người làm việc. '
                    'Vui lòng kiểm tra lại organization chart.'
                ))
            recipients = manager_user
            role_label = _('quản lý')
        notified_names = []
        for recipient in recipients:
            posted = self._post_ctkm_bot_dm(
                recipient, self._ctkm_manager_confirm_message_body()
            )
            if not posted:
                raise UserError(_(
                    'Không gửi được thông báo Discuss tới %s %s.'
                ) % (role_label, recipient.name))
            notified_names.append(recipient.name)
        self.message_post(
            body=_(
                'Đã gửi yêu cầu xác nhận tới %s <b>%s</b> qua OdooBot CTKM.'
            ) % (role_label, escape(', '.join(notified_names))),
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return recipients

    def _ctkm_store_verifier_message_body(self, line):
        """Nội dung tin OdooBot CTKM gửi Quản lý cửa hàng của MỘT cửa hàng.

        Chỉ ghi đúng Phụ trách (Cửa hàng trưởng) của cửa hàng đó, theo đúng
        chuỗi: Cửa hàng A (CHT A) → Quản lý cửa hàng A. Bước 10 (Bàn giao Tem
        Tag) không có Phụ trách riêng nên không ghi tên nhân viên.
        """
        self.ensure_one()
        store_label = line.store_label or ''
        program_name = self.program_name or self.program_id.name or self.name or ''
        if line.no_assignee:
            store_line = Markup(
                'Cửa hàng <b>%s</b> đã bấm Gửi dữ liệu.'
            ) % escape(store_label)
        else:
            worker = line.assignee_user_id
            worker_name = worker.name or ''
            store_line = Markup(
                'Cửa hàng <b>%s</b>: Nhân viên <b>%s</b> (Phụ trách) '
                'đã bấm Gửi dữ liệu.'
            ) % (escape(store_label), escape(worker_name))
        lines = [
            Markup('<b>Yêu cầu xác nhận hoàn thành công việc CTKM</b>'),
            store_line,
            Markup('Công việc: <b>%s</b>') % escape(program_name),
            Markup(
                'Vui lòng vào form và tick <b>Xác nhận theo cửa hàng</b> '
                'để hoàn tất trạng thái.'
            ),
            self._ctkm_manager_confirm_button_markup(),
        ]
        return Markup('<br/>').join(lines)

    def _notify_store_verifiers(self, store_keys=None):
        """Gửi tin OdooBot CTKM tới từng Quản lý cửa hàng (theo cửa hàng).

        Mỗi Quản lý cửa hàng chỉ nhận thông báo của cửa hàng mình, ghi rõ
        đúng Phụ trách (Cửa hàng trưởng) của cửa hàng đó đã bấm Gửi dữ liệu
        (không lấy nhân viên đầu tiên của toàn công việc).
        """
        self.ensure_one()
        if not self.store_verifier_ids:
            return self.env['res.users']
        recipients = self.env['res.users']
        notified_names = []
        allowed = set(store_keys) if store_keys else None
        for line in self.store_verifier_ids:
            if allowed is not None and not self._ctkm_store_in_allowed(
                allowed, line.store_key, line.store_label,
            ):
                continue
            user = line.verifier_user_id
            if not user or line.verified or user.share or not user.active:
                continue
            try:
                self._post_ctkm_bot_dm(
                    user, self._ctkm_store_verifier_message_body(line)
                )
                recipients |= user
                notified_names.append(user.name)
            except UserError:
                continue
        if notified_names:
            self.message_post(
                body=_(
                    'Đã gửi yêu cầu xác nhận theo cửa hàng tới Quản lý cửa hàng: '
                    '%(users)s.'
                ) % {'users': escape(', '.join(notified_names))},
                subtype_xmlid='mail.mt_note',
                body_is_html=True,
            )
        return recipients

    def _notify_worker_manager_confirmed(self, manager_user=None):
        """Gửi tin OdooBot CTKM cho người nhận việc khi quản lý đã xác nhận."""
        self.ensure_one()
        worker = self.user_ids[:1]
        if not worker or worker.share or not worker.partner_id:
            return self.env['res.users']
        manager_user = manager_user or self.env.user
        # Tránh spam nếu người nhận việc chính là người xác nhận.
        if worker == manager_user:
            return worker
        posted = self._post_ctkm_bot_dm(
            worker,
            self._ctkm_worker_manager_confirmed_message_body(manager_user),
        )
        if not posted:
            _logger.warning(
                'ctkm_core: cannot notify worker task_id=%s user_id=%s',
                self.id,
                worker.id,
            )
            return self.env['res.users']
        self.message_post(
            body=_(
                'Đã thông báo cho <b>%s</b> rằng quản lý đã xác nhận '
                '(qua OdooBot CTKM).'
            ) % worker.name,
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return worker

    def _ctkm_receive_notice_body(self, store, lines):
        """Nội dung tin OdooBot CTKM báo Quản lý cửa hàng về tem/tag đã nhận."""
        self.ensure_one()
        program_name = self.program_name or self.program_id.name or self.name or ''
        items = []
        for idx, line in enumerate(
            lines.filtered(lambda l: l.received).sorted(
                key=lambda l: (l.material_code or '', not l.is_tag)
            ),
            start=1,
        ):
            kind = _('Tag') if line.is_tag else _('Tem')
            qty = line.total_quantity or 0.0
            items.append(Markup(
                '%(n)s. %(code)s (%(kind)s) với số lượng %(qty)s'
            ) % {
                'n': idx,
                'code': escape(line.material_code or ''),
                'kind': kind,
                'qty': '{:,}'.format(qty),
            })
        body_lines = [
            Markup('<b>Thông báo nhận tem/tag mới</b>'),
            Markup('Chương trình khuyến mãi: <b>%s</b>') % escape(program_name),
            Markup('Cửa hàng: <b>%s</b>') % escape(store or ''),
            Markup('đã nhận số tem/tag là:'),
        ]
        if items:
            body_lines.append(Markup('<br/>').join(items))
        else:
            body_lines.append(Markup('(chưa có hạng mục nào được tick Đã nhận).'))
        body_lines.append(self._ctkm_worker_confirmed_button_markup())
        return Markup('<br/>').join(body_lines)

    def _ctkm_notify_store_received(self, store, lines):
        """Báo Quản lý cửa hàng (Người kiểm soát) của cửa hàng khi Cửa hàng
        trưởng đã tick Đã nhận toàn bộ tem/tag của cửa hàng đó (bước 11).

        Chỉ gửi 1 lần mỗi cửa hàng (cờ received_notified trên dòng store_verifier).
        """
        self.ensure_one()
        if not self.is_tem_receive_task:
            return self.env['res.users']
        store_key = _ctkm_normalize_store_key(store)
        sv = self.sudo().store_verifier_ids.filtered(
            lambda l: l.store_key == store_key
        )[:1]
        if not sv or sv.received_notified:
            return self.env['res.users']
        verifier_user = sv.verifier_user_id
        if not verifier_user or verifier_user.share or not verifier_user.active:
            return self.env['res.users']
        body = self._ctkm_receive_notice_body(store, lines)
        try:
            posted = self._post_ctkm_bot_dm(verifier_user, body)
        except UserError:
            posted = self.env['mail.message']
        if posted:
            sv.sudo().write({'received_notified': True})
            self.message_post(
                body=_(
                    'Đã gửi thông báo cửa hàng <b>%(store)s</b> đã nhận tem/tag '
                    'tới Quản lý cửa hàng <b>%(user)s</b> (OdooBot CTKM).'
                ) % {
                    'store': escape(store or ''),
                    'user': escape(verifier_user.name),
                },
                subtype_xmlid='mail.mt_note',
                body_is_html=True,
            )
        return verifier_user

    @api.onchange('print_store_search')
    def _onchange_print_store_search(self):
        search = (self.print_store_search or '').strip()
        pending = [('done', '=', False)]
        done = [('done', '=', True)]
        if search:
            pending.append(('store', 'ilike', search))
            done.append(('store', 'ilike', search))
        return {'domain': {
            'print_store_pending_ids': pending,
            'print_store_done_ids': done,
            'print_store_ids': [('store', 'ilike', search)] if search else [],
        }}

    @api.onchange('handover_store_search')
    def _onchange_handover_store_search(self):
        search = (self.handover_store_search or '').strip()
        domain = [('line_type', '=', 'handover')]
        if search:
            domain.append(('store', 'ilike', search))
        return {'domain': {'handover_store_ids': domain}}

    @api.onchange('collect_store_search')
    def _onchange_collect_store_search(self):
        search = (self.collect_store_search or '').strip()
        domain = [('line_type', '=', 'collect')]
        if search:
            domain.append(('store', 'ilike', search))
        return {'domain': {'collect_store_ids': domain}}

    @api.onchange('postcheck_store_search')
    def _onchange_postcheck_store_search(self):
        search = (self.postcheck_store_search or '').strip()
        domain = [('store', 'ilike', search)] if search else []
        return {'domain': {'postcheck_store_ids': domain}}

    @api.onchange('price_store_search')
    def _onchange_price_store_search(self):
        search = (self.price_store_search or '').strip()
        domain = [('store', 'ilike', search)] if search else []
        return {'domain': {'price_store_ids': domain}}

    @api.onchange('state')
    def _onchange_state(self):
        if self.state in ('waiting_confirm', 'done') and not self.done_date:
            self.done_date = fields.Date.context_today(self)

    @api.model_create_multi
    def create(self, vals_list):
        tasks = super().create(vals_list)
        # Công việc bước 4 / bước 12 được tạo sau khi đã import kho Tem/Tag
        # (hoặc bước trước chuyển tiếp) → dựng bảng "Chi tiết tem/tag" ngay.
        tasks.sudo()._ctkm_sync_tem_tag_lines()
        tasks.sudo()._ctkm_ensure_time_lines()
        return tasks

    @staticmethod
    def _ctkm_keep_print_split_commands(commands):
        """Hai One2many cùng inverse: bỏ (5)/(6) để không xóa dòng bảng kia."""
        kept = []
        for cmd in commands or []:
            code = cmd[0] if isinstance(cmd, (list, tuple)) and cmd else None
            if code in (3, 5, 6):
                continue
            kept.append(cmd)
        return kept

    def write(self, vals):
        vals = dict(vals)
        for fname in ('print_store_pending_ids', 'print_store_done_ids'):
            if fname in vals:
                vals[fname] = self._ctkm_keep_print_split_commands(vals[fname])
        # Đồng bộ từ checklist / nút workflow dùng context nội bộ.
        internal = (
            self.env.context.get('ctkm_internal_state_write')
            or self.env.context.get('ctkm_task_sync')
        )
        user_set_state = 'state' in vals

        if user_set_state and not internal:
            raise UserError(_(
                'Không thể đổi trạng thái trực tiếp. '
                'Dùng nút Hoàn thành hoặc Xác nhận quản lý.'
            ))

        if 'manager_confirmed' in vals and not internal:
            for task in self:
                if not task._user_can_confirm_as_manager(self.env.user):
                    raise UserError(_(
                        'Chỉ quản lý trực tiếp (theo organization chart) '
                        'mới được bấm Xác nhận quản lý.'
                    ))

        if vals.get('manager_confirmed'):
            for task in self:
                next_state = vals.get('state', task.state)
                if next_state not in ('waiting_confirm', 'done'):
                    raise UserError(_(
                        'Chỉ xác nhận quản lý được sau khi đã bấm Hoàn thành.'
                    ))
            vals['state'] = 'done'
        elif 'manager_confirmed' in vals and not vals['manager_confirmed']:
            # Bỏ tick xác nhận → hạ về chờ xác nhận.
            # Không ghi đè khi workflow nội bộ chủ động đánh dấu done
            # (bước không cần xác nhận quản lý: state=done + manager_confirmed=False).
            if not (internal and vals.get('state') == 'done'):
                next_state = vals.get('state')
                if next_state is None or next_state == 'done':
                    vals['state'] = 'waiting_confirm'

        if 'state' in vals and vals['state'] not in ('waiting_confirm', 'done'):
            vals['manager_confirmed'] = False

        if vals.get('state') == 'done' and not internal:
            for task in self:
                confirmed = (
                    vals['manager_confirmed']
                    if 'manager_confirmed' in vals
                    else task.manager_confirmed
                )
                if not confirmed:
                    raise UserError(_(
                        'Trạng thái Hoàn thành chỉ khi đã bấm Hoàn thành '
                        'và có Xác nhận quản lý.'
                    ))

        newly_confirmed = self.browse()
        if vals.get('manager_confirmed'):
            newly_confirmed = self.filtered(lambda t: not t.manager_confirmed)

        res = super().write(vals)

        if newly_confirmed:
            manager_user = self.env.user
            for task in newly_confirmed:
                task._notify_worker_manager_confirmed(manager_user)

        if 'checklist_line_id' in vals or 'state' in vals or 'done_date' in vals:
            self._ctkm_sync_checklist_from_task(
                checklist_line_id=vals.get('checklist_line_id'),
                state=vals.get('state'),
                done_date=vals.get('done_date'),
            )

        # Đổi CTKM / người nhận việc / bước → bảng "Chi tiết tem/tag" phải đổi theo
        # (bước 12 lọc theo cửa hàng của người nhận việc).
        if {'program_id', 'user_ids', 'name', 'checklist_line_id'} & set(vals):
            self.sudo()._ctkm_sync_tem_tag_lines()
            self.sudo().filtered(
                lambda t: t.is_tem_photo_task or t.is_tem_check_task
            )._ctkm_sync_tem_photo_lines()

        if {'process_date', 'done_date'} & set(vals):
            self.sudo()._ctkm_touch_main_time_line(vals)

        if 'state' in vals:
            self.sudo()._ctkm_sync_price_lines_for_programs(self.mapped('program_id'))

        return res

    def _ctkm_task_needs_manager_confirm(self):
        """Nguồn đúng: toggle 'Cần quản lý xác nhận' trên bước checklist.

        Nếu công việc có Xác nhận theo cửa hàng (store_verifier_ids), bắt buộc
        phải có xác nhận từ Quản lý cửa hàng theo từng cửa hàng (vd. bước 10,
        11, 12) dù toggle trên checklist có tắt.
        """
        self.ensure_one()
        if self.store_verifier_ids:
            return True
        if self.checklist_line_id:
            return bool(self.checklist_line_id.need_manager_confirm)
        return True

    def _ctkm_notify_reload(self, title, message, notif_type='success'):
        """Toast + reload form để hiện/ẩn đúng nút Hoàn thành / Chuyển tiếp."""
        params = {
            'title': title,
            'message': message,
            'type': notif_type,
            'sticky': False,
        }
        if len(self) == 1:
            params['next'] = {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'res_id': self.id,
                'view_mode': 'form',
                'views': [(False, 'form')],
                'target': 'current',
            }
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': params,
        }

    def action_mark_done(self):
        """Mỗi người nhận việc báo hoàn thành phần của mình. Công việc chỉ chuyển
        sang chờ xác nhận / hoàn tất khi TẤT CẢ người nhận việc đã hoàn thành.
        """
        notified = self.browse()
        already_done = self.browse()
        directly_done = self.browse()
        for task in self:
            if self.env.user not in task.user_ids and not self.env.user.has_group(
                'ctkm_core.group_ctkm_manager'
            ):
                raise UserError(_(
                    'Chỉ người nhận việc mới được bấm Hoàn thành.'
                ))
            if task.state == 'done':
                already_done |= task
                continue
            today = fields.Date.context_today(task)
            # Ghi nhận hoàn thành của người hiện tại (mỗi người 1 bản ghi).
            already_completed = bool(
                task.completion_ids.filtered(
                    lambda c: c.user_id == self.env.user and c.done
                )
            )
            if not already_completed:
                completion = task.completion_ids.filtered(
                    lambda c: c.user_id == self.env.user
                )[:1]
                if not completion:
                    completion = self.env['ctkm.task.user.completion'].sudo().create({
                        'task_id': task.id,
                        'user_id': self.env.user.id,
                    })
                completion.sudo().write({'done': True, 'done_date': today})
                task.invalidate_recordset(['completion_ids'])
            # Đã đủ người nhận việc hoàn thành chưa?
            if task.user_ids:
                done_user_ids = task.completion_ids.filtered('done').mapped('user_id.id')
                all_done = set(done_user_ids) >= set(task.user_ids.ids)
            else:
                all_done = True
            if not all_done:
                # Chờ các người nhận việc còn lại; chưa đổi trạng thái công việc.
                if already_completed:
                    already_done |= task
                continue
            # Bước 4 đổ BB: gán cửa hàng trưởng (bước 11–12), quản lý cửa hàng
            # làm Phụ trách bước 13 (chụp ảnh), giám sát (hậu kiểm) và Quản lý
            # cửa hàng làm Người kiểm soát bước 11–12.
            if task.is_tem_tag_import_task:
                task._ctkm_assign_store_managers_after_bb_import()
                task._ctkm_assign_photo_store_managers_after_bb_import()
                task._ctkm_assign_supervisors_after_bb_import()
                task._ctkm_assign_store_managers_verifier_after_bb_import()
            # Tất cả đã hoàn thành → tiếp tục luồng xác nhận như cũ.
            checklist = task.checklist_line_id
            need_confirm = task._ctkm_task_needs_manager_confirm()
            if not need_confirm:
                # Không cần xác nhận quản lý: nhân viên bấm Hoàn thành là xong,
                # không tìm quản lý trên org chart và không gửi tin xác nhận.
                task.with_context(ctkm_internal_state_write=True).write({
                    'state': 'done',
                    'manager_confirmed': False,
                    'done_date': task.done_date or today,
                })
                directly_done |= task
                continue
            if task.state == 'waiting_confirm' and not task.manager_confirmed:
                # Đang chờ xác nhận nhưng bước đã tắt "Cần quản lý xác nhận".
                if checklist and not checklist.need_manager_confirm:
                    task.with_context(ctkm_internal_state_write=True).write({
                        'state': 'done',
                        'manager_confirmed': False,
                        'done_date': task.done_date or today,
                    })
                    directly_done |= task
                    continue
                already_done |= task
                continue
            # Xác nhận theo cửa hàng (bước 10–12): mỗi Quản lý cửa hàng xác nhận
            # phần cửa hàng của mình. Gửi yêu cầu cho từng Quản lý cửa hàng.
            if task.store_verifier_ids:
                vals = {
                    'state': 'waiting_confirm',
                    'manager_confirmed': False,
                }
                if not task.done_date:
                    vals['done_date'] = today
                task.with_context(ctkm_internal_state_write=True).write(vals)
                task._notify_store_verifiers()
                notified |= task
                continue
            manager_user = task._get_org_chart_manager_user()
            if not manager_user:
                if task.sudo().verifier_ids:
                    raise UserError(_(
                        'Người kiểm soát (%s) chưa có tài khoản Odoo '
                        'hoặc tài khoản không hoạt động.'
                    ) % ', '.join(task.sudo().verifier_ids.mapped('display_name')))
                raise UserError(_(
                    'Không tìm thấy quản lý trực tiếp trên organization chart '
                    '(hoặc quản lý chưa có tài khoản Odoo).'
                ))
            vals = {
                'state': 'waiting_confirm',
                'manager_confirmed': False,
            }
            if not task.done_date:
                vals['done_date'] = today
            task.with_context(ctkm_internal_state_write=True).write(vals)
            task._notify_org_manager_confirm()
            notified |= task

        # Bước 10 "Thu hồi tem": xóa Tem/Tag đã chọn khỏi Kho ngay khi bấm Hoàn thành.
        done_tasks = (directly_done | already_done).filtered(
            lambda t: t.state == 'done' and t.is_tem_handover_task and t.recover_ids
        )
        if done_tasks:
            done_tasks.recover_ids._ctkm_recover_inventory()

        target = (directly_done or notified or already_done)[:1]
        if directly_done:
            return target._ctkm_notify_reload(
                _('Đã hoàn thành'),
                _(
                    'Công việc đã hoàn thành (không cần xác nhận quản lý). '
                    'Bấm Chuyển tiếp để giao bước sau.'
                ),
            )
        if notified:
            confirm_msg = _(
                'Đã bấm Hoàn thành. OdooBot CTKM đã gửi yêu cầu xác nhận '
                'tới quản lý trực tiếp.'
            )
            if target.sudo().verifier_ids:
                confirm_msg = _(
                    'Đã bấm Hoàn thành. OdooBot CTKM đã gửi yêu cầu xác nhận '
                    'tới Người kiểm soát.'
                )
            elif target.store_verifier_ids:
                confirm_msg = _(
                    'Đã bấm Hoàn thành. OdooBot CTKM đã gửi yêu cầu xác nhận '
                    'tới Quản lý cửa hàng.'
                )
            return target._ctkm_notify_reload(
                _('Đã gửi hoàn thành'),
                confirm_msg,
            )
        if already_done:
            pending_forward = already_done.filtered(
                lambda t: t.state == 'done' and not t.forwarded
            )
            if pending_forward:
                return pending_forward[:1]._ctkm_notify_reload(
                    _('Đã hoàn thành'),
                    _('Công việc đã xong. Bấm Chuyển tiếp để giao bước sau.'),
                    'warning',
                )
            return target._ctkm_notify_reload(
                _('Đã xử lý'),
                _('Công việc này đã ở trạng thái hoàn thành.'),
                'warning',
            )
        return True

    def action_manager_confirm(self):
        """Quản lý xác nhận hoàn thành → thông báo cho người nhận việc.

        Với xác nhận theo cửa hàng (bước 10/11/12): mỗi Quản lý cửa hàng chỉ
        xác nhận được phần của cửa hàng ĐÚNG mình quản lí. Bước 10 (Bàn giao Tem
        Tag) không có Phụ trách riêng nên Quản lý cửa hàng xác nhận trực tiếp
        phần cửa hàng của mình. Công việc chỉ chuyển Hoàn thành khi TẤT CẢ cửa
        hàng đã được xác nhận.
        """
        user = self.env.user
        for task in self:
            if task.store_verifier_ids:
                if task.state not in ('waiting_confirm', 'done'):
                    raise UserError(_(
                        'Chỉ xác nhận được sau khi nhân viên đã bấm Hoàn thành.'
                    ))
                pending = task.store_verifier_ids.filtered(
                    lambda l: l.assignee_completed and not l.verified
                )
                if not user.has_group('ctkm_core.group_ctkm_manager'):
                    pending = pending.filtered(
                        lambda l: l.verifier_user_id == user
                    )
                if not pending:
                    if all(l.verified for l in task.store_verifier_ids):
                        # Đã xong hết, không còn gì để xác nhận.
                        continue
                    raise UserError(_(
                        'Bạn không phải Quản lý cửa hàng của cửa hàng có công '
                        'việc đang chờ xác nhận, hoặc cửa hàng đó đã xác nhận.'
                    ))
                pending._ctkm_verify(user)
                if all(l.verified for l in task.store_verifier_ids):
                    task.with_context(ctkm_internal_state_write=True).write({
                        'manager_confirmed': True,
                        'state': 'done',
                    })
                    # Bước 10 "Thu hồi tem": xóa Tem/Tag đã chọn khỏi Kho
                    # ngay khi TẤT CẢ Quản lý cửa hàng đã xác nhận xong.
                    if task.is_tem_handover_task and task.recover_ids:
                        task.recover_ids._ctkm_recover_inventory()
                else:
                    remaining = task.store_verifier_ids.filtered(
                        lambda l: not l.verified
                    ).mapped('store_label')
                    task.message_post(
                        body=_(
                            'Đã xác nhận phần cửa hàng của bạn. Còn chờ Quản lý '
                            'cửa hàng xác nhận các cửa hàng: %s.'
                        ) % escape(', '.join(remaining)),
                        subtype_xmlid='mail.mt_note',
                        body_is_html=True,
                    )
                continue
            if not task._user_can_confirm_as_manager(user):
                if task.sudo().verifier_ids:
                    raise UserError(_(
                        'Chỉ Người kiểm soát (%s) mới được xác nhận hoàn thành.'
                    ) % ', '.join(task.sudo().verifier_ids.mapped('display_name')))
                raise UserError(_(
                    'Chỉ quản lý trực tiếp (theo organization chart) '
                    'mới được xác nhận hoàn thành.'
                ))
            if task.state not in ('waiting_confirm', 'done'):
                raise UserError(_(
                    'Chỉ xác nhận được sau khi nhân viên đã bấm Hoàn thành.'
                ))
            if task.manager_confirmed and task.state == 'done':
                continue
            task.with_context(ctkm_internal_state_write=True).write({
                'manager_confirmed': True,
                'state': 'done',
            })
        target = self[:1]
        return target._ctkm_notify_reload(
            _('Đã xác nhận'),
            _(
                'Đã xác nhận quản lý. OdooBot CTKM đã thông báo '
                'cho người nhận việc. Người nhận việc bấm Chuyển tiếp '
                'để giao bước sau.'
            ),
        )

    def action_notify_support(self):
        """Gửi thông báo Discuss cho người hỗ trợ / bàn giao."""
        self.ensure_one()
        partners = self.support_employee_ids.mapped('user_id.partner_id')
        if self.handover_employee_id.user_id.partner_id:
            partners |= self.handover_employee_id.user_id.partner_id
        if not partners:
            raise UserError(_(
                'Chưa chọn người hỗ trợ / bàn giao có tài khoản Odoo để gửi thông báo.'
            ))
        body = _(
            'Bạn được nhờ hỗ trợ công việc CTKM:<br/>'
            '<b>%(content)s</b><br/>'
            'Ngày xử lý: %(process_date)s — Trạng thái: %(state)s'
        ) % {
            'content': self.program_name or self.name or '',
            'process_date': self.process_date or '',
            'state': dict(self._fields['state'].selection).get(self.state, ''),
        }
        self.message_post(
            body=body,
            partner_ids=partners.ids,
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Đã gửi thông báo'),
                'message': _('Đã thông báo tới %s người.') % len(partners),
                'type': 'success',
                'sticky': False,
            },
        }

    def action_advance_stage(self):
        """Chuyển tiếp → giao việc bước tiến độ tiếp theo (kèm ghi chú/file)."""
        self.ensure_one()
        if self.state != 'done':
            raise UserError(_(
                'Chỉ chuyển tiếp được sau khi đã bấm Hoàn thành '
                '(trạng thái công việc phải là Hoàn thành).'
            ))
        if self._ctkm_task_needs_manager_confirm() and not self.manager_confirmed:
            raise UserError(_(
                'Cần có Xác nhận quản lý trước khi chuyển tiếp.'
            ))
        if self.forwarded:
            raise UserError(_('Công việc này đã chuyển tiếp rồi.'))
        if not self.program_id:
            raise UserError(_('Công việc chưa gắn chương trình khuyến mãi.'))

        program = self.program_id.sudo()
        current_checklist = self.checklist_line_id.sudo()
        current_notify = self.notify_line_id.sudo()
        step_label = (
            current_checklist.name
            or current_notify.step_label
            or _('(không xác định)')
        )
        self.write({'forwarded': True})
        self.message_post(
            body=_('Đã bấm <b>Chuyển tiếp</b> cho bước <b>%s</b>.')
            % escape(step_label),
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )

        # Ưu tiên chuỗi Tiến độ thực hiện
        if current_checklist:
            return self._ctkm_advance_checklist_step(program, current_checklist)

        # Tương thích task cũ gắn Phạm vi thông báo
        if current_notify:
            return self._ctkm_advance_notify_line_step(program, current_notify)

        return self._ctkm_advance_program_stage(program)

    def _ctkm_advance_checklist_step(self, program, current_line):
        """Sau Chuyển tiếp: gửi tin việc cho bước tiến độ kế tiếp."""
        self.ensure_one()
        next_line = program._ctkm_next_checklist_line(current_line)
        if next_line and not next_line.user_ids:
            raise UserError(_(
                'Bước kế tiếp "%s" chưa có người phụ trách. '
                'Gán người phụ trách trên Tiến độ thực hiện rồi bấm Chuyển tiếp lại.'
            ) % (next_line.name or ''))
        if next_line:
            handover_note, handover_docs = self._ctkm_collect_handover_from_checklist(
                program, current_line
            )
            sent = program.sudo()._ctkm_send_checklist_step_notify(
                next_line,
                handover_note=handover_note,
                handover_attachments=handover_docs.sudo() if handover_docs else handover_docs,
            )
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Đã chuyển bước'),
                    'message': _(
                        'Đã gửi OdooBot CTKM giao việc bước "%(step)s" cho %(user)s, '
                        'kèm ghi chú và tài liệu đã đẩy qua.'
                    ) % {
                        'step': next_line.name or '',
                        'user': ", ".join(sent.mapped('name')) if sent else (next_line.user_ids[:1].name or ''),
                    },
                    'type': 'success',
                    'sticky': False,
                },
            }
        return self._ctkm_advance_program_stage(program)

    def _ctkm_advance_notify_line_step(self, program, current_line):
        """Giữ luồng Chuyển tiếp theo phạm vi thông báo (task cũ)."""
        self.ensure_one()
        siblings = self.sudo().search([
            ('program_id', '=', program.id),
            ('notify_line_id', '=', current_line.id),
        ])
        pending = siblings.filtered(
            lambda t: not t.forwarded or t.state != 'done' or not t.manager_confirmed
        )
        if pending:
            names = ', '.join(pending.mapped('user_id.name'))
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Đã chuyển tiếp phần của bạn'),
                    'message': _(
                        'Chờ các người còn lại hoàn tất bước "%(step)s": %(names)s'
                    ) % {
                        'step': current_line.step_label or '',
                        'names': names,
                    },
                    'type': 'warning',
                    'sticky': False,
                },
            }

        lines = program.notify_line_ids.sorted(lambda l: (l.sequence, l.id))
        next_line = self.env['ctkm.program.notify.line']
        found_current = False
        for line in lines:
            if found_current:
                next_line = line
                break
            if line.id == current_line.id:
                found_current = True

        if next_line:
            handover_note, handover_docs = self._ctkm_collect_handover_from_line(
                program, current_line
            )
            sent = program.sudo()._ctkm_send_notify_line(
                next_line,
                handover_note=handover_note,
                handover_attachments=handover_docs.sudo() if handover_docs else handover_docs,
            )
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Đã chuyển bước'),
                    'message': _(
                        'Đã gửi OdooBot CTKM tới bước "%(step)s" (%(count)s người), '
                        'kèm ghi chú và tài liệu đã đẩy qua.'
                    ) % {
                        'step': next_line.step_label or '',
                        'count': len(sent),
                    },
                    'type': 'success',
                    'sticky': False,
                },
            }

        return self._ctkm_advance_program_stage(program)

    def _ctkm_advance_program_stage(self, program):
        """Chuyển giai đoạn CTKM khi hết bước phạm vi / task không gắn dòng."""
        self.ensure_one()
        current = program.stage_id
        Stage = self.env['ctkm.stage'].sudo()

        if current and current.pipe_end:
            raise UserError(_(
                'Chương trình "%(program)s" đã ở giai đoạn kết thúc "%(stage)s".'
            ) % {
                'program': program.display_name,
                'stage': current.display_name,
            })

        domain = [('sequence', '>', current.sequence)] if current else []
        next_stage = Stage.search(domain, order='sequence, id', limit=1)
        if not next_stage and current:
            next_stage = Stage.search(
                [('sequence', '=', current.sequence), ('id', '>', current.id)],
                order='id',
                limit=1,
            )
        if not next_stage:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Hoàn tất các bước phạm vi'),
                    'message': _(
                        'Đã chuyển tiếp hết các bước người nhận. '
                        'Không còn giai đoạn chương trình tiếp theo.'
                    ),
                    'type': 'success',
                    'sticky': False,
                },
            }

        old_name = current.display_name if current else _('(chưa có)')
        program.write({
            'stage_id': next_stage.id,
            'kanban_state': 'normal',
        })
        self.message_post(
            body=_(
                'Đã chuyển bước chương trình <b>%(program)s</b>: '
                '%(old)s → <b>%(new)s</b>'
            ) % {
                'program': program.display_name,
                'old': old_name,
                'new': next_stage.display_name,
            },
            subtype_xmlid='mail.mt_note',
            body_is_html=True,
        )
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Đã chuyển bước'),
                'message': _('%(old)s → %(new)s') % {
                    'old': old_name,
                    'new': next_stage.display_name,
                },
                'type': 'success',
                'sticky': False,
            },
        }

    @api.model
    def _task_content_from_program(self, program):
        content = program.name or _('Công việc CTKM')
        description = html2plaintext(program.description or '').replace('\xa0', ' ').strip()
        if description:
            content = '%s\n%s' % (content, description)
        return content

    @api.model
    def _duplicate_attachments_for_task(self, attachments, task, description_prefix=None):
        """Copy file để bước sau / nhân viên đọc được (ACL theo task mới).

        Luôn sudo: người bấm Chuyển tiếp không có quyền ghi attachment
        lên task của người bước tiếp theo.
        """
        Attachment = self.env['ir.attachment'].sudo()
        copies = Attachment
        for att in attachments.sudo():
            vals = {
                'res_model': 'ctkm.task',
                'res_id': task.id,
                'name': att.name,
            }
            if description_prefix:
                vals['description'] = '%s%s' % (description_prefix, att.id)
            copies |= att.copy(vals)
        return copies

    def _ensure_program_notify_documents(self):
        """Copy tài liệu CTKM vào task để nhân viên không bị chặn ACL program."""
        for task in self.sudo():
            program = task.program_id
            if not program or not task.id:
                continue
            Attachment = self.env['ir.attachment'].sudo()
            copies = Attachment
            for doc in program.notify_document_ids:
                marker = 'ctkm_program_doc:%s' % doc.id
                existing = Attachment.search([
                    ('res_model', '=', 'ctkm.task'),
                    ('res_id', '=', task.id),
                    ('description', '=', marker),
                ], limit=1)
                if not existing:
                    existing = doc.copy({
                        'res_model': 'ctkm.task',
                        'res_id': task.id,
                        'name': doc.name,
                        'description': marker,
                    })
                copies |= existing
            task.program_notify_document_ids = [(6, 0, copies.ids)]

    @api.model
    def _ctkm_collect_handover_from_line(self, program, notify_line):
        """Gom ghi chú/file chương trình + task đã chuyển tiếp của bước phạm vi."""
        program.ensure_one()
        note_parts = []
        docs = self.env['ir.attachment']
        tasks = self.sudo().search([
            ('program_id', '=', program.id),
            ('notify_line_id', '=', notify_line.id),
            ('forwarded', '=', True),
        ], order='id')
        # Handover đã có trên bước này (gồm chương trình + bước trước hơn)
        if tasks and tasks[0].handover_note:
            note_parts.append(Markup(tasks[0].handover_note))
            docs |= tasks[0].handover_document_ids
        elif program.note:
            note_parts.append(
                Markup('<p><b>%s</b></p>%s')
                % (escape(_('Ghi chú chương trình')), Markup(program.note))
            )
        if not tasks or not tasks[0].handover_document_ids:
            docs |= program.notify_document_ids

        for task in tasks:
            step = task.notify_step_label or _('Bước trước')
            worker = task.user_ids[:1].name or ''
            if task.work_note:
                note_parts.append(
                    Markup('<p><b>%s</b></p>%s')
                    % (
                        escape(_('Ghi chú từ %s (%s)') % (worker, step)),
                        Markup(task.work_note),
                    )
                )
            docs |= task.work_document_ids
        handover_note = Markup('<hr/>').join(note_parts) if note_parts else False
        return handover_note, docs

    @api.model
    def _ctkm_collect_handover_from_checklist(self, program, checklist_line):
        """Gom ghi chú/file từ bước tiến độ đã chuyển tiếp."""
        program.ensure_one()
        note_parts = []
        docs = self.env['ir.attachment']
        tasks = self.sudo().search([
            ('program_id', '=', program.id),
            ('checklist_line_id', '=', checklist_line.id),
            ('forwarded', '=', True),
        ], order='id')
        if tasks and tasks[0].handover_note:
            note_parts.append(Markup(tasks[0].handover_note))
            docs |= tasks[0].handover_document_ids
        elif program.note:
            note_parts.append(
                Markup('<p><b>%s</b></p>%s')
                % (escape(_('Ghi chú chương trình')), Markup(program.note))
            )
        if not tasks or not tasks[0].handover_document_ids:
            docs |= program.notify_document_ids

        for task in tasks:
            step = task.checklist_step_name or _('Bước trước')
            worker = task.user_ids[:1].name or ''
            if task.work_note:
                note_parts.append(
                    Markup('<p><b>%s</b></p>%s')
                    % (
                        escape(_('Ghi chú từ %s (%s)') % (worker, step)),
                        Markup(task.work_note),
                    )
                )
            docs |= task.work_document_ids
        handover_note = Markup('<hr/>').join(note_parts) if note_parts else False
        return handover_note, docs

    def _ctkm_sync_checklist_from_task(self, checklist_line_id=None, state=None, done_date=None):
        """Đồng bộ trạng thái/ngày xong từ công việc về bước checklist."""
        self.ensure_one()
        checklist = self.checklist_line_id
        if not checklist and checklist_line_id:
            checklist = self.env['ctkm.program.checklist.line'].browse(checklist_line_id)
        if not checklist or not checklist.exists():
            return
        vals = {}
        if state:
            mapped = {'todo': 'todo', 'progress': 'progress', 'waiting_confirm': 'progress', 'done': 'done'}
            vals['state'] = mapped.get(state, state)
        if done_date:
            vals['done_date'] = done_date
        if vals and not self.env.context.get('ctkm_task_sync'):
            checklist.sudo().with_context(ctkm_task_sync=True).write(vals)

    @api.model
    def _ctkm_sync_task_from_checklist(self, checklist):
        """Đồng bộ công việc từ bước checklist (gán người / tiến độ trên CTKM).

        Trạng thái Hoàn thành của task vẫn đi qua nút Hoàn thành / Xác nhận quản lý.
        Checklist chỉ đẩy todo/progress; cột Xong cập nhật done_date, không ép task = done.
        """
        if not checklist or not checklist.exists():
            return self.browse()
        Task = self.sudo()
        task = Task.search([
            ('program_id', '=', checklist.program_id.id),
            ('checklist_line_id', '=', checklist.id),
        ], limit=1)
        if not task:
            return self.browse()
        vals = {}
        if checklist.user_ids and task.user_ids != checklist.user_ids:
            vals['user_ids'] = [(6, 0, checklist.user_ids.ids)]
        if checklist.state in ('todo', 'progress'):
            if task.state not in ('waiting_confirm', 'done'):
                vals['state'] = checklist.state
        if checklist.done_date and task.done_date != checklist.done_date:
            vals['done_date'] = checklist.done_date
        if checklist.name and task.name != checklist.name:
            vals['name'] = checklist.name
        verifier_ids = checklist.verifier_ids
        if (verifier_ids or task.verifier_ids) and task.verifier_ids != verifier_ids:
            vals['verifier_ids'] = [(6, 0, verifier_ids.ids)]
        if vals and not task.env.context.get('ctkm_task_sync'):
            task.with_context(
                ctkm_task_sync=True,
                ctkm_internal_state_write=True,
            ).write(vals)
        # Nếu đang chờ xác nhận quản lý mà giờ không còn cần, tự động hoàn thành.
        # Invalidate compute để form hiện nút Chuyển tiếp đúng.
        task.invalidate_recordset(['checklist_need_manager_confirm'])
        if not checklist.need_manager_confirm and task.state == 'waiting_confirm':
            task.with_context(
                ctkm_task_sync=True,
                ctkm_internal_state_write=True,
            ).write({
                'state': 'done',
                'manager_confirmed': False,
                'done_date': task.done_date or fields.Date.context_today(task),
            })
        return task

    @api.model
    def _get_or_create_for_program_user(
        self, program, user, notify_line=None,
        handover_note=False, handover_attachments=None,
    ):
        """Một bước (phạm vi / checklist) = một công việc chia sẻ nhiều người nhận."""
        program.ensure_one()
        if not user or not user.exists():
            return self.browse()
        Task = self.sudo()
        domain = [('program_id', '=', program.id)]
        if notify_line:
            domain.append(('notify_line_id', '=', notify_line.id))
        else:
            domain.append(('notify_line_id', '=', False))
        task = Task.search(domain, limit=1)
        if not task and not notify_line:
            checklist = self._ctkm_pick_checklist_line_for_user(program, user)
            if checklist:
                task = Task.search([
                    ('program_id', '=', program.id),
                    ('checklist_line_id', '=', checklist.id),
                ], limit=1)
        if task:
            # Đảm bảo user nằm trong danh sách người nhận việc.
            if user not in task.user_ids:
                task.with_context(
                    ctkm_task_sync=True, ctkm_internal_state_write=True
                ).write({'user_ids': [(4, user.id)]})
            vals = {}
            if handover_note and not task.handover_note:
                vals['handover_note'] = handover_note
            if vals:
                task.write(vals)
            if handover_attachments and not task.handover_document_ids:
                copies = self._duplicate_attachments_for_task(handover_attachments, task)
                task.handover_document_ids = [(6, 0, copies.ids)]
            task._ensure_program_notify_documents()
            return task
        vals = {
            'program_id': program.id,
            'user_ids': [(6, 0, [user.id])],
            'process_date': fields.Date.context_today(self),
            'name': self._task_content_from_program(program),
            'state': 'todo',
            'company_id': program.company_id.id or self.env.company.id,
            'notify_line_id': notify_line.id if notify_line else False,
            'handover_note': handover_note or False,
        }
        if not notify_line:
            checklist = self._ctkm_pick_checklist_line_for_user(program, user)
            if checklist:
                vals['checklist_line_id'] = checklist.id
                vals['name'] = checklist.name
        try:
            with self.env.cr.savepoint():
                task = Task.create(vals)
        except IntegrityError:
            task = Task.search(domain, limit=1)
            if not task and checklist:
                task = Task.search([
                    ('program_id', '=', program.id),
                    ('checklist_line_id', '=', checklist.id),
                ], limit=1)
            if task:
                if user not in task.user_ids:
                    task.with_context(
                        ctkm_task_sync=True, ctkm_internal_state_write=True
                    ).write({'user_ids': [(4, user.id)]})
                task._ensure_program_notify_documents()
            return task
        if handover_attachments:
            copies = self._duplicate_attachments_for_task(handover_attachments, task)
            task.handover_document_ids = [(6, 0, copies.ids)]
        task._ensure_program_notify_documents()
        return task

    @api.model
    def _ctkm_pick_checklist_line_for_user(self, program, user):
        """Bước đúng lượt theo STT: không nhảy qua bước trước còn chưa xong.

        Nếu user đã xong bước của mình nhưng chưa bấm Chuyển tiếp, giữ bước đó
        để còn giao việc cho người phụ trách bước kế. Không mở bước sau khi
        chương trình vẫn đang dừng ở một bước của người khác.
        """
        empty = self.env['ctkm.program.checklist.line']
        all_lines = program.checklist_line_ids.sorted(
            lambda line: (line.sequence, line.id)
        )
        mine = all_lines.filtered(lambda line: user in line.user_ids)
        if not mine:
            return empty

        Task = self.sudo()
        pending_forward = Task.search([
            ('program_id', '=', program.id),
            ('user_ids', 'in', [user.id]),
            ('checklist_line_id', '!=', False),
            ('state', '=', 'done'),
            ('forwarded', '=', False),
        ], order='checklist_line_id, id', limit=1)
        if pending_forward:
            return pending_forward.checklist_line_id

        current = all_lines.filtered(lambda line: line.state != 'done')[:1]
        if current and user in current.user_ids:
            return current
        return empty

    @api.model
    def _ctkm_find_current_checklist_task(self, program, user):
        """Mở task của bước đang tới lượt, không lấy task id mới nhất."""
        Task = self.sudo()
        checklist = self._ctkm_pick_checklist_line_for_user(program, user)
        if not checklist:
            return Task.browse()
        task = Task.search([
            ('program_id', '=', program.id),
            ('user_ids', 'in', [user.id]),
            ('checklist_line_id', '=', checklist.id),
            ('forwarded', '=', False),
        ], limit=1)
        if task:
            return task
        return checklist._ctkm_ensure_task()

    @api.model
    def action_open_for_program(self, program_id):
        """Tạo/mở công việc khi bấm nút trong Discuss (không phụ thuộc ACL form CTKM)."""
        program_id = int(program_id or 0)
        if not program_id:
            raise UserError(_('Thiếu mã chương trình khuyến mãi.'))
        program = self.env['ctkm.program'].sudo().browse(program_id)
        if not program.exists():
            raise UserError(_('Không tìm thấy chương trình khuyến mãi.'))

        user = self.env.user
        notified_users = program.notify_line_ids.notify_employee_ids.mapped('user_id')
        checklist_users = program.checklist_line_ids.mapped('user_ids')
        allowed = (
            user.has_group('ctkm_core.group_ctkm_user')
            or program.user_id == user
            or user in notified_users
            or user in checklist_users
        )
        if not allowed:
            raise UserError(_('Bạn không có quyền mở công việc của chương trình này.'))

        # Chọn đúng bước đang tới lượt (theo STT), không lấy task id mới nhất
        # (Hạnh phụ trách nhiều bước → id mới nhất dễ nhảy thẳng bước sau).
        task = self._ctkm_find_current_checklist_task(program, user)
        if not task and not program.checklist_line_ids:
            line = program.notify_line_ids.filtered(
                lambda l: user in l.notify_employee_ids.mapped('user_id') and l.notified
            )[:1]
            if not line:
                line = program.notify_line_ids.filtered(
                    lambda l: user in l.notify_employee_ids.mapped('user_id')
                ).sorted(lambda l: (l.sequence, l.id))[:1]
            task = self._get_or_create_for_program_user(
                program, user, notify_line=line or None
            )
        if not task:
            raise UserError(_(
                'Hiện chưa tới lượt công việc của bạn trên chương trình này. '
                'Chờ người phụ trách bước trước hoàn thành và bấm Chuyển tiếp.'
            ))

        task._ensure_program_notify_documents()
        url, app_menu_id = task._ctkm_task_form_url()
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
            'task_id': task.id,
            'menu_id': app_menu_id,
        }

    @api.model
    def action_open_task(self, task_id):
        """Mở đúng công việc đã gắn trên nút Discuss (bước 9, 10, …)."""
        task_id = int(task_id or 0)
        if not task_id:
            raise UserError(_('Thiếu mã công việc CTKM.'))
        task = self.sudo().browse(task_id)
        if not task.exists():
            raise UserError(_('Không tìm thấy công việc CTKM.'))
        user = self.env.user
        allowed = (
            user in task.user_ids
            or task._user_can_confirm_as_manager(user)
            or user.has_group('ctkm_core.group_ctkm_manager')
            or user.has_group('ctkm_core.group_ctkm_user')
            or (task.program_id and task.program_id.user_id == user)
        )
        if not allowed:
            raise UserError(_('Bạn không có quyền mở công việc này.'))
        task._ensure_program_notify_documents()
        url, app_menu_id = task._ctkm_task_form_url()
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
            'task_id': task.id,
            'menu_id': app_menu_id,
        }

    @api.model
    def action_open_for_manager_confirm(self, task_id):
        """Mở form công việc để quản lý xác nhận (từ tin OdooBot CTKM)."""
        task_id = int(task_id or 0)
        if not task_id:
            raise UserError(_('Thiếu mã công việc CTKM.'))
        task = self.sudo().browse(task_id)
        if not task.exists():
            raise UserError(_('Không tìm thấy công việc CTKM.'))
        user = self.env.user
        allowed = (
            task._user_can_confirm_as_manager(user)
            or user in task.user_ids
            or user.has_group('ctkm_core.group_ctkm_manager')
        )
        if not allowed:
            raise UserError(_(
                'Bạn không có quyền mở công việc này để xác nhận.'
            ))
        task._ensure_program_notify_documents()
        url, app_menu_id = task._ctkm_task_form_url()
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
            'task_id': task.id,
            'menu_id': app_menu_id,
        }

    def action_open_tem_tag_import(self):
        self.ensure_one()
        if not self.program_id:
            raise UserError(_('Công việc chưa gắn chương trình khuyến mãi.'))
        if not self.is_tem_tag_import_task:
            raise UserError(_('Công việc này không phải bước import Tem/Tag.'))
        action = self.env.ref('ctkm_inventory.action_ctkm_inventory_import_wizard', raise_if_not_found=False)
        if not action:
            raise UserError(_('Module Kho Tem/Tag chưa cài đặt.'))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Import Tem/Tag'),
            'res_model': 'ctkm.inventory.import.wizard',
            'view_mode': 'form',
            'view_id': self.env.ref('ctkm_inventory.view_ctkm_inventory_import_wizard_form').id,
            'target': 'new',
            'context': {
                'default_program_id': self.program_id.id,
                'default_replace_existing': True,
                'ctkm_import_task_id': self.id,
            },
        }


    @api.model
    def _ctkm_get_user_managed_stores(self, user=None):
        """Danh sách cửa hàng mà user quản lý hoặc được phân quyền xem."""
        user = user or self.env.user
        stores = []
        seen = set()

        def _add(key, code, name):
            if not key:
                return
            canon = self._ctkm_store_canonical_key(key, code, name) or key
            if canon in seen:
                if name and name != canon:
                    for s in stores:
                        if s['key'] == canon and (s['name'] == canon or len(name) > len(s['name'])):
                            s['name'] = name
                return
            seen.add(canon)
            stores.append({
                'key': canon,
                'code': code or canon,
                'name': name or code or canon,
            })

        # 1. Cửa hàng từ hồ sơ nhân viên (hr.employee)
        employees = self.env['hr.employee'].sudo()
        if 'employee_id' in user._fields and user.employee_id:
            employees |= user.employee_id.sudo()
        if user.employee_ids:
            employees |= user.employee_ids.sudo()

        for emp in employees:
            if 'managed_store_ids' in emp._fields:
                for s in emp.managed_store_ids:
                    k = _ctkm_normalize_store_key(s.code or s.name)
                    _add(k, s.code, s.name)
            if 'store_id' in emp._fields and emp.store_id:
                s = emp.store_id
                k = _ctkm_normalize_store_key(s.code or s.name)
                _add(k, s.code, s.name)
            if 'ma_bo_phan_id' in emp._fields and emp.ma_bo_phan_id:
                s = emp.ma_bo_phan_id
                k = _ctkm_normalize_store_key(s.code or s.name)
                _add(k, s.code, s.name)
            if 'ma_bo_phan' in emp._fields and emp.ma_bo_phan:
                k = _ctkm_normalize_store_key(emp.ma_bo_phan)
                _add(k, emp.ma_bo_phan, emp.ma_bo_phan)

        # 2. LUG Permission: user.assigned_ma_bo_phan_ids
        if 'assigned_ma_bo_phan_ids' in user._fields:
            for s in user.sudo().assigned_ma_bo_phan_ids:
                k = _ctkm_normalize_store_key(s.code or s.name)
                _add(k, s.code, s.name)

        # 3. Cửa hàng từ ctkm.task.store.verifier (Người kiểm soát / Quản lý cửa hàng / Phụ trách)
        if 'ctkm.task.store.verifier' in self.env:
            sv_records = self.env['ctkm.task.store.verifier'].sudo().search([
                '|', '|',
                    ('verifier_user_id', '=', user.id),
                    ('verifier_id.user_id', '=', user.id),
                    ('assignee_user_id', '=', user.id),
            ])
            for sv in sv_records:
                k = _ctkm_normalize_store_key(sv.store_key or sv.store_label)
                _add(k, sv.store_key, sv.store_label or sv.store_key)

        # 4. Cửa hàng từ mã bộ phận của user (_ctkm_user_department_store_keys)
        dept_keys = self._ctkm_user_department_store_keys(user)
        for dk in dept_keys:
            canon = self._ctkm_store_canonical_key(dk) or dk
            if canon not in seen:
                s_name = dk
                if 'hr.store' in self.env:
                    st_rec = self.env['hr.store'].sudo().search([
                        '|', ('code', '=', dk), ('name', '=', dk),
                    ], limit=1)
                    if st_rec:
                        s_name = st_rec.name
                _add(canon, dk, s_name)

        # 5. Cửa hàng từ các công việc mà user là Người kiểm soát (verifier_ids)
        TaskModel = self.env['ctkm.task'].sudo()
        verifier_tasks = TaskModel.search([
            ('verifier_ids.user_id', 'in', [user.id]),
            ('program_id.state', 'not in', ['draft', 'cancel']),
        ])
        for vt in verifier_tasks:
            for sv in vt.store_verifier_ids:
                k = _ctkm_normalize_store_key(sv.store_key or sv.store_label)
                _add(k, sv.store_key, sv.store_label or sv.store_key)
            for d in vt.store_dispatch_ids:
                k = _ctkm_normalize_store_key(d.store_key or d.store_label)
                _add(k, d.store_key, d.store_label or d.store_key)

        stores.sort(key=lambda s: s.get('name') or s.get('code') or s.get('key'))
        return stores

    @api.model
    def get_user_store_progress_matrix(self):
        """Tính toán ma trận tiến độ CTKM × Cửa hàng cho người dùng hiện tại."""
        from odoo.addons.ctkm_core.models.ctkm_task_store_dispatch import _CTKM_DISPATCH_CHAIN

        user = self.env.user
        stores = self._ctkm_get_user_managed_stores(user)

        # Tìm các công việc CTKM liên quan đến user (hoặc tất cả nếu là manager/admin)
        Task = self.sudo()
        user_tasks = Task.search([
            ('program_id.state', 'not in', ['draft', 'cancel']),
            '|', '|', '|',
                ('user_ids', 'in', [user.id]),
                ('user_ids.employee_ids.parent_id.user_id', 'in', [user.id]),
                ('verifier_ids.user_id', 'in', [user.id]),
                ('store_verifier_ids.verifier_user_id', 'in', [user.id]),
        ])
        programs = user_tasks.mapped('program_id').filtered(
            lambda p: p.state not in ('draft', 'cancel')
        )

        if not programs:
            if user.has_group('ctkm_core.group_ctkm_user') or user.has_group('ctkm_core.group_ctkm_manager'):
                programs = self.env['ctkm.program'].sudo().search([
                    ('state', 'not in', ['draft', 'cancel'])
                ], order='date_begin desc, id desc', limit=20)
            elif stores:
                store_keys_set = {s['key'] for s in stores}
                candidates = self.env['ctkm.program'].sudo().search([
                    ('state', 'not in', ['draft', 'cancel'])
                ], order='date_begin desc, id desc', limit=20)
                matched_programs = self.env['ctkm.program']
                for cand in candidates:
                    step4_map = cand._ctkm_step4_store_qty_map()
                    if step4_map and any(self._ctkm_store_in_allowed(store_keys_set, k) for k in step4_map):
                        matched_programs |= cand
                    elif cand.task_ids.filtered(lambda t: any(
                        self._ctkm_store_in_allowed(store_keys_set, d.store_key, d.store_label)
                        for d in t.store_dispatch_ids
                    )):
                        matched_programs |= cand
                programs = matched_programs

        # Nếu user chưa có store cụ thể (ví dụ tài khoản admin/manager chung),
        # lấy các cửa hàng xuất hiện trong các chương trình này
        if not stores and programs:
            seen_keys = set()

            def _add_store(key, code, name):
                if not key:
                    return
                canon = self._ctkm_store_canonical_key(key, code, name) or key
                if canon in seen_keys:
                    if name and name != canon:
                        for s in stores:
                            if s['key'] == canon and (s['name'] == canon or len(name) > len(s['name'])):
                                s['name'] = name
                    return
                seen_keys.add(canon)
                stores.append({'key': canon, 'code': code or canon, 'name': name or code or canon})

            for p in programs:
                step4_map = p.sudo()._ctkm_step4_store_qty_map()
                if step4_map:
                    for k in step4_map:
                        canon = self._ctkm_store_canonical_key(k) or k
                        lbl = p.task_ids[:1]._ctkm_store_label_for_key(canon) if p.task_ids else canon
                        _add_store(canon, canon, lbl or canon)
                for t in p.task_ids:
                    for d in t.sudo().store_dispatch_ids:
                        k = d.store_key
                        _add_store(k, k, d.store_label or k)
                    for sv in t.sudo().store_verifier_ids:
                        k = sv.store_key
                        _add_store(k, k, sv.store_label or k)
                    if t.is_tem_print_task:
                        for pl in t.print_store_ids:
                            k = pl.store_key or _ctkm_normalize_store_key(pl.store)
                            _add_store(k, pl.store_code, pl.store_id.name if pl.store_id else pl.store)
            if not stores and 'hr.store' in self.env:
                for s in self.env['hr.store'].sudo().search([], order='code, name', limit=15):
                    k = _ctkm_normalize_store_key(s.code or s.name)
                    _add_store(k, s.code, s.name)
            stores.sort(key=lambda s: s.get('name') or s.get('code') or s.get('key'))

        if not programs or not stores:
            return {
                'has_data': False,
                'stores': stores,
                'programs': [],
            }

        # Sắp xếp các chương trình theo ngày bắt đầu giảm dần
        programs = programs.sorted(lambda p: (p.date_begin or fields.Date.today(), p.id), reverse=True)

        program_rows = []
        for program in programs:
            tasks = program.task_ids.sorted(lambda t: (t.checklist_line_id.sequence or 999, t.id))
            stores_map = program.sudo()._ctkm_step4_store_qty_map()
            cells = {}

            # Cache task của chuỗi bước 10-16
            chain_tasks = {}
            for rank, flag in _CTKM_DISPATCH_CHAIN:
                t_match = tasks.filtered(lambda t: t[flag])
                if t_match:
                    chain_tasks[rank] = t_match[:1]

            for st in stores:
                st_key = st['key']
                st_aliases = self._ctkm_store_key_aliases(st_key, st.get('code'), st.get('name'))

                # 1. Kiểm tra xem cửa hàng này có thuộc chương trình không
                is_store_in_prog = False
                if stores_map:
                    bucket = program._ctkm_match_store_bucket(stores_map, st_key)
                    if bucket or (st_aliases & set(stores_map.keys())):
                        is_store_in_prog = True
                if not is_store_in_prog:
                    for t in tasks:
                        if t.store_dispatch_ids.filtered(lambda d: self._ctkm_store_in_allowed(st_aliases, d.store_key, d.store_label)):
                            is_store_in_prog = True
                            break
                        if t.store_verifier_ids.filtered(lambda sv: self._ctkm_store_in_allowed(st_aliases, sv.store_key, sv.store_label)):
                            is_store_in_prog = True
                            break
                        if t.is_tem_print_task and t.print_store_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break
                        if t.is_tem_handover_task and t.handover_store_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break
                        if (t.is_tem_receive_task or t.is_tem_replace_task) and t.tem_tag_replace_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break
                        if (t.is_tem_photo_task or t.is_tem_check_task) and t.tem_photo_check_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break
                        if t.is_tem_price_task and t.price_store_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break
                        if t.is_tem_postcheck_task and t.postcheck_store_ids.filtered(lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)):
                            is_store_in_prog = True
                            break

                if not is_store_in_prog:
                    cells[st_key] = None
                    continue

                # 2. Kiểm tra nếu cả chương trình đang ở các bước chuẩn bị (sequence < 9)
                has_started_dispatch = any(
                    t._ctkm_is_store_dispatch_step() and t.store_dispatch_ids
                    for t in tasks
                )
                early_tasks = [t for t in tasks if t.checklist_line_id.sequence and t.checklist_line_id.sequence < 9 and t.state != 'done']
                if early_tasks and not has_started_dispatch:
                    et = early_tasks[0]
                    s_label = 'Đang làm' if et.state == 'progress' else ('Chờ xác nhận' if et.state == 'waiting_confirm' else 'Chưa bắt đầu')
                    b_class = 'warning' if et.state == 'progress' else ('info' if et.state == 'waiting_confirm' else 'secondary')
                    cells[st_key] = {
                        'task_id': et.id,
                        'stage_name': et.checklist_line_id.name or et.name,
                        'state': et.state,
                        'state_label': s_label,
                        'badge_class': b_class,
                    }
                    continue

                # 3. Bước 9 (In tem tag):
                print_task = tasks.filtered('is_tem_print_task')[:1]
                if print_task and print_task.state != 'done':
                    plines = print_task.print_store_ids.filtered(
                        lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)
                    )
                    if plines and not all(l.done for l in plines):
                        s_label = 'Đang làm' if print_task.state == 'progress' else ('Chờ xác nhận' if print_task.state == 'waiting_confirm' else 'Chưa bắt đầu')
                        b_class = 'warning' if print_task.state == 'progress' else ('info' if print_task.state == 'waiting_confirm' else 'secondary')
                        cells[st_key] = {
                            'task_id': print_task.id,
                            'stage_name': print_task.checklist_line_id.name or print_task.name,
                            'state': print_task.state,
                            'state_label': s_label,
                            'badge_class': b_class,
                        }
                        continue

                # 4. Chuỗi phân nhánh từng cửa hàng (Ranks 10..15):
                cell_found = False
                for rank in range(10, 16):
                    step_task = chain_tasks.get(rank)
                    if not step_task:
                        continue

                    if step_task.state == 'done':
                        # Toàn bộ task đã xong -> cửa hàng này đã vượt qua bước này, xét tiếp bước sau
                        continue

                    # Kiểm tra store_dispatch_ids của cửa hàng trên bước này
                    dispatches = step_task.store_dispatch_ids.filtered(
                        lambda d: self._ctkm_store_in_allowed(st_aliases, d.store_key, d.store_label)
                    )
                    if dispatches:
                        if any(d.state == 'sent' for d in dispatches):
                            # Đã xác nhận gửi xuống bước sau -> xét bước kế tiếp
                            continue
                        if any(d.state == 'pending' for d in dispatches):
                            # Đang chờ xác nhận gửi dữ liệu ở bước này!
                            cells[st_key] = {
                                'task_id': step_task.id,
                                'stage_name': step_task.checklist_line_id.name or step_task.name,
                                'state': 'waiting_confirm',
                                'state_label': 'Chờ xác nhận',
                                'badge_class': 'info',
                            }
                            cell_found = True
                            break

                    # Chưa có dispatch sent/pending -> Cửa hàng hiện đang dừng ở bước này!
                    sv = step_task.store_verifier_ids.filtered(
                        lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store_label)
                    )
                    if sv and any(not l.verified and (l.assignee_completed or step_task.state == 'waiting_confirm') for l in sv):
                        s_label = 'Chờ xác nhận'
                        b_class = 'info'
                    elif step_task.state == 'waiting_confirm':
                        s_label = 'Chờ xác nhận'
                        b_class = 'info'
                    elif step_task.state == 'progress':
                        s_label = 'Đang làm'
                        b_class = 'warning'
                    else:
                        s_label = 'Chưa bắt đầu'
                        b_class = 'secondary'

                    cells[st_key] = {
                        'task_id': step_task.id,
                        'stage_name': step_task.checklist_line_id.name or step_task.name,
                        'state': step_task.state,
                        'state_label': s_label,
                        'badge_class': b_class,
                    }
                    cell_found = True
                    break

                if cell_found:
                    continue

                # 5. Bước 16 (Hậu kiểm CTKM) - Bước cuối cùng của chuỗi
                postcheck_task = chain_tasks.get(16)
                if postcheck_task:
                    plines = postcheck_task.postcheck_store_ids.filtered(
                        lambda l: self._ctkm_store_in_allowed(st_aliases, l.store_key, l.store)
                    )
                    is_postcheck_done = postcheck_task.state == 'done' or bool(plines and all(l.replaced or l.tag_replaced for l in plines))
                    if is_postcheck_done:
                        cells[st_key] = {
                            'task_id': postcheck_task.id,
                            'stage_name': 'Hoàn thành',
                            'state': 'done',
                            'state_label': 'Hoàn thành',
                            'badge_class': 'success',
                        }
                    else:
                        s_label = 'Chờ xác nhận' if postcheck_task.state == 'waiting_confirm' else ('Đang làm' if postcheck_task.state == 'progress' else 'Chưa bắt đầu')
                        b_class = 'info' if postcheck_task.state == 'waiting_confirm' else ('warning' if postcheck_task.state == 'progress' else 'secondary')
                        cells[st_key] = {
                            'task_id': postcheck_task.id,
                            'stage_name': postcheck_task.checklist_line_id.name or postcheck_task.name,
                            'state': postcheck_task.state,
                            'state_label': s_label,
                            'badge_class': b_class,
                        }
                else:
                    last_t = tasks[-1] if tasks else False
                    cells[st_key] = {
                        'task_id': last_t.id if last_t else False,
                        'stage_name': 'Hoàn thành',
                        'state': 'done',
                        'state_label': 'Hoàn thành',
                        'badge_class': 'success',
                    }

            program_rows.append({
                'id': program.id,
                'name': program.name,
                'code': program.notify_code or '',
                'cells': cells,
            })

        return {
            'has_data': True,
            'stores': stores,
            'programs': program_rows,
        }


class CtkmTaskUserCompletion(models.Model):
    _name = 'ctkm.task.user.completion'
    _description = 'Hoàn thành của từng người nhận việc'
    _rec_name = 'user_id'
    _order = 'done_date desc, id'

    task_id = fields.Many2one(
        'ctkm.task', string='Công việc',
        required=True, ondelete='cascade', index=True,
    )
    user_id = fields.Many2one(
        'res.users', string='Người nhận việc',
        required=True, index=True,
    )
    done = fields.Boolean(string='Đã hoàn thành', default=False)
    done_date = fields.Date(string='Ngày hoàn thành')

    _sql_constraints = [
        ('task_user_uniq', 'unique(task_id, user_id)',
         'Mỗi người nhận việc chỉ có một bản ghi hoàn thành.'),
    ]
