# -*- coding: utf-8 -*-

import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .ctkm_inventory_tem_tag import _normalize_store_code

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# Định dạng số "Giá KM" (accounting, không dấu phẩy nghìn chuẩn).
ACCOUNTING_NUMBER_FORMAT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'
# Định dạng số lượng (SL bàn giao).
QUANTITY_NUMBER_FORMAT = '#,##0'

# Bố cục biên bản (theo mẫu Bien_Ban_In_Va_Ban_Giao_Tag.xlsx): 5 cột A-E.
COL_STT = 1
COL_MA_VAT_TU = 2
COL_GIA_KM = 3
COL_SL_BAN_GIAO = 4
COL_GHI_CHU = 5
HEADER_ROW = 8
DATA_START_ROW = 9

# Ký tự không hợp lệ trong tên sheet Excel.
_INVALID_SHEET_CHARS = (':', '\\', '/', '?', '*', '[', ']')


class CtkmTask(models.Model):
    _inherit = 'ctkm.task'

    store_ids = fields.Many2many(
        'hr.store',
        string='Cửa hàng',
        help='Chọn một hoặc nhiều cửa hàng để xuất biên bản thay tem (bước 6).',
    )
    bb_export_store_domain_ids = fields.Many2many(
        'hr.store',
        string='Cửa hàng được xuất BB',
        compute='_compute_bb_export_store_domain_ids',
        help='Các cửa hàng vừa có dữ liệu Tem/Tag import bước 4, vừa thuộc Cửa hàng quản lí.',
    )

    @api.depends('program_id', 'user_ids')
    def _compute_bb_export_store_domain_ids(self):
        for task in self:
            task.bb_export_store_domain_ids = task._bb_export_allowed_stores()

    @api.onchange('bb_export_store_domain_ids')
    def _onchange_bb_export_store_domain_ids(self):
        for task in self:
            allowed = task.bb_export_store_domain_ids
            if allowed:
                task.store_ids &= allowed
            else:
                task.store_ids = False

    def action_export_bb_file(self):
        self.ensure_one()
        if not (self.is_tem_bb_replace_task or self.is_tem_replace_task):
            raise UserError(_(
                'Chỉ công việc bước "Lập BB thay tem" (hoặc "Thay tem Tag") mới '
                'được xuất biên bản thay tem.'
            ))
        if not self.store_ids:
            raise UserError(_('Vui lòng chọn ít nhất một Cửa hàng trước khi xuất file.'))
        self._check_bb_export_store_ids()
        if Workbook is None:
            raise UserError(_('Thiếu thư viện openpyxl để xuất file Excel.'))
        data = self._build_bb_xlsx()
        date_str = fields.Date.context_today(self).strftime('%d_%m_%Y')
        filename = 'BB_thay_tem_%s_%s.xlsx' % (
            self.program_id.notify_code or self.program_id.id or 'CTKM',
            date_str,
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'datas': base64.b64encode(data),
            'res_model': 'ctkm.task',
            'res_id': self.id,
            'type': 'binary',
            'public': True,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def _bb_export_allowed_store_keys(self):
        self.ensure_one()
        if not self.program_id:
            return set()
        Inventory = self.env['ctkm.inventory.tem.tag'].sudo()
        groups = Inventory.read_group(
            [('program_id', '=', self.program_id.id)],
            ['store_key'], ['store_key'],
        )
        inventory_keys = {
            row['store_key']
            for row in groups
            if row.get('store_key')
        }
        managed_keys = set()
        if hasattr(self, '_ctkm_tem_tag_managed_store_keys'):
            managed_keys = set(self._ctkm_tem_tag_managed_store_keys())
        return inventory_keys & managed_keys

    def _bb_export_allowed_stores(self):
        self.ensure_one()
        allowed_keys = self._bb_export_allowed_store_keys()
        if not allowed_keys or 'hr.store' not in self.env:
            return self.env['hr.store']
        stores = self.env['hr.store'].sudo().search([])
        return stores.filtered(
            lambda store: bool({
                _normalize_store_code(store.code),
                _normalize_store_code(store.name),
            } & allowed_keys)
        )

    def _check_bb_export_store_ids(self):
        self.ensure_one()
        allowed_keys = self._bb_export_allowed_store_keys()
        invalid = self.env['hr.store']
        for store in self.store_ids.sudo():
            store_keys = {
                _normalize_store_code(store.code),
                _normalize_store_code(store.name),
            }
            if not (store_keys & allowed_keys):
                invalid |= store
        if invalid:
            raise UserError(_(
                'Bạn chỉ được xuất BB cho cửa hàng vừa có dữ liệu Tem/Tag bước 4 '
                'vừa thuộc "Cửa hàng quản lí": %s'
            ) % ', '.join(invalid.mapped('display_name')))

    @api.model
    def _bb_sheet_title(self, store_code):
        """Tên sheet hợp lệ (tối đa 31 ký tự, không chứa ký tự cấm)."""
        title = (store_code or 'CH').strip()
        for ch in _INVALID_SHEET_CHARS:
            title = title.replace(ch, '_')
        title = title[:31] or 'CH'
        # Đảm bảo tên không được đặt trong dấu ngoặc đơn/đơn.
        return title

    def _build_bb_xlsx(self):
        """Xuất biên bản thay tem theo mẫu Bien_Ban_In_Va_Ban_Giao_Tag.xlsx.

        Mỗi cửa hàng được chọn xuất thành một sheet riêng. Bố cục mỗi sheet:
            Dòng 1: "Cộng Hòa Xã Hội Chủ Nghĩa Việt Nam" (merge A1:E1)
            Dòng 2: "Độc lập – Tự do – Hạnh phúc" (merge A2:E2)
            Dòng 3: "BIÊN BẢN IN VÀ BÀN GIAO TAG" (tiêu đề, merge A3:E3)
            Dòng 4: "SỐ : .............." (merge A4:E4)
            Dòng 5: "Hôm nay, Ngày ... Tháng ... Năm ..."
            Dòng 6: "TẠI CH:<mã cửa hàng>"
            Dòng 7: "Nội dung công việc:BIÊN BẢN THAY TAG THEO TB <mã> NGÀY <ngày>" (merge A7:E7)
            Dòng 8: Tiêu đề cột: STT | MÃ VẬT TƯ | GIÁ KM | SL BÀN GIAO | GHI CHÚ
            Dòng 9+: Chi tiết từng Mã vật tư của cửa hàng đó
            Cột GHI CHÚ ghi nội dung CTKM (cột "CTKM" của file import bước 4).
        """
        self.ensure_one()
        program = self.program_id
        inventory_date = fields.Date.context_today(self)
        date_str = inventory_date.strftime('%d/%m/%Y')
        date_vn = 'Ngày %s Tháng %s Năm %s' % (
            inventory_date.day, inventory_date.month, inventory_date.year,
        )
        notify_code = (program.notify_code or '').strip()

        # --- Các cửa hàng: lấy từ selection, hoặc toàn bộ cửa hàng có trong
        #     kho Tem/Tag (import bước 4) khi không chọn cửa hàng nào. ---
        selected = self.store_ids.sudo()
        store_list = []  # (store_key, mã hiển thị)
        seen = set()
        if selected:
            for store in selected:
                key = _normalize_store_code(store.code or store.name)
                if key and key not in seen:
                    seen.add(key)
                    store_list.append((key, store.code or store.name or key))
        else:
            Inventory = self.env['ctkm.inventory.tem.tag'].sudo()
            groups = Inventory.read_group(
                [('program_id', '=', program.id)],
                ['store_key'], ['store_key'],
            )
            for row in groups:
                key = row['store_key']
                if key and key not in seen:
                    seen.add(key)
                    store_list.append((key, key))
        if not store_list:
            raise UserError(_(
                'Không có dữ liệu kho Tem/Tag (import bước 4) cho chương trình này.'
            ))

        # --- Đọc 1 lần kho Tem/Tag của chương trình ---
        Inventory = self.env['ctkm.inventory.tem.tag'].sudo()
        inv_records = Inventory.search([('program_id', '=', program.id)])

        # --- Vẽ workbook: mỗi cửa hàng 1 sheet ---
        wb = Workbook()
        used_titles = set()
        first = True
        for key, store_code in store_list:
            materials = {}
            work_contents = []
            for rec in inv_records:
                if rec.store_key != key:
                    continue
                code = rec.material_code
                if not code:
                    continue
                mat = materials.setdefault(code, {
                    'promo': rec.promo_price or 0.0,
                    'qty': 0.0,
                    'notes': [],
                })
                mat['qty'] += rec.quantity or 0.0
                note = (rec.ctkm_name or '').strip()
                if note and note not in mat['notes']:
                    mat['notes'].append(note)
                work_content = (rec.bb_work_content or '').strip()
                if work_content and work_content not in work_contents:
                    work_contents.append(work_content)
            material_codes = sorted(materials.keys())

            ws = wb.active if first else wb.create_sheet()
            first = False

            # Tên sheet: mã cửa hàng, đảm bảo không trùng.
            base_title = self._bb_sheet_title(store_code)
            sheet_title = base_title
            suffix = 1
            while sheet_title in used_titles:
                suffix += 1
                sheet_title = '%s_%s' % (base_title[:31 - len('_%s' % suffix)], suffix)
            used_titles.add(sheet_title)
            ws.title = sheet_title

            self._build_bb_sheet(
                ws, store_code, material_codes, materials,
                date_vn, date_str, notify_code, work_contents[:1],
            )

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _build_bb_sheet(self, ws, store_code, material_codes, materials,
                       date_vn, date_str, notify_code, work_contents):
        """Vẽ nội dung một sheet biên bản cho một cửa hàng."""
        thin = Side(style='thin', color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        bold = Font(bold=True, size=11)
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=11)
        cell_font = Font(size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Dòng 1-2: tiêu ngữ
        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = 'Cộng Hòa Xã Hội Chủ Nghĩa Việt Nam'
        c.font = bold
        c.alignment = center
        ws.merge_cells('A2:E2')
        c = ws['A2']
        c.value = 'Độc lập – Tự do – Hạnh phúc'
        c.font = bold
        c.alignment = center

        # Dòng 3: tiêu đề
        ws.merge_cells('A3:E3')
        c = ws['A3']
        c.value = 'BIÊN BẢN IN VÀ BÀN GIAO TAG'
        c.font = title_font
        c.alignment = center
        ws.row_dimensions[3].height = 24

        # Dòng 4: số
        ws.merge_cells('A4:E4')
        c = ws['A4']
        c.value = 'SỐ : ..............'
        c.font = bold
        c.alignment = center

        # Dòng 5: ngày
        ws['A5'] = date_vn
        ws['A5'].font = cell_font

        # Dòng 6: tại cửa hàng
        ws['A6'] = 'TẠI CH:%s' % store_code
        ws['A6'].font = cell_font

        # Dòng 7: nội dung công việc
        ws.merge_cells('A7:E7')
        c = ws['A7']
        work_content = work_contents[0] if work_contents else ''
        if not work_content:
            work_content = 'BIÊN BẢN THAY TAG THEO TB %s NGÀY %s' % (
                notify_code, date_str,
            )
        c.value = 'Nội dung công việc: %s' % work_content
        c.font = cell_font
        c.alignment = left_top
        ws.row_dimensions[7].height = 28

        # Dòng 8: tiêu đề cột
        headers = ['STT', 'MÃ VẬT TƯ', 'GIÁ KM', 'SL BÀN GIAO', 'GHI CHÚ']
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
            cell.font = header_font
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[HEADER_ROW].height = 28

        # Dòng 9+: chi tiết
        r = DATA_START_ROW
        for idx, code in enumerate(material_codes, start=1):
            mat = materials[code]
            ws.cell(row=r, column=COL_STT, value=idx)
            ws.cell(row=r, column=COL_MA_VAT_TU, value=code)
            gc = ws.cell(row=r, column=COL_GIA_KM, value=mat['promo'])
            gc.number_format = ACCOUNTING_NUMBER_FORMAT
            qc = ws.cell(row=r, column=COL_SL_BAN_GIAO, value=mat['qty'])
            qc.number_format = QUANTITY_NUMBER_FORMAT
            ws.cell(row=r, column=COL_GHI_CHU, value='\n'.join(mat['notes']))
            for col_idx in range(1, COL_GHI_CHU + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.font = cell_font
                cell.border = border
                if col_idx in (COL_STT, COL_GIA_KM, COL_SL_BAN_GIAO):
                    cell.alignment = center
                else:
                    cell.alignment = left_top
            ws.row_dimensions[r].height = 24
            r += 1

        # Độ rộng cột (khớp mẫu: A=8, B=28, C=15, D=14, E=85)
        widths = {1: 8, 2: 28, 3: 15, 4: 14, 5: 85}
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
